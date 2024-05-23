import anvil.files
from anvil.files import data_files
# Copyright(C) Val-Cloud Ltd 2023. All rights reserved
import anvil.secrets
import anvil.email
import anvil.users
import anvil.tables as tables
import anvil.tables.query as q
from anvil.tables import app_tables
import anvil.server
from openpyxl import load_workbook
import json
import pyodbc
import io
import anvil.media
import Connections
import pyodbc
import sqlalchemy as salch
import urllib.parse

#from ServerModule1 import initialise_database_connection 

from datetime import datetime, timedelta
from datetime import date
from dateutil.relativedelta import *
import calendar
from typing import Optional
import numpy as np
import decimal
from rich import box
from rich.console import Console
from rich.table import Table
import Project_types as pt
import Benchmarks as bm

import pandas as pd
from pandas.api.types import is_numeric_dtype
from io import StringIO
import pandas_schema as ps
from pandas_schema import Column, Schema
from pandas_schema.validation import LeadingWhitespaceValidation, TrailingWhitespaceValidation, CanConvertValidation, MatchesPatternValidation, InRangeValidation, InListValidation, CustomElementValidation
import sys, traceback
import math

from . import Solar_factors as sf

import cryptography
from cryptography.fernet import Fernet

# This is a server package. It runs on the Anvil server,
# rather than in the user's browser.
#
# To allow anvil.server.call() to call functions here, we mark
# them with @anvil.server.callable.

def get_dec_letter(dec_number):
  # Translates a provided dec_number into a dec_letter
  dec_letter = "NO DEC"
  
  if dec_number >= 1 and dec_number < 26: 
    dec_letter = "A"
  if dec_number >= 26 and dec_number < 51: 
    dec_letter = "B"    
  if dec_number >= 51 and dec_number < 76: 
    dec_letter = "C"
  if dec_number >= 76 and dec_number < 101: 
    dec_letter = "D"
  if dec_number >= 101 and dec_number < 126: 
    dec_letter = "E"
  if dec_number >= 126 and dec_number < 151: 
    dec_letter = "F"
  if dec_number >= 151: 
    dec_letter = "G" 
    
  return dec_letter
  
def round_half_up(n, decimals=0):
# This function is provided to enable rounding of 0.5 to be compatible with Excel (i.e. rounds 0.5 up). The Python
# built-in function, round(), rounds 0.5 down but 0.51 up.
  multiplier = 10**decimals
  return math.floor(n * multiplier + 0.5) / multiplier

def encrypt(message: bytes, key: bytes) -> bytes:
    return Fernet(key).encrypt(message)

def decrypt(token: bytes, key: bytes) -> bytes:
    return Fernet(key).decrypt(token)

def create_upload_auth_key(partner, client, entity):
  if (partner == None) or (client == None) or (entity == None):
    return 'Error - 1 or more null inputs'
  else:
    stored_key = anvil.secrets.get_secret('upload_auth_key')
    key = bytes(stored_key, 'utf-8')
    message = f"{partner}/{client}/{entity}"
    bauth_key = encrypt(message.encode(), key)
    auth_key  = bauth_key.decode("utf-8")
  return auth_key
  
def decode_upload_auth_key(auth_key):
  stored_key = anvil.secrets.get_secret('upload_auth_key')
  key = bytes(stored_key, 'utf-8') 
  message = decrypt(auth_key, key).decode()
  parts_list = message.split('/')
  return parts_list
  
def write_auth_sheets(workbook, worksheetau, worksheetky, partner, client, entity):
  ret_mess = {'ef':0, 'em': ''}
  print('In write_auth_sheets ======xxxxxxx')
  print(f"Partner - {partner}")
  print(f"Client - {client}")
  print(f"Entity - {entity}")
  try:
    auth_key = create_upload_auth_key(partner, client, entity)
    if 'Error' in auth_key:
      ret_mess['ef'] = 2
      ret_mess['em'] = 'Error creating auth key.'
      return ret_mess

 # Write the Auth worksheet, visible to user, containing Partner, Client and Entity details
    
    cell_format_LABEL    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'black', 'bg_color': 'white', 'text_wrap':'false', 'bold': True, 'valign':'vcenter', 'align':'left'}) 
    worksheetau.write('A1','This workbook has been generated for exclusive use by: -', cell_format_LABEL)
    worksheetau.write('A2','OPF Partner:', cell_format_LABEL)
    worksheetau.write('A4','Client:', cell_format_LABEL)
    worksheetau.write('A6','Entity', cell_format_LABEL)

    cell_format_NORM    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'black', 'bg_color': 'white', 'text_wrap':'false', 'bold': False, 'valign':'vcenter', 'align':'left'}) 
    worksheetau.write('A3',f"{partner}", cell_format_NORM)
    worksheetau.write('A5',f"{client}", cell_format_NORM)
    worksheetau.write('A7',f"{entity}", cell_format_NORM)
  #  worksheetau.write('A9',f"{auth_key}", cell_format_NORM) 

    worksheetau.set_column('A:A',75) 

 # Write the Key worksheet, hide from user, containing the encrypted key made up of Partner, Client and Entity. 
    
    worksheetky.write('A1','Key:', cell_format_LABEL)
    worksheetky.write('A2',f"{auth_key}", cell_format_NORM)
    
    worksheetky.set_column('A:A',150)
    worksheetky.hide()

    wspwd = anvil.secrets.get_secret('protect_workbook')
    # Protect worksheets

    optionsau = {
    'format_cells':          False,
    'format_columns':        False,
    'format_rows':           False,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           False,
    'select_locked_cells':   False,
    'sort':                  False,
    'autofilter':            False,
    'pivot_tables':          False,
    'select_unlocked_cells': False,
    }        
    worksheetky.protect(wspwd,optionsau)
    worksheetau.protect(wspwd,optionsau)
    
    return ret_mess 
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = f"****Error - exception occured in write_auth_sheets. \n {msg}"
    return ret_mess
  
def validate_estate_upload(conn, entity, entity_number, df):
  print('At top validate_estate_upload')
  # Validates estate upload dataframe (df) for entity with entity_number read from upload file. Conn is the connection object to the decarb database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()
    #print('df_original at start validate')
    #print(df_original.to_string())

    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce warnings
    
    ids                 = df_in["uprn"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="uprn")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["uprn","excel_row_num", "building_name", "address","postcode"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"-----WARNING - {num_dup_uprn} occurrences of duplicate UPRNs have been found - review advised\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"
      validation['nvw']   = validation['nvw'] + num_dup_uprn
    
#  Validate action codes 

    df_actions            = df_in[['action','excel_row_num']].copy()
    df_actions            = df_actions.set_index('excel_row_num')
    schema = Schema([
      Column('action', [InListValidation(['DELETE', 'UPDATE'])]),])
    
    numactionerrs         = 0
    errors                = schema.validate(df_actions)
    errors_index_rows     = [e.row for e in errors]

    numactionerrs         = len(errors)
    if numactionerrs > 0:
      vm                = vm + f"{numactionerrs} errors have been found in ACTION codes:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"    

#    Remove DELETE records from further validation to stop spurious messages - DELETE only needs action code and uprn

    df_in.drop(df_in[df_in['action'] == "DELETE"].index, inplace = True)

#   Validate numeric columns don't contain strings. If they do then report and return  
    numeric_columns_list =      ['entering_estate_year','entering_estate_month','exiting_estate_year','exiting_estate_month','gia_m2','roof_space_m2','data_year','annual_elec_kwh','annual_gas_kwh',\
      'annual_oil_kwh','annual_lpg_kwh', 'dec_score','epc','elec_cost_per_kwh','gas_cost_per_kwh','oil_cost_per_kwh','lpg_cost_per_kwh','annual_generation_kwh','decarbonised_heat_annual_output_kwh',\
      'number_of_ev_charge_sockets', 'charging_capacity_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh' ]
    
    ret               = log_str_types_in_numeric_cols(df_in, numeric_columns_list)

    ef                = ret['ef']
    em                = ret['em']
    nerrs             = ret['nerrs']
    out_log           = ret['out_log']

    if ef == 2:
      validation['ef']                  = 2
      validation['em']                 = f"**Error while checking for presence of strings in numeric fields - \n {em}"
      validation['validated_df']        = ''
      validation['validation_messages'] = ''
      return validation
    
    if nerrs > 0:
      validation['ef']                  = 0
      validation['em']                 = ''
      validation['validated_df']        = ''
      validation['validation_messages'] = out_log
      validation['nve']                 = nerrs
      return validation      
    
#    Read building types from table cibse_benchmarks
    
    with conn.cursor() as cursor:
      sql1             = f"SELECT building_type FROM {bm.benchmark_table_name};"
      cursor.execute(sql1)
      t_building_types = cursor.fetchall()
      keys             = ("building_type","dummy_key")
      building_types   = [dict(zip(keys, values)) for values in t_building_types]

#    Validate building_type  

    bt_mess                = ''
    nbterr                 = 0
    num_miss_building_type = 0
    
    for index, row in df_in.iterrows():
      excel_row_num        = row['excel_row_num']
      bt                   = row['building_type']

      if bt == None or bt == '':
        num_miss_building_type = num_miss_building_type + 1
        bt_mess          = bt_mess + f"++++Warning - Building type on Excel row {excel_row_num} is missing \n"        
      elif not any(d['building_type'] == bt for d in building_types):
        nbterr           = nbterr + 1
        bt_mess          = bt_mess + f"****ERROR - Building type - {bt} - on Excel row {excel_row_num} is not valid \n"
        
    if nbterr > 0 or num_miss_building_type > 0:
      vm                 = vm + f"{nbterr} errors and {num_miss_building_type} missing building type warnings have been found in BUILDING TYPES:-\n"
      vm                 = vm + bt_mess
      validation['nve']  = validation['nve'] + nbterr

    thisyear = datetime.now ().year 

    # Validate remaining fields in dataframe
  
    df_s                 = df_in[['excel_row_num','building_name','address','postcode','under_control','remain_in_portfolio','entering_estate_year','entering_estate_month','exiting_estate_year','exiting_estate_month'\
                                ,'listed','construction_year','gia_m2','roof_space_m2','data_year','annual_elec_kwh','annual_gas_kwh','annual_oil_kwh','annual_lpg_kwh','source_of_heating','source_of_dhw','dec_score'\
                                ,'epc','elec_cost_per_kwh','gas_cost_per_kwh','oil_cost_per_kwh','lpg_cost_per_kwh','onsite_generation_asset','annual_generation_kwh','decarbonised_heat_asset','decarbonised_heat_annual_output_kwh'\
                                ,'car_park_available','number_of_ev_charge_sockets','charging_capacity_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh']].copy()

    # Reset index to excel_row_number
    
    df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    schema = Schema([
      Column('building_name',[CustomElementValidation(lambda d: len(d) > 0, "Building_name is missing ")]),
      Column('address',[CustomElementValidation(lambda d: len(d) > 0, "Address is missing ")]),
      Column('postcode',[CustomElementValidation(lambda d: len(d) > 0, "Postcode is missing ")]),
      Column('under_control',[InListValidation(['YES', 'NO',''])]),
      Column('remain_in_portfolio',[InListValidation(['YES', 'NO',''])]),
      Column('entering_estate_year',[InRangeValidation(0, 2056)]),
      Column('entering_estate_month',[InRangeValidation(0, 12)]),
      Column('exiting_estate_year',[InRangeValidation(0, 2056)]),
      Column('exiting_estate_month',[InRangeValidation(0, 12)]),      
      Column('listed',[InListValidation(['YES', 'NO',''])]),
      Column('construction_year',[InListValidation(["2020","2015","2010","2005","2000","1990S","1980S","1970S","1960S","1950S","PRE 1950","PRE 1900S",""])]),
      Column('gia_m2',   [CustomElementValidation(lambda d: d >= 0, "gia_m2 must be >= to zero")]),
      Column('roof_space_m2', [CustomElementValidation(lambda d: d >= 0, "roof_space_m2 must be >= to zero")]),
      Column('data_year',   [InRangeValidation(2017, 2030)]),
      Column('annual_elec_kwh', [CustomElementValidation(lambda d: d >= 0, "annual_elec_kwh must be a >= to zero")]),      
      Column('annual_gas_kwh', [CustomElementValidation(lambda d: d >= 0, "annual_gas_kwh must be >= to zero")]), 
      Column('annual_oil_kwh', [CustomElementValidation(lambda d: d >= 0, "annual_oil_kwh must be >= to zero")]), 
      Column('annual_lpg_kwh', [CustomElementValidation(lambda d: d >= 0, "annual_lpg_kwh must be >= to zero")]),
      Column('source_of_heating',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('source_of_dhw',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('dec_score', [CustomElementValidation(lambda d: d >= 0, "dec_score must be >= to zero")]),
      Column('epc', [CustomElementValidation(lambda d: d >= 0, "epc must be >= to zero")]),
      Column('elec_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "elec_cost_per_kwh must be >= to zero")]),
      Column('gas_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "gas_cost_per_kwh must be >= to zero")]),
      Column('oil_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "oil_cost_per_kwh must be >= to zero")]),
      Column('lpg_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "lpg_cost_per_kwh must be >= to zero")]),
      Column('onsite_generation_asset',[InListValidation(['YES', 'NO',''])]),
      Column('annual_generation_kwh', [CustomElementValidation(lambda d: d >= 0, "annual_generation_kwh must be >= to zero")]),
      Column('decarbonised_heat_asset',[InListValidation(['YES', 'NO',''])]),
      Column('decarbonised_heat_annual_output_kwh', [CustomElementValidation(lambda d: d >= 0, "decarbonised_heat_annual_output_kwh must be >= to zero")]),
      Column('car_park_available',[InListValidation(['YES', 'NO',''])]),
      Column('number_of_ev_charge_sockets', [CustomElementValidation(lambda d: vcl_check_int(d), "number_of_ev_charge_sockets must be an whole number")]),      
      Column('charging_capacity_kwh', [CustomElementValidation(lambda d: d >= 0, "charging_capacity_kwh must be >= to zero")]),
      Column('exist_solar_pv_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_pv_annual_kwh must be >= to zero")]),
      Column('exist_solar_thermal_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_thermal_annual_kwh must be >= to zero")]),
    ])

    errors            = schema.validate(df_s)
    errors_index_rows = [e.row for e in errors]
    nscerrs           = len(errors)
    
    if nscerrs > 0:
      vm                = vm + f"{nscerrs} errors have been found in other fields:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"
#{'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}    
    nvw                 = num_dup_uprn + num_miss_building_type
    nve                 = nuprnerrs + numactionerrs + nbterr + nscerrs
    #print('df_original at end validation')
    #print(df_original.to_string())
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def validate_estate_upload_H4_PC_001_1(conn, entity, entity_number, df):
  print('validate_estate_upload_H4_PC_001_1')
  # Partner Channel Solution November 2023
  # Validates estate upload dataframe (df) for entity with entity_number read from estate upload file. Conn is the connection object to the OPF database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()
    print('In top ')

    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce warnings
    
    ids                 = df_in["uprn"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="uprn")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["uprn","excel_row_num", "building_name", "address","postcode"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"****ERROR - {num_dup_uprn} occurrences of duplicate UPRNs have been found - please correct and re-submit.\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"
      validation['nve']   = validation['nve'] + num_dup_uprn
      validation['validation_messages']    = vm
      validation['ef']    = 0 
      validation['em']    = '******Input upload estate Lite file has failed validation - see messages in upload log' 
      validation['validated_df'] = df
      return validation

#   Validate numeric columns don't contain strings. If they do then report and return  
    numeric_columns_list =      ['uprn','gia_m2','roof_space_m2','data_year','latitude_dd', 'longitude_dd','baseline_annual_elec_kwh','baseline_annual_gas_kwh',\
      'baseline_annual_oil_kwh','baseline_annual_lpg_kwh','onsite_generation_annual_kwh', 'baseline_annual_cert_green_kwh', 'dec_score','epc','baseline_elec_cost_per_kwh',\
      'baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','baseline_cert_green_cost_per_kwh',\
      'exist_non_solar_decarb_heat_annual_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh' ]
    
    ret               = log_str_types_in_numeric_cols(df_in, numeric_columns_list)

    ef                = ret['ef']
    em                = ret['em']
    nerrs             = ret['nerrs']
    out_log           = ret['out_log']

    if ef == 2:
      validation['ef']                  = 2
      validation['em']                 = f"**Error while checking for presence of strings in numeric fields - \n {em}"
      validation['validated_df']        = ''
      validation['validation_messages'] = ''
      return validation
    
    if nerrs > 0:
      validation['ef']                  = 0
      validation['em']                 = ''
      validation['validated_df']        = ''
      validation['validation_messages'] = out_log
      validation['nve']                 = nerrs
      return validation      
    
#    Read building types from table cibse_benchmarks
    
    with conn.cursor() as cursor:
      sql1             = f"SELECT building_type FROM {bm.benchmark_table_name};"
      cursor.execute(sql1)
      t_building_types = cursor.fetchall()
      keys             = ("building_type","dummy_key")
      building_types   = [dict(zip(keys, values)) for values in t_building_types]

#    Validate building_type  

    bt_mess                = ''
    fuw_mess               = ''
    ecw_mess               = ''
    ll_mess                = ''
    nbterr                 = 0
    nxffuw                 = 0
    nxfecw                 = 0
    nlatlongw              = 0
    
    for index, row in df_in.iterrows():
      excel_row_num        = row['excel_row_num']
      bt                   = row['building_type']

      elec_kwh        = row['baseline_annual_elec_kwh']
      gas_kwh         = row['baseline_annual_gas_kwh']
      oil_kwh         = row['baseline_annual_oil_kwh']
      lpg_kwh         = row['baseline_annual_lpg_kwh']
      cert_e_kwh      = row['baseline_annual_cert_green_kwh']
      onsite_kwh      = row['onsite_generation_annual_kwh']    

      elec_cost       = row['baseline_elec_cost_per_kwh']
      gas_cost        = row['baseline_gas_cost_per_kwh']
      oil_cost        = row['baseline_oil_cost_per_kwh']
      lpg_cost        = row['baseline_lpg_cost_per_kwh']
      cert_e_cost     = row['baseline_cert_green_cost_per_kwh'] 

      latitude        = row['latitude_dd']
      longitude       = row['longitude_dd']
      
    # Validate building types are valid. Issue validation error if they are not valid.
      if not any(d['building_type'] == bt for d in building_types):
        nbterr           = nbterr + 1
        bt_mess          = bt_mess + f"****ERROR - Building type - {bt} - on Excel row {excel_row_num} is not valid \n"
    # Do cross field validations: -
    # 1 - Issue warning if all fossil fuel energy usage are zero that emissions can't be calculated
    # 2 - Issue warning an energy usage figure has been provided but no coresponding energy unit cost has been provided.
      if (elec_kwh == 0) and (gas_kwh == 0) and (oil_kwh == 0) and (lpg_kwh == 0) and (onsite_kwh == 0):
        nxffuw           = nxffuw + 1
        fuw_mess         = fuw_mess + f"----Warning: No fossil fuel usage has been reported for building on Excel row {excel_row_num} which means emissions cannot be calculated \n"
      if (elec_kwh > 0) and (elec_cost == 0):
        nxfecw           = nxfecw + 1
        ecw_mess         = ecw_mess + f"----Warning: Electricity usage has been reported for building on Excel row {excel_row_num} but no associated unit cost which means electricity costs cannot be calculated. \n"

      if (gas_kwh > 0) and (gas_cost == 0):
        nxfecw           = nxfecw + 1
        ecw_mess         = ecw_mess + f"----Warning: Gas usage has been reported for building on Excel row {excel_row_num} but no associated unit cost which means gas costs cannot be calculated. \n"

      if (oil_kwh > 0) and (oil_cost == 0):
        nxfecw           = nxfecw + 1
        ecw_mess         = ecw_mess + f"----Warning: Oil usage has been reported for building on Excel row {excel_row_num} but no associated unit cost which means oil costs cannot be calculated. \n"

      if (lpg_kwh > 0) and (lpg_cost == 0):
        nxfecw           = nxfecw + 1
        ecw_mess         = ecw_mess + f"----Warning: LPG usage has been reported for building on Excel row {excel_row_num} but no associated unit cost which means LPG costs cannot be calculated. \n"

      if (cert_e_kwh > 0) and (cert_e_cost == 0):
        nxfecw           = nxfecw + 1
        ecw_mess         = ecw_mess + f"----Warning: Certified green electricity usage has been reported for building on Excel row {excel_row_num} but no associated unit cost which means certified green electricity costs cannot be calculated. \n"        

      if latitude == 0:
        nlatlongw        = nlatlongw + 1
        ll_mess          = ll_mess + f"----Warning: Latitude on Excel row {excel_row_num} is zero\n" 
        
      if longitude == 0:
        nlatlongw        = nlatlongw + 1
        ll_mess          = ll_mess + f"----Warning: Longitude on Excel row {excel_row_num} is zero\n" 

    if nbterr > 0 : 
      vm                 = vm + f"{nbterr} invalid building types have been found:-\n"
      vm                 = vm + bt_mess
      validation['nve']  = validation['nve'] + nbterr

    if nxffuw > 0 : 
      vm                 = vm + f"{nxffuw} warnings for buildings with no fossil fuel energy usage:-\n"
      vm                 = vm + fuw_mess
      validation['nve']  = validation['nve'] + nxffuw

    if nxfecw > 0 : 
      vm                 = vm + f"{nxfecw} warnings for buildings with energy usage data with no corresponding unit costs:-\n"
      vm                 = vm + ecw_mess
      validation['nve']  = validation['nve'] + nxfecw    

    if nlatlongw > 0 : 
      vm                 = vm + f"{nlatlongw} warnings for latitude or longitude set to zero:-\n"
      vm                 = vm + ll_mess
      validation['nve']  = validation['nve'] + nlatlongw  
    
    thisyear = datetime.now ().year 

    # Validate remaining fields in dataframe
  
    df_s                 = df_in[['excel_row_num','uprn','building_name','address','postcode','latitude_dd','longitude_dd','gia_m2','roof_space_m2','data_year'\
                                ,'baseline_annual_elec_kwh','baseline_annual_gas_kwh','baseline_annual_oil_kwh','baseline_annual_lpg_kwh','onsite_generation_annual_kwh'\
                                ,'exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh','baseline_annual_cert_green_kwh','exist_non_solar_decarb_heat_annual_kwh'\
                                ,'baseline_elec_cost_per_kwh','baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','baseline_cert_green_cost_per_kwh'\
                                ,'source_of_heating','source_of_dhw','dec_score','epc'\
                                ]].copy()

    # Reset index to excel_row_number
    
    df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    schema = Schema([
      Column('uprn',[CustomElementValidation(lambda d: d > 0, "Building ID must be greater than zero")]),
      Column('building_name',[CustomElementValidation(lambda d: len(d) > 0, "Building_name is missing ")]),
      Column('address',[CustomElementValidation(lambda d: len(d) > 0, "Address is missing ")]),
      Column('postcode',[CustomElementValidation(lambda d: len(d) > 0, "Postcode is missing ")]),
      Column('latitude_dd', [InRangeValidation(-90, 90)]),
      Column('longitude_dd', [InRangeValidation(-180, 180)]),
      Column('gia_m2',   [CustomElementValidation(lambda d: d >= 0, "Gross internal area m2 must be >= to zero")]),
      Column('roof_space_m2', [CustomElementValidation(lambda d: d >= 0, "Roof space m2 must be >= to zero")]),
      Column('data_year',   [InRangeValidation(2017, 2050)]),
      Column('baseline_annual_elec_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual electricity usage must be >= to zero")]),      
      Column('baseline_annual_gas_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual gas usage must be >= to zero")]), 
      Column('baseline_annual_oil_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual oil usage must be >= to zero")]), 
      Column('baseline_annual_lpg_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual LPG usage must be >= to zero")]),
      Column('baseline_annual_cert_green_kwh', [CustomElementValidation(lambda d: d >= 0, "Electricity purchase from REGO sources must be >= to zero")]),
      Column('source_of_heating',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('source_of_dhw',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('dec_score', [CustomElementValidation(lambda d: d >= 0, "DEC score must be >= to zero")]),
      Column('epc', [CustomElementValidation(lambda d: d >= 0, "EPC must be >= to zero")]),
      Column('baseline_elec_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Electricity cost per kWh must be >= to zero")]),
      Column('baseline_gas_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Gas cost per kWh must be >= to zero")]),
      Column('baseline_oil_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Oil cost per kWh must be >= to zero")]),
      Column('baseline_lpg_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "LPG cost per kWh must be >= to zero")]),
      Column('baseline_cert_green_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "REGO electricity cost per kWh must be >= to zero")]),
      Column('onsite_generation_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "Total onsite generation kWh must be >= to zero")]),
      Column('exist_non_solar_decarb_heat_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual non-solar decarbonised heat usage kWh must be >= to zero")]),
      Column('exist_solar_pv_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual solar PV usage kWh >= to zero")]),
      Column('exist_solar_thermal_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "Annual solar thermal usage kWh must be >= to zero")]),
    ])

    errors            = schema.validate(df_s)
    errors_index_rows = [e.row for e in errors]
    nscerrs           = len(errors)
    
    if nscerrs > 0:
      vm                = vm + f"{nscerrs} field validation errors have been found:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"

    nvw                 = nxffuw + nxfecw + nlatlongw
    nve                 = num_dup_uprn + nbterr + nscerrs
    #print('df_original at end validation')
    #print(df_original.to_string())
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def validate_estate_lite_upload_v002(conn, entity, entity_number, df):
  print('At top validate_estate_upload_v002')
  # Tactical Solution August 2023
  # Validates estate upload dataframe (df) for entity with entity_number read from estate lite upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()

    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce warnings
    
    ids                 = df_in["uprn"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="uprn")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["uprn","excel_row_num", "building_name", "address","postcode"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"****ERROR - {num_dup_uprn} occurrences of duplicate UPRNs have been found - please correct and re-submit.\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"
      validation['nve']   = validation['nve'] + num_dup_uprn
      validation['validation_messages']    = vm
      validation['ef']    = 0 
      validation['em']    = '******Input upload estate Lite file has failed validation - see messages in upload log' 
      validation['validated_df'] = df
      return validation
#  Validate action codes 

    df_actions            = df_in[['action','excel_row_num']].copy()
    df_actions            = df_actions.set_index('excel_row_num')
    schema = Schema([
      Column('action', [InListValidation(['DELETE', 'UPDATE'])]),])
    
    numactionerrs         = 0
    errors                = schema.validate(df_actions)
    errors_index_rows     = [e.row for e in errors]

    numactionerrs         = len(errors)
    if numactionerrs > 0:
      vm                = vm + f"{numactionerrs} errors have been found in ACTION codes:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"    

#    Remove DELETE records from further validation to stop spurious messages - DELETE only needs action code and uprn

    df_in.drop(df_in[df_in['action'] == "DELETE"].index, inplace = True)

#   Validate numeric columns don't contain strings. If they do then report and return  
    numeric_columns_list =      ['gia_m2','roof_space_m2','data_year','baseline_annual_elec_kwh','baseline_annual_gas_kwh',\
      'baseline_annual_oil_kwh','baseline_annual_lpg_kwh', 'baseline_annual_cert_green_kwh', 'dec_score','epc','baseline_elec_cost_per_kwh',\
      'baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','baseline_cert_green_cost_per_kwh',\
      'onsite_annual_generation_kwh','exist_non_solar_decarb_heat_annual_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh' ]
    
    ret               = log_str_types_in_numeric_cols(df_in, numeric_columns_list)

    ef                = ret['ef']
    em                = ret['em']
    nerrs             = ret['nerrs']
    out_log           = ret['out_log']

    if ef == 2:
      validation['ef']                  = 2
      validation['em']                 = f"**Error while checking for presence of strings in numeric fields - \n {em}"
      validation['validated_df']        = ''
      validation['validation_messages'] = ''
      return validation
    
    if nerrs > 0:
      validation['ef']                  = 0
      validation['em']                 = ''
      validation['validated_df']        = ''
      validation['validation_messages'] = out_log
      validation['nve']                 = nerrs
      return validation      
    
#    Read building types from table cibse_benchmarks
    
    with conn.cursor() as cursor:
      sql1             = f"SELECT building_type FROM {bm.benchmark_table_name};"
      cursor.execute(sql1)
      t_building_types = cursor.fetchall()
      keys             = ("building_type","dummy_key")
      building_types   = [dict(zip(keys, values)) for values in t_building_types]

#    Validate building_type  

    bt_mess                = ''
    nbterr                 = 0
    num_miss_building_type = 0
    
    for index, row in df_in.iterrows():
      excel_row_num        = row['excel_row_num']
      bt                   = row['building_type']

#      if bt == None or bt == '':
#        num_miss_building_type = num_miss_building_type + 1
#        bt_mess          = bt_mess + f"++++Warning - Building type on Excel row {excel_row_num} is missing \n"        
#      elif not any(d['building_type'] == bt for d in building_types):
      if not any(d['building_type'] == bt for d in building_types):
        nbterr           = nbterr + 1
        bt_mess          = bt_mess + f"****ERROR - Building type - {bt} - on Excel row {excel_row_num} is not valid \n"
        
    if nbterr > 0 or num_miss_building_type > 0:
      vm                 = vm + f"{nbterr} errors and {num_miss_building_type} missing building type warnings have been found in BUILDING TYPES:-\n"
      vm                 = vm + bt_mess
      validation['nve']  = validation['nve'] + nbterr

    thisyear = datetime.now ().year 

    # Validate remaining fields in dataframe
  
    df_s                 = df_in[['excel_row_num','building_name','address','postcode','listed','construction_year','gia_m2','roof_space_m2','data_year'\
                                ,'baseline_annual_elec_kwh','baseline_annual_gas_kwh','baseline_annual_oil_kwh','baseline_annual_lpg_kwh','baseline_annual_cert_green_kwh','source_of_heating','source_of_dhw','dec_score'\
                                ,'epc','baseline_elec_cost_per_kwh','baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','baseline_cert_green_cost_per_kwh','onsite_annual_generation_kwh','exist_non_solar_decarb_heat_annual_kwh'\
                                ,'exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh']].copy()

    # Reset index to excel_row_number
    
    df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    schema = Schema([
      Column('building_name',[CustomElementValidation(lambda d: len(d) > 0, "Building_name is missing ")]),
      Column('address',[CustomElementValidation(lambda d: len(d) > 0, "Address is missing ")]),
      Column('postcode',[CustomElementValidation(lambda d: len(d) > 0, "Postcode is missing ")]),
      Column('listed',[InListValidation(['YES', 'NO',''])]),
      Column('construction_year',[InListValidation(["2020","2015","2010","2005","2000","1990S","1980S","1970S","1960S","1950S","PRE 1950","PRE 1900S",""])]),
      Column('gia_m2',   [CustomElementValidation(lambda d: d >= 0, "gia_m2 must be >= to zero")]),
      Column('roof_space_m2', [CustomElementValidation(lambda d: d >= 0, "roof_space_m2 must be >= to zero")]),
      Column('data_year',   [InRangeValidation(2017, 2030)]),
      Column('baseline_annual_elec_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_elec_kwh must be a >= to zero")]),      
      Column('baseline_annual_gas_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_gas_kwh must be >= to zero")]), 
      Column('baseline_annual_oil_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_oil_kwh must be >= to zero")]), 
      Column('baseline_annual_lpg_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_lpg_kwh must be >= to zero")]),
      Column('baseline_annual_cert_green_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_cert_green_kwh must be >= to zero")]),
      Column('source_of_heating',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('source_of_dhw',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('dec_score', [CustomElementValidation(lambda d: d >= 0, "dec_score must be >= to zero")]),
      Column('epc', [CustomElementValidation(lambda d: d >= 0, "epc must be >= to zero")]),
      Column('baseline_elec_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline elec_cost_per_kwh must be >= to zero")]),
      Column('baseline_gas_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline gas_cost_per_kwh must be >= to zero")]),
      Column('baseline_oil_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline oil_cost_per_kwh must be >= to zero")]),
      Column('baseline_lpg_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline lpg_cost_per_kwh must be >= to zero")]),
      Column('baseline_cert_green_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline cert_green_cost_per_kwh must be >= to zero")]),
      Column('onsite_annual_generation_kwh', [CustomElementValidation(lambda d: d >= 0, "onsite_annual_generation_kwh must be >= to zero")]),
      Column('exist_non_solar_decarb_heat_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_non_solar_decarb_heat_annual_kwh must be >= to zero")]),
      Column('exist_solar_pv_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_pv_annual_kwh must be >= to zero")]),
      Column('exist_solar_thermal_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_thermal_annual_kwh must be >= to zero")]),
    ])

    errors            = schema.validate(df_s)
    errors_index_rows = [e.row for e in errors]
    nscerrs           = len(errors)
    
    if nscerrs > 0:
      vm                = vm + f"{nscerrs} errors have been found in other fields:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"
#{'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}    
    nvw                 = num_dup_uprn + num_miss_building_type
    nve                 = nuprnerrs + numactionerrs + nbterr + nscerrs
    #print('df_original at end validation')
    #print(df_original.to_string())
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def validate_estate_upload_v002(conn, entity, entity_number, df):
  print('At top validate_estate_upload_v002')
  # Rationalised version Jan 2023
  # Validates estate upload dataframe (df) for entity with entity_number read from upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()
    #print('df_original at start validate')
    #print(df_original.to_string())
   
    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce warnings
    
    ids                 = df_in["uprn"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="uprn")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["uprn","excel_row_num", "building_name", "address","postcode"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"-----WARNING - {num_dup_uprn} occurrences of duplicate UPRNs have been found - review advised\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"
      validation['nvw']   = validation['nvw'] + num_dup_uprn
    
#  Validate action codes 

    df_actions            = df_in[['action','excel_row_num']].copy()
    df_actions            = df_actions.set_index('excel_row_num')
    schema = Schema([
      Column('action', [InListValidation(['DELETE', 'UPDATE'])]),])
    
    numactionerrs         = 0
    errors                = schema.validate(df_actions)
    errors_index_rows     = [e.row for e in errors]

    numactionerrs         = len(errors)
    if numactionerrs > 0:
      vm                = vm + f"{numactionerrs} errors have been found in ACTION codes:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"    

#    Remove DELETE records from further validation to stop spurious messages - DELETE only needs action code and uprn

    df_in.drop(df_in[df_in['action'] == "DELETE"].index, inplace = True)

#   Validate numeric columns don't contain strings. If they do then report and return  
    numeric_columns_list =      ['gia_m2','roof_space_m2','data_year','baseline_annual_elec_kwh','baseline_annual_gas_kwh',\
      'baseline_annual_oil_kwh','baseline_annual_lpg_kwh', 'dec_score','epc','baseline_elec_cost_per_kwh','baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','onsite_generation_annual_kwh','exist_non_solar_decarb_heat_annual_kwh',\
      'number_of_ev_charge_sockets', 'charging_capacity_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh' ]
    
    ret               = log_str_types_in_numeric_cols(df_in, numeric_columns_list)

    ef                = ret['ef']
    em                = ret['em']
    nerrs             = ret['nerrs']
    out_log           = ret['out_log']

    if ef == 2:
      validation['ef']                  = 2
      validation['em']                 = f"**Error while checking for presence of strings in numeric fields - \n {em}"
      validation['validated_df']        = ''
      validation['validation_messages'] = ''
      return validation
    
    if nerrs > 0:
      validation['ef']                  = 0
      validation['em']                 = ''
      validation['validated_df']        = ''
      validation['validation_messages'] = out_log
      validation['nve']                 = nerrs
      return validation      
    
#    Read building types from table cibse_benchmarks
    
    with conn.cursor() as cursor:
      sql1             = f"SELECT building_type FROM {bm.benchmark_table_name};"
      cursor.execute(sql1)
      t_building_types = cursor.fetchall()
      keys             = ("building_type","dummy_key")
      building_types   = [dict(zip(keys, values)) for values in t_building_types]

#    Validate building_type  

    bt_mess                = ''
    nbterr                 = 0
    num_miss_building_type = 0
    
    for index, row in df_in.iterrows():
      excel_row_num        = row['excel_row_num']
      bt                   = row['building_type']

#      if bt == None or bt == '':
#        num_miss_building_type = num_miss_building_type + 1
#        bt_mess          = bt_mess + f"++++Warning - Building type on Excel row {excel_row_num} is missing \n"        
#      elif not any(d['building_type'] == bt for d in building_types):
      if not any(d['building_type'] == bt for d in building_types):
        nbterr           = nbterr + 1
        bt_mess          = bt_mess + f"****ERROR - Building type - {bt} - on Excel row {excel_row_num} is not valid \n"
        
    if nbterr > 0 or num_miss_building_type > 0:
      vm                 = vm + f"{nbterr} errors and {num_miss_building_type} missing building type warnings have been found in BUILDING TYPES:-\n"
      vm                 = vm + bt_mess
      validation['nve']  = validation['nve'] + nbterr

    thisyear = datetime.now ().year 

    # Validate remaining fields in dataframe
  
    df_s                 = df_in[['excel_row_num','building_name','address','postcode','under_control','remain_in_portfolio','listed','construction_year','gia_m2','roof_space_m2','data_year'\
                                ,'baseline_annual_elec_kwh','baseline_annual_gas_kwh','baseline_annual_oil_kwh','baseline_annual_lpg_kwh','source_of_heating','source_of_dhw','dec_score'\
                                ,'epc','baseline_elec_cost_per_kwh','baseline_gas_cost_per_kwh','baseline_oil_cost_per_kwh','baseline_lpg_cost_per_kwh','onsite_generation_annual_kwh','exist_non_solar_decarb_heat_annual_kwh'\
                                ,'car_park_available','number_of_ev_charge_sockets','charging_capacity_kwh','exist_solar_pv_annual_kwh','exist_solar_thermal_annual_kwh']].copy()

    # Reset index to excel_row_number
    
    df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    schema = Schema([
      Column('building_name',[CustomElementValidation(lambda d: len(d) > 0, "Building_name is missing ")]),
      Column('address',[CustomElementValidation(lambda d: len(d) > 0, "Address is missing ")]),
      Column('postcode',[CustomElementValidation(lambda d: len(d) > 0, "Postcode is missing ")]),
      Column('under_control',[InListValidation(['YES', 'NO',''])]),
      Column('remain_in_portfolio',[InListValidation(['YES', 'NO',''])]),
      Column('listed',[InListValidation(['YES', 'NO',''])]),
      Column('construction_year',[InListValidation(["2020","2015","2010","2005","2000","1990S","1980S","1970S","1960S","1950S","PRE 1950","PRE 1900S",""])]),
      Column('gia_m2',   [CustomElementValidation(lambda d: d >= 0, "gia_m2 must be >= to zero")]),
      Column('roof_space_m2', [CustomElementValidation(lambda d: d >= 0, "roof_space_m2 must be >= to zero")]),
      Column('data_year',   [InRangeValidation(2017, 2030)]),
      Column('baseline_annual_elec_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_elec_kwh must be a >= to zero")]),      
      Column('baseline_annual_gas_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_gas_kwh must be >= to zero")]), 
      Column('baseline_annual_oil_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_oil_kwh must be >= to zero")]), 
      Column('baseline_annual_lpg_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline annual_lpg_kwh must be >= to zero")]),
      Column('source_of_heating',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('source_of_dhw',[InListValidation(['ELECTRICITY','GAS','OIL','LPG',''])]),
      Column('dec_score', [CustomElementValidation(lambda d: d >= 0, "dec_score must be >= to zero")]),
      Column('epc', [CustomElementValidation(lambda d: d >= 0, "epc must be >= to zero")]),
      Column('baseline_elec_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline elec_cost_per_kwh must be >= to zero")]),
      Column('baseline_gas_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline gas_cost_per_kwh must be >= to zero")]),
      Column('baseline_oil_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline oil_cost_per_kwh must be >= to zero")]),
      Column('baseline_lpg_cost_per_kwh', [CustomElementValidation(lambda d: d >= 0, "Baseline lpg_cost_per_kwh must be >= to zero")]),
      Column('onsite_generation_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "onsite_generation_annual_kwh must be >= to zero")]),
      Column('exist_non_solar_decarb_heat_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_non_solar_decarb_heat_annual_kwh must be >= to zero")]),
      Column('car_park_available',[InListValidation(['YES', 'NO',''])]),
      Column('number_of_ev_charge_sockets', [CustomElementValidation(lambda d: vcl_check_int(d), "number_of_ev_charge_sockets must be an whole number")]),      
      Column('charging_capacity_kwh', [CustomElementValidation(lambda d: d >= 0, "charging_capacity_kwh must be >= to zero")]),
      Column('exist_solar_pv_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_pv_annual_kwh must be >= to zero")]),
      Column('exist_solar_thermal_annual_kwh', [CustomElementValidation(lambda d: d >= 0, "exist_solar_thermal_annual_kwh must be >= to zero")]),
    ])

    errors            = schema.validate(df_s)
    errors_index_rows = [e.row for e in errors]
    nscerrs           = len(errors)
    
    if nscerrs > 0:
      vm                = vm + f"{nscerrs} errors have been found in other fields:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"
#{'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}    
    nvw                 = num_dup_uprn + num_miss_building_type
    nve                 = nuprnerrs + numactionerrs + nbterr + nscerrs
    #print('df_original at end validation')
    #print(df_original.to_string())
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def vcl_check_decimal(dec):
  if round(dec) != dec:
   return True
  else:
   return False
def vcl_check_int(dec):
  if round(dec) == dec:
   return True
  else:
   return False
def validate_projects_initialisation_upload(conn, entity, entity_number, df):
  print('At top validate_projects_initialisation_upload')
  # Validates projects initialisation upload dataframe (df) for entity with entity_number read from upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()
    assessed_as          =["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FTHR IMPV","ASSESSED/NV"]

    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce error
    
    ids                 = df_in["Building ID"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="Building ID")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["Building ID","excel_row_num", "Building name", "Building type"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"****ERROR - {num_dup_uprn} occurrences of duplicate UPRNs have been found - upload cannot proceed. Please review and correct\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"
      validation['nve']   = validation['nve'] + num_dup_uprn
    
    # Validate remaining fields in dataframe
  
    df_s                 = df_in[['excel_row_num','Fabric Roof','Fabric Windows','Fabric Doors','Fabric (Walls)',\
                               'Pipe Insulation','Heating Controls','LED Lighting','BMS Upgrade/Controls','Variable Speed Drives','Voltage optimisation','Smart Microgrid',\
                               'Energy Efficient Chillers/Ventilation','Boiler Upgrade','CHP','Heat pump (GAS SAVING)','Battery Demand Response','Thermal energy store',\
                               'Solar PV Power','Solar Thermal','Wind Power','Hydropower','Bioenergy','Heat Network']].copy()

    # Reset index to excel_row_number

    df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    schema = Schema([
    #  Column('Full Retrofit',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Fabric Roof',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Fabric Windows',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Fabric Doors',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Fabric (Walls)',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Pipe Insulation',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Heating Controls',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('LED Lighting',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('BMS Upgrade/Controls',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Variable Speed Drives',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Voltage optimisation',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Smart Microgrid',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Energy Efficient Chillers/Ventilation',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Boiler Upgrade',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('CHP',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Heat pump (GAS SAVING)',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Battery Demand Response',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Thermal energy store',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Solar PV Power',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Solar Thermal',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Wind Power',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Hydropower',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Bioenergy',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
      Column('Heat Network',[InListValidation(['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'IN PLACE', 'FTHR IMPV', 'ASSESSED/NV',""])]),
     ])

    errors            = schema.validate(df_s)
    errors_index_rows = [e.row for e in errors]
    nscerrs           = len(errors)
    
    if nscerrs > 0:
      vm                = vm + f"{nscerrs} errors have been found in other fields:-\n"
      for error in errors:
        vm              = vm + f"****ERROR - {error}\n"
#{'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}    

    nve                 = nuprnerrs + nscerrs + num_dup_uprn

    validation['ef'] = 0
    validation['em'] = ''
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = 0
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def uprn_exists_in_raw_data(conn, uprn, entity_number):
  
  # Tests whether a record with uprn/entity number composite primary key provided exists 
  # in the raw_estate_data table of database connected by connection object conn
  
  with conn.cursor() as cursor:
        dsql1 = "SELECT * FROM raw_estate_data WHERE (entity_number ="
        dsql2 = f"{entity_number}"
        dsql3 = f") and (uprn ="
        dsql4 = f"{uprn})"
        dsql  = f"{dsql1} {dsql2} {dsql3} {dsql4}"

        cursor.execute(dsql)
        output = cursor.fetchall() 
        if len(output) > 0:
          return True
        else:
          return False

def cons_valid_date(datestr):
  # Validates a date provided as a date string against the standard date format used in Clean Onshore
  import datetime
  print('In cons valid date - 1')
  date_format = '%d/%m/%Y'
  try:
    if datestr == None:
      return True
    else:
      print('In cons valid date - 2')
      date_obj = datetime.datetime.strptime(datestr, date_format)
      print('In cons valid date - 3')
    return True
  except ValueError:
    return False

def remove_spurious_projects(conn, dfin, project_types):
  # Conn is the database connection object. dfin is the input dataframe containing project data. project_types is a list of dicts containing project type name and project type id.
  # Removes those projects from a project details dataframe, dfin, read from an Excel project details upload workbook, which do not have an entry already in the database
  # (i.e. most likely added manually to the project details upload without being created correctly via the project assessment upload).
  # Returns the cleaned up dataframe in dictionary ret['df_nospur'].
  try:
    ret       = {'ef':0, 'em':'', 'df_nospur': '','log':'', 'nspur': 0}
    bd        = []
    log       = f"Spurious projects removed:-\n"
    uprn_col  = dfin['Building ID']
    tyid_col  = dfin['Project type']
    nspur     = 0
    
    # Loop down the input dataframe. For each record (project) convert the project_type name to project_type_id then execute a query on the projects table
    # to find a record with the same uprn and project_type_id as the input dataframe record. If one can't be found then this is a spurious entry so mark it for deletion.
    with conn.cursor() as cursor:
      for index, row in dfin.iterrows():
        foundtype     = False
        typeisvalid   = False
        project_type  = row['Project type']
        uprn          = row['Building ID']
        excel_row_num = row['excel_row_num']
        
        # Convert project type to project type id. If project type from input dataframe can't be found in list of valid project_types then it is illegal so mark for deletion
        for p in project_types:
          typeisvalid = False
          if p['name'] == project_type:
            project_type_id = p['project_type_id']
            typeisvalid = True
            break
        if not typeisvalid:
          nspur     = nspur + 1
          bd.append(False) # This row will be deleted
          log       = log + f"Project type on row {excel_row_num} for Building ID {uprn} is invalid - this row has been removed\n"

        else:
          qsql      = f"SELECT project_type_id FROM projects WHERE (uprn = {uprn}) AND (project_type_id = {project_type_id});"
          cursor.execute(qsql)
          t_output  = cursor.fetchall()
          keys      = ("project_type_id")
          output_en = [dict(zip(keys, values)) for values in t_output]
          
          if len(output_en) > 0: # Found project with this project_type_id for this uprn
            bd.append(True) # This row will remain
          else:
            nspur     = nspur + 1
            bd.append(False) # This row will be deleted
            log       = log + f"Project type {project_type} on row {excel_row_num} for Building ID {uprn} has not been set up on the database - this row has been removed\n"

            
      df_nospur = dfin[bd] # Remove spurious projects
      ret['df_nospur'] = df_nospur
      ret['log']       = log
      ret['nspur']     = nspur
     # print(f"End of remove_spurious_projects*********\n dfin = \n {dfin.to_string()} \n df_nospur = \n {df_nospur.to_string()}\n ******")
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
    return ret
  
def validate_non_solar_projects_details_upload(conn, entity, entity_number, df):
  print('At top validate non solar projects details_upload')
  #import kv_calcs as kc
  import datetime
  # Validates non_solar project details dataframe (df) for entity with entity_number read from project details upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Non Solar projects validation completed successfully','validated_df':'','validation_messages':'','missing_data_messages':'','nvw':0,'nve':0,'mdw':0}
    vm                   = ''
    mdm                  = ''
    df_in                = df.copy()
    df_original          = df.copy()
    assessed_as          =["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FTHR IMPV","ASSESSED/NV"]

    # Exit if dataframe has no records
    nrows                = df_in.shape[0]
    if nrows == 0:
      return validation
      
    nuprnerrs            = 0
    nderrs               = 0
    ndwars               = 0
    mdwars               = 0
    num_rows             = df_in.shape[0]

    df_s                 = df_in[['excel_row_num','Assessed','Status','Utility','Lifetime (yrs)','Saving %','CAPEX','Delivery date']].copy()

    # Reset index to excel_row_number
    #df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    print('Start of non solar validation')
    
    for d in df_s.to_dict(orient="records"):

      utillist = ['GAS','ELEC']
      util     = d['Utility']
      if utillist.count(util) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid utility {d['utility']} on row {d['excel_row_num']}. \n"      

      assed    = d['Assessed']
      if assessed_as.count(assed) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid assessed as value {d['assessed']} on row {d['excel_row_num']}. \n" 
# project_status follows Catapult convention
      stlist   = ['Concept','Feasibility','Business Case','Procurement','Abandoned','Benefits Realisation' ]
      status   = d['Status']
      if stlist.count(status) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid status {d['status']} on row {d['excel_row_num']}. \n"       
      
      pfact    = d['Lifetime (yrs)'] 
      if isinstance(pfact,float) or isinstance(pfact,int):
        if pfact == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning persistence factor is zero on row {d['excel_row_num']}. \n"
        if pfact < 0:
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['persistence_factor'] }  for persistence factor on row {d['excel_row_num']}. \n" 
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Persistence factor is not numeric on row {d['excel_row_num']}. \n"          
          
      savng    = d['Saving %'] 
      if isinstance(savng,float) or isinstance(savng,int):
        if savng == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning saving_percent is zero on row {d['excel_row_num']}. \n"        
        if savng < 0 or savng > 100:
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['saving_percent'] } for percent saving on row {d['excel_row_num']}. \n"
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Percent saving is not numeric on row {d['excel_row_num']}. \n" 
          
      capex    = d['CAPEX'] 
      if isinstance(capex,float) or isinstance(capex,int):
        if capex == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning capex is zero on row {d['excel_row_num']}. \n"          
        if capex < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['capex'] } for capex on row {d['excel_row_num']}. \n"  
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Capex is not numeric on row {d['excel_row_num']}. \n" 
          
      ddate = d['Delivery date']
#      if ddate == '0000-00-00':
#        print(f"ddate is 0000-00-00 on row {d['excel_row_num']}. \n" )
      if ddate != '0000-00-00' :
        if not isinstance(ddate, datetime.date):
          nderrs            = nderrs + 1
          vm                = vm + f"****Invalid date {d['delivery_date']} on row {d['excel_row_num']}. \n"
        
    nve                 = nuprnerrs + nderrs
    nvw                 = ndwars
    mdw                 = mdwars
    print('nve nvw mdw')
    print(nve)
    print(nvw)
    print(mdw)
    if nve == 0 and nvw == 0:
      validation['ef'] = 0
      validation['em'] = '-----Non Solar projects validation completed successfully.------'
    if nve == 0 and nvw > 0:
      validation['ef'] = 1
      validation['em'] = '++++Warning - there are validation warnings for non solar projects but updates will be performed.'      
    if nve > 0:  
      validation['ef'] = 2
      validation['em'] = '****ERROR - upload has failed non solar projects validation. No updates will be performed.'
      
    validation['validated_df']          = df_original
    print('validation messages')
    print(vm)
    validation['validation_messages']   = vm
    validation['missing_data_messages'] = mdm
    validation['nvw']                   = nvw
    validation['nve']                   = nve
    validation['mdw']                   = mdw
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - Non solar projects validation terminated with exception: \n {msg}"
    validation['validated_df']          = ''
    validation['validation_messages']   = ''
    validation['missing_data_messages'] = ''
    validation['nvw']                   = 0
    validation['nve']                   = 0
    validation['mdw']                   = 0    
    return validation

def validate_solar_pv_projects_details_upload(conn, entity, entity_number, df):
  print('At top validate_solar_pv_projects details_upload')
  #import kv_calcs as kc
  import datetime
  # Validates solar pv project details dataframe (df) for entity with entity_number read from project details upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Solar pv validation completed successfully','validated_df':'','validation_messages':'','missing_data_messages':'','nvw':0,'nve':0,'mdw':0}
    vm                   = ''
    mdm                  = ''
    df_in                = df.copy()
    df_original          = df.copy()
    assessed_as          =["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FTHR IMPV","ASSESSED/NV"]
#    print('df_original at start validate solar pv projects upload')
#    print(df_original.to_string())
    
    # Exit if dataframe has no records
    nrows                = df_in.shape[0]
    if nrows == 0:
      return validation
    
    # Validate uprns are in allowed range(s) for this entity
   
    uprn_mess            = ''
    nuprnerrs            = 0
    nderrs               = 0
    ndwars               = 0
    mdwars               = 0
    num_rows             = df_in.shape[0]

    # Reset index to excel_row_number
    #df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook
    
    df_s                 = df_in[['excel_row_num','Assessed','Status','Utility','Lifetime (yrs)','CAPEX','Delivery date','Solar roof type','Solar angle','Solar area m2','Solar KW peak','Corrected annual gen kWh']].copy()
#    print('Start of validation')
    
    for d in df_s.to_dict(orient="records"):

      utillist = ['GAS','ELEC']
      util     = d['Utility']
      if utillist.count(util) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid utility {d['Utility']} on row {d['excel_row_num']}. \n"      

      assed    = d['Assessed']
      
      if assessed_as.count(assed) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid assessed as value {d['Assessed']} on row {d['excel_row_num']}. \n" 

      status   = d['Status']
      
# project_status follows Catapult convention
      stlist   = ['Concept','Feasibility','Business Case','Procurement','Abandoned','Benefits Realisation' ]
      if stlist.count(status) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid status {d['Status']} on row {d['excel_row_num']}. \n"       
      
      pfact    = d['Lifetime (yrs)']
      
      if isinstance(pfact,float) or isinstance(pfact,int):
        if pfact == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning Lifetime is zero on row {d['excel_row_num']}. \n"
        if pfact < 0:
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['Lifetime (yrs)'] }  for Lifetime on row {d['excel_row_num']}. \n" 
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Lifetime is not numeric on row {d['excel_row_num']}. \n"          
          
      capex    = d['CAPEX']
      
      if isinstance(capex,float) or isinstance(capex,int):
        if capex == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning capex is zero on row {d['excel_row_num']}. \n"          
        if capex < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['CAPEX'] } for capex on row {d['excel_row_num']}. \n"  
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Capex is not numeric on row {d['excel_row_num']}. \n" 
          
      ddate = d['Delivery date']
#      if ddate == '0000-00-00':
#        print(f"ddate is 0000-00-00 on row {d['excel_row_num']}. \n" )
      if ddate != '0000-00-00' :
        if not isinstance(ddate, datetime.date):
          nderrs            = nderrs + 1
          vm                = vm + f"****Invalid date {d['Delivery date']} on row {d['excel_row_num']}. \n"
          
      coragk  = d['Corrected annual gen kWh']
      if isinstance(coragk,float) or isinstance(coragk,int):
        if coragk < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['Corrected annual gen kWh'] } for solar PV Corrected annual gen kWh on row {d['excel_row_num']}. \n"
        if coragk <= 0:  
          sangle  = d['Solar angle']
          
          anglist = ['Flat', 'Pitched']
          if anglist.count(sangle) == 0:
            nderrs            = nderrs + 1
            vm                = vm + f"****Invalid solar angle {d['Solar angle']} on row {d['excel_row_num']}. \n"       
            
          rtype   = d['Solar roof type']
          
          typlist = ['Flat Roof (Membrane)','Flat Roof (Deck)','Profile sheet','Concrete Tile','Clay tile','Slate']
          if typlist.count(rtype) == 0:
            nderrs            = nderrs + 1
            vm                = vm + f"****Invalid solar roof type {d['Solar roof type']} on row {d['excel_row_num']}. \n"  
    
          aream2  = d['Solar area m2']
          
          if isinstance(aream2,float) or isinstance(aream2,int):
            if aream2 == 0:
              mdwars = mdwars + 1
              mdm             = mdm + f"++++Missing data warning solar area m2 is zero on row {d['excel_row_num']}. \n"          
            if aream2 < 0 :
              nderrs = nderrs + 1
              vm              = vm + f"****Invalid value {d['Solar area m2'] } for solar area m2 on row {d['excel_row_num']}. \n"  
          else:
              nderrs = nderrs + 1
              vm              = vm + f"****Solar area m2 is not numeric on row {d['excel_row_num']}. \n"    
    
          solkwp  = d['Solar KW peak']
          
          if isinstance(solkwp,float) or isinstance(solkwp,int):
            if solkwp < 0 :
              nderrs = nderrs + 1
              vm              = vm + f"****Invalid value {d['Solar KW peak'] } for solar kw peak on row {d['excel_row_num']}. \n"  
          else:
              nderrs = nderrs + 1
              vm              = vm + f"****Solar kw peak is not numeric on row {d['excel_row_num']}. \n" 
      
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Solar Corrected annual gen kWh is not numeric on row {d['excel_row_num']}. \n" 

    nve                 = nuprnerrs + nderrs
    nvw                 = ndwars
    mdw                 = mdwars
#    print('nve nvw mdw')
#    print(nve)
#    print(nvw)
#    print(mdw)
    if nve == 0 and nvw == 0:
      validation['ef'] = 0
      validation['em'] = '-----Solar pv projects passed validation successfully.------'
    if nve == 0 and nvw > 0:
      validation['ef'] = 1
      validation['em'] = '++++Warning - there are validation warnings for solar pv projects but updates will be performed'      
    if nve > 0:  
      validation['ef'] = 2
      validation['em'] = '****ERROR - upload has failed solar pv validation. No updates will be performed.'
      
    validation['validated_df']          = df_original
#    print('validation messages')
#    print(vm)
    validation['validation_messages']   = vm
    validation['missing_data_messages'] = mdm
    validation['nvw']                   = nvw
    validation['nve']                   = nve
    validation['mdw']                   = mdw
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - Solar pv projects validation terminated with exception: \n {msg}"
    validation['validated_df']          = ''
    validation['validation_messages']   = ''
    validation['missing_data_messages'] = ''
    validation['nvw']                   = 0
    validation['nve']                   = 0
    validation['mdw']                   = 0    
    return validation
  
def validate_solar_thermal_projects_details_upload(conn, entity, entity_number, df):
  print('At top validate_solar_thermal_projects details_upload')
  #import kv_calcs as kc
  import datetime
  # Validates solar thermal project details dataframe (df) for entity with entity_number read from project details upload file. Conn is the connection object to the Decarb database.
  try:
    validation           = {'ef':0,'em':'Solar thermal validation completed successfully','validated_df':'','validation_messages':'','missing_data_messages':'','nvw':0,'nve':0,'mdw':0}
    vm                   = ''
    mdm                  = ''
    df_in                = df.copy()
    df_original          = df.copy()
    assessed_as          =["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FTHR IMPV","ASSESSED/NV"]
 #   print('df_original at start validate solar thermal projects upload')
 #   print(df_original.to_string())

    # Exit if dataframe has no records
    nrows                = df_in.shape[0]
    if nrows == 0:
      return validation

    uprn_mess            = ''
    nuprnerrs            = 0
    nderrs               = 0
    ndwars               = 0
    mdwars               = 0
    num_rows             = df_in.shape[0]

    df_s                 = df_in[['excel_row_num','Assessed','Status','Utility','Lifetime (yrs)','CAPEX','Delivery date', 'Solar area_m2', 'Corrected annual gen kWh']].copy()

    # Reset index to excel_row_number
    #df_s                 = df_s.set_index('excel_row_num') #So errors generated by Pandas validation show the row number as seen by user in Excel data load workbook

    print('Start of validation')
    
    for d in df_s.to_dict(orient="records"):

      utillist = ['GAS','ELEC']
      util     = d['Utility']
      if utillist.count(util) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid utility {d['utility']} on row {d['excel_row_num']}. \n"      

      assed    = d['Assessed']
      
      if assessed_as.count(assed) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid assessed as value {d['assessed']} on row {d['excel_row_num']}. \n" 

      status   = d['Status']
# project_status follows Catapult convention
      stlist   = ['Concept','Feasibility','Business Case','Procurement','Abandoned','Benefits Realisation' ]
      if stlist.count(status) == 0:
        nderrs            = nderrs + 1
        vm                = vm + f"****Invalid status {d['status']} on row {d['excel_row_num']}. \n"       
      
      pfact    = d['Lifetime (yrs)']
      
      if isinstance(pfact,float) or isinstance(pfact,int):
        if pfact == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning persistence factor is zero on row {d['excel_row_num']}. \n"
        if pfact < 0:
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['Lifetime (yrs)'] }  for Lifetime on row {d['excel_row_num']}. \n" 
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Lifetime is not numeric on row {d['excel_row_num']}. \n"          
          
      capex    = d['CAPEX']
      
      if isinstance(capex,float) or isinstance(capex,int):
        if capex == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning capex is zero on row {d['excel_row_num']}. \n"          
        if capex < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['CAPEX'] } for capex on row {d['excel_row_num']}. \n"  
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Capex is not numeric on row {d['excel_row_num']}. \n" 
          
      ddate = d['Delivery date']
#      if ddate == '0000-00-00':
#        print(f"ddate is 0000-00-00 on row {d['excel_row_num']}. \n" )
      if ddate != '0000-00-00' :
        if not isinstance(ddate, datetime.date):
          nderrs            = nderrs + 1
          vm                = vm + f"****Invalid date {d['Delivery date']} on row {d['excel_row_num']}. \n"
          
#      print('In validate thermal projects - d')
#      print(d)
      aream2  = d['Solar area m2']
      
      if isinstance(aream2,float) or isinstance(aream2,int):
        if aream2 == 0:
          mdwars = mdwars + 1
          mdm             = mdm + f"++++Missing data warning solar thermal area is zero on row {d['excel_row_num']}. \n"          
        if aream2 < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['Solar area m2'] } for solar thermal area on row {d['excel_row_num']}. \n"  
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Solar thermal area is not numeric on row {d['excel_row_num']}. \n"    

      coragk  = d['Corrected annual gen kWh']
      
      if isinstance(coragk,float) or isinstance(coragk,int):
        if coragk < 0 :
          nderrs = nderrs + 1
          vm              = vm + f"****Invalid value {d['Corrected annual gen kWh'] } for solar thermal Corrected annual gen kWh on row {d['excel_row_num']}. \n"  
      else:
          nderrs = nderrs + 1
          vm              = vm + f"****Solar thermal Corrected annual gen kWh is not numeric on row {d['excel_row_num']}. \n" 
          
    nve                 = nuprnerrs + nderrs
    nvw                 = ndwars
    mdw                 = mdwars
#    print('nve nvw mdw')
#    print(nve)
 #   print(nvw)
 #   print(mdw)
    if nve == 0 and nvw == 0:
      validation['ef'] = 0
      validation['em'] = '-----Solar thermal projects passed validation successfully.------'
    if nve == 0 and nvw > 0:
      validation['ef'] = 1
      validation['em'] = '++++Warning - there are validation warnings for solar thermal projects but updates will be performed'      
    if nve > 0:  
      validation['ef'] = 2
      validation['em'] = '****ERROR - upload has failed solar thermal validation. No updates will be performed.'
      
    validation['validated_df']          = df_original
  #  print('validation messages')
 #   print(vm)
    validation['validation_messages']   = vm
    validation['missing_data_messages'] = mdm
    validation['nvw']                   = nvw
    validation['nve']                   = nve
    validation['mdw']                   = mdw
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - Solar thermal projects validation terminated with exception: \n {msg}"
    validation['validated_df']          = ''
    validation['validation_messages']   = ''
    validation['missing_data_messages'] = ''
    validation['nvw']                   = 0
    validation['nve']                   = 0
    validation['mdw']                   = 0    
    return validation
  
  
def project_exists(project_type_id, uprn, entity_number):
  return

def calc_building_co2_emissions(dbconnect, entity_number, uprn):
  
  # Calculates baseline building CO2 emissions for electricity, gas, oil amd LPG (i.e. emissions for the data year)
  
  try:
    import Benchmarks as bm
    emissions = {'ef':0,'em':'','elec_co2':0,'gas_co2':0,'oil_co2':0,'lpg_co2':0, 'gas_wtt_scope_3':0,'elec_t_d_scope_3':0,\
                 'elec_wtt_t_d_scope_3':0,'elec_wtt_gen_scope_3':0,'oil_wtt_scope_3':0,'lpg_wtt_scope_3':0,'total_scope_1':0,\
                 'total_scope_2':0,'total_scope_3':0,'total_co2_tco2e':0,'annual_elec_cost':0,'annual_gas_cost':0,'annual_oil_cost':0,\
                 'annual_lpg_cost':0,'annual_energy_cost':0,'total_kwh':0, 'elec_kwh_m2':0, 'gas_kwh_m2':0, 'bmark_elec_kwh_m2b':0,\
                 'bmark_gas_kwh_m2b':0, 'elec_2b_saved_2_typical':0, 'gas_2b_saved_2_typical':0}
  
    with dbconnect.cursor() as cur:
    
  # Get raw building data for this building
  
      sql_raw_data_1 = "SELECT building_type, data_year, baseline_annual_elec_kwh, baseline_annual_gas_kwh, baseline_annual_oil_kwh, baseline_annual_lpg_kwh, gia_m2, baseline_elec_cost_per_kwh, baseline_gas_cost_per_kwh, baseline_oil_cost_per_kwh, baseline_lpg_cost_per_kwh FROM raw_estate_data WHERE entity_number = " 
      sql_raw_data_2 = f"{entity_number} AND uprn =  {uprn};"
      sqlr           = f"{sql_raw_data_1} {sql_raw_data_2} "
      #print('sqlr')
      #print(sqlr)
      cur.execute(sqlr)
      t_brl           = cur.fetchall()
      keys            = ("building_type", "data_year", "baseline_annual_elec_kwh", "baseline_annual_gas_kwh", "baseline_annual_oil_kwh", "baseline_annual_lpg_kwh", "gia_m2", "baseline_elec_cost_per_kwh", "baseline_gas_cost_per_kwh", "baseline_oil_cost_per_kwh", "baseline_lpg_cost_per_kwh")
      brl             = [dict(zip(keys, values)) for values in t_brl]
      br              = brl[0]
    
      building_type      = br['building_type']
      year_of_data       = br['data_year']
      annual_elec_kwh    = br['baseline_annual_elec_kwh']
      annual_gas_kwh     = br['baseline_annual_gas_kwh']
      annual_oil_kwh     = br['baseline_annual_oil_kwh']
      annual_lpg_kwh     = br['baseline_annual_lpg_kwh']
      gia_m2             = br['gia_m2']
      elec_cost_per_kwh  = br['baseline_elec_cost_per_kwh']
      gas_cost_per_kwh   = br['baseline_gas_cost_per_kwh']
      oil_cost_per_kwh   = br['baseline_oil_cost_per_kwh']
      lpg_cost_per_kwh   = br['baseline_lpg_cost_per_kwh']
      
      #print('xxxxx printing at BRL:')
      #print(f" annual_elec_kwh: {br['annual_elec_kwh']}  annual_gas_kwh: {br['annual_gas_kwh']}")
      #print(' ')
    
      if annual_elec_kwh == 0 and annual_gas_kwh == 0 and annual_oil_kwh == 0 and annual_lpg_kwh == 0:
        emissions['ef'] = 1
        emissions['em'] = "At least one (elec, gas, oil or lpg) annual kwh figure must be provided - none provided so all emission figures will be set to zero"
    
  # Get emission factors for year of building data
  
      sql_factors_1  = "SELECT elec_consumed, gas_consumed, oil_consumed, lpg_consumed,  electricity_t_d,  electricity_t_d_wtt,  electricity_gen_wtt, overall_elec_scope_3, gas_wtt, heating_oil_wtt, lpg_wtt  FROM emission_factors WHERE data_year ="
      sql_factors_2  = f"{year_of_data}"
      sqlf           = f"{sql_factors_1} {sql_factors_2}"
      #print('sqlf')
      #print(sqlf)
      cur.execute(sqlf)
      t_fal          = cur.fetchall()
      keys           = ("elec_consumed", "gas_consumed", "oil_consumed", "lpg_consumed",  "electricity_t_d",  "electricity_t_d_wtt",  "electricity_gen_wtt", "overall_elec_scope_3", "gas_wtt", "heating_oil_wtt", "lpg_wtt") 
      fal            = [dict(zip(keys, values)) for values in t_fal]
      fa             = fal[0]
      
      sql_bmark_1    = f"SELECT elec_good, gas_good FROM {bm.benchmark_table_name} WHERE building_type ="
      sql_bmark_2    = f"\'{building_type}\'"
      sqlb           = f"{sql_bmark_1} {sql_bmark_2}"
      
      #print('sqlb')
      #print(sqlb)
      cur.execute(sqlb)
      t_bml          = cur.fetchall()
      keys           = ("elec_good", "gas_good")
      bml            = [dict(zip(keys, values)) for values in t_bml]
      bm             = bml[0]
      
      elec_consumed        = fa['elec_consumed']
      gas_consumed         = fa['gas_consumed']
      oil_consumed         = fa['oil_consumed']
      lpg_consumed         = fa['lpg_consumed']
      electricity_t_d      = fa['electricity_t_d']
      electricity_t_d_wtt  = fa['electricity_t_d_wtt']
      electricity_gen_wtt  = fa['electricity_gen_wtt']
      overall_elec_scope_3 = fa['overall_elec_scope_3']
      gas_wtt              = fa['gas_wtt']
      heating_oil_wtt      = fa['heating_oil_wtt']
      lpg_wtt              = fa['lpg_wtt']
      elec_good            = bm['elec_good']
      gas_good             = bm['gas_good']
   
  # Calculate emissions
    
      elec_co2             = (elec_consumed  *  annual_elec_kwh)/1000
      gas_co2              = (gas_consumed  *  annual_gas_kwh)/1000
      oil_co2              = (oil_consumed  *  annual_oil_kwh)/1000
      lpg_co2              = (lpg_consumed  *  annual_lpg_kwh)/1000
      
      total_kwh            = annual_elec_kwh + annual_gas_kwh + annual_oil_kwh + annual_lpg_kwh

      if gia_m2 == 0:
         elec_kwh_m2       = 0
         gas_kwh_m2        = 0
      else:
         elec_kwh_m2       = annual_elec_kwh/gia_m2
         gas_kwh_m2        = annual_gas_kwh/gia_m2
  
  # Benchmarks
  
      bmark_elec_kwh_m2b   = elec_good
      bmark_gas_kwh_m2b    = gas_good
      
      elec_2b_saved_2_typical = (elec_kwh_m2 - bmark_elec_kwh_m2b) * gia_m2
      gas_2b_saved_2_typical  = (gas_kwh_m2 - bmark_gas_kwh_m2b) * gia_m2

      gas_wtt_scope_3      = (annual_gas_kwh *  gas_wtt)/1000
      elec_t_d_scope_3     = (annual_elec_kwh *  electricity_t_d)/1000
      elec_wtt_t_d_scope_3 = (annual_elec_kwh * electricity_t_d_wtt)/1000
      elec_wtt_gen_scope_3 = (annual_elec_kwh * electricity_gen_wtt)/1000
      oil_wtt_scope_3      = (annual_oil_kwh * heating_oil_wtt)/1000
      lpg_wtt_scope_3      = (annual_lpg_kwh * lpg_wtt)/1000
    
      total_scope_1        = gas_co2 + oil_co2 + lpg_co2
      total_scope_2        = elec_co2
      total_scope_3        = gas_wtt_scope_3 + elec_t_d_scope_3 + elec_wtt_t_d_scope_3 + elec_wtt_gen_scope_3 + oil_wtt_scope_3 + lpg_wtt_scope_3
    
      total_co2_tco2e      = total_scope_3 + elec_co2 + gas_co2 + oil_co2 + lpg_co2
    
      annual_elec_cost     = annual_elec_kwh * elec_cost_per_kwh
      annual_gas_cost      = annual_gas_kwh * gas_cost_per_kwh
      annual_oil_cost      = annual_oil_kwh * oil_cost_per_kwh
      annual_lpg_cost      = annual_lpg_kwh * lpg_cost_per_kwh
      annual_energy_cost   = annual_elec_cost + annual_gas_cost + annual_oil_cost + annual_lpg_cost
    
  # Build output dictionary of emission results for this building
    
      emissions['elec_co2']                = elec_co2
      emissions['gas_co2']                 = gas_co2
      emissions['oil_co2']                 = oil_co2
      emissions['lpg_co2']                 = lpg_co2
      emissions['gas_wtt_scope_3']         = gas_wtt_scope_3
      emissions['elec_t_d_scope_3']        = elec_t_d_scope_3
      emissions['elec_wtt_t_d_scope_3']    = elec_wtt_t_d_scope_3
      emissions['elec_wtt_gen_scope_3']    = elec_wtt_gen_scope_3
      emissions['oil_wtt_scope_3']         = oil_wtt_scope_3
      emissions['lpg_wtt_scope_3']         = lpg_wtt_scope_3
      emissions['total_scope_1']           = total_scope_1
      emissions['total_scope_2']           = total_scope_2
      emissions['total_scope_3']           = total_scope_3
      emissions['total_co2_tco2e']         = total_co2_tco2e
      emissions['annual_elec_cost']        = annual_elec_cost
      emissions['annual_gas_cost']         = annual_gas_cost
      emissions['annual_oil_cost']         = annual_oil_cost
      emissions['annual_lpg_cost']         = annual_lpg_cost
      emissions['annual_energy_cost']      = annual_energy_cost
      emissions['total_kwh']               = total_kwh
      emissions['elec_kwh_m2']             = elec_kwh_m2
      emissions['gas_kwh_m2']              = gas_kwh_m2
      emissions['bmark_elec_kwh_m2b']      = bmark_elec_kwh_m2b
      emissions['bmark_gas_kwh_m2b']       = bmark_gas_kwh_m2b
      emissions['elec_2b_saved_2_typical'] = elec_2b_saved_2_typical
      emissions['gas_2b_saved_2_typical']  = gas_2b_saved_2_typical      
      return emissions 
    
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    emissions['ef'] = 2
    emissions['em'] = f"**** A fatal error has occured in emission calculations - please report to your support team\n{msg}"
  return emissions

def calc_controlled_estate_summary(dbconnect, entity_number):
  
  # Calculates results summarising all controlled buildings within an estate and writes these to the controlled_estate_summary table
  #import kv_calcs as kc
  print('In calc_controlled_estate_summary')

  try:
    
    summary = {'ef':0,'em':'all ok','entity_number':0,'total_elec_kwh':0,'total_gas_kwh':0,'total_oil_kwh':0,'total_lpg_kwh':0,'total_solar_pv_kwh':0,'total_solar_thermal_kwh':0,'total_energy_kwh':0,\
               'total_area_m2':0,'total_estates_portfolio':0,'elec_cost_kwh_gbp':0,'gas_cost_kwh_gbp':0,'average_dec_score':0,'average_dec_rating':0,\
               'co2_scope_1':0,'co2_scope_2':0,'co2_scope_3':0,'co2_total':0,'stock_ave_elec_use_kwh_m2':0,'stock_ave_gas_use_kwh_m2':0,'stock_ave_total_use_kwh_m2':0,\
               'use_good_bm_elec_use_kwh_m2':0,'use_good_bm_gas_use_kwh_m2':0,'use_good_bm_total_use_kwh_m2':0,'stock_ave_vs_good_elec_kwh_m2':0,'stock_ave_vs_good_gas_kwh_m2':0,\
               'stock_ave_vs_good_total_kwh_m2':0,'elec_2b_saved_2get_good':0,'gas_2b_saved_2get_good':0,'pc_tot_energy_elec':0,'pc_tot_energy_gas':0,'pc_tot_energy_oil':0,'pc_tot_energy_lpg':0,\
               'pc_tot_energy_zero_carbon_elec':0,'pc_tot_energy_zero_carbon_heat':0}
    
    with dbconnect.cursor() as cur:
    
        # Get results for the controlled estate for this entity from the results_raw_data table

        csql1  = "SELECT * FROM results_raw_estate_data WHERE entity_number ="
        csql2  = f"{entity_number} AND under_control = \'YES\';"
        csql   = f"{csql1} {csql2}"
        rasql1 = "SELECT uprn,gia_m2,dec_score,baseline_annual_elec_kwh,baseline_annual_gas_kwh,baseline_annual_oil_kwh,baseline_annual_lpg_kwh,baseline_elec_cost_per_kwh,baseline_gas_cost_per_kwh, exist_solar_pv_annual_kwh, exist_solar_thermal_annual_kwh FROM raw_estate_data WHERE entity_number ="
        rasql  = f"{rasql1} {csql2}"
        
        cur.execute(csql)
        keys     = [column[0] for column in cur.description]
        t_cesres = cur.fetchall()
        cesres   = [dict(zip(keys, values)) for values in t_cesres]
      
        cur.execute(rasql)
        keys     = ("uprn","gia_m2","dec_score","baseline_annual_elec_kwh","baseline_annual_gas_kwh","baseline_annual_oil_kwh","baseline_annual_lpg_kwh","baseline_elec_cost_per_kwh","baseline_gas_cost_per_kwh","exist_solar_pv_annual_kwh","exist_solar_thermal_annual_kwh")
        t_cesraw = cur.fetchall()
        cesraw   = [dict(zip(keys, values)) for values in t_cesraw]
      
        if len(cesres) != len(cesraw):
          summary['ef'] = 2
          summary['em'] = "**** Number of controlled estate records in raw data table does not match that in the results data table - please report to your support team"
          return summary                
        if len(cesraw) == 0:
          summary['ef'] = 2
          summary['em'] = f"**** No buildings have been found in the raw estate data table - please report to your support team \n Entity number : {entity_number}\n Sql : {rasql}"
          return summary    
          
        t_total_elec_kwh                 = 0
        t_total_gas_kwh                  = 0
        t_total_oil_kwh                  = 0
        t_total_lpg_kwh                  = 0
        t_total_solar_pv_kwh             = 0
        t_total_solar_thermal_kwh        = 0
        t_total_energy_kwh               = 0
        t_total_area_m2                  = 0
        t_total_estates_portfolio        = len(cesraw)
        t_elec_cost_kwh_gbp              = 0
        t_gas_cost_kwh_gbp               = 0
        t_average_dec_score              = 0
        t_average_dec_rating             = 0
        t_co2_scope_1                    = 0
        t_co2_scope_2                    = 0
        t_co2_scope_3                    = 0
        t_co2_total                      = 0
        t_stock_ave_elec_use_kwh_m2      = 0
        t_stock_ave_gas_use_kwh_m2       = 0
        t_stock_ave_total_use_kwh_m2     = 0
        t_use_good_bm_elec_use_kwh_m2    = 0
        t_use_good_bm_gas_use_kwh_m2     = 0
        t_use_good_bm_total_use_kwh_m2   = 0
        t_stock_ave_vs_good_elec_kwh_m2  = 0
        t_stock_ave_vs_good_gas_kwh_m2   = 0
        t_stock_ave_vs_good_total_kwh_m2 = 0
        t_total_zero_carbon_elec         = 0
        t_total_zero_carbon_heat         = 0
        
        t_elec_2b_saved_2get_good        = 0
        t_gas_2b_saved_2get_good         = 0
        
        t_pc_tot_energy_elec             = 0
        t_pc_tot_energy_gas              = 0
        t_pc_tot_energy_oil              = 0
        t_pc_tot_energy_lpg              = 0
        t_pc_tot_energy_zero_carbon_elec = 0
        t_pc_tot_energy_zero_carbon_heat = 0
        
        tecpkwh                          = 0
        tgcpkwh                          = 0
        tdecscore                        = 0
        bidgiadict                       = {}
        biddecdict                       = {}
        sum_dec_x_gia                    = 0
        cd1                              = 0
        
        for raw in cesraw:

          t_total_elec_kwh                 = t_total_elec_kwh     + raw['baseline_annual_elec_kwh']
          t_total_gas_kwh                  = t_total_gas_kwh      + raw['baseline_annual_gas_kwh']
          t_total_oil_kwh                  = t_total_oil_kwh      + raw['baseline_annual_oil_kwh']
          t_total_lpg_kwh                  = t_total_lpg_kwh      + raw['baseline_annual_lpg_kwh']
          t_total_zero_carbon_elec         = t_total_zero_carbon_elec + raw['exist_solar_pv_annual_kwh']
          t_total_zero_carbon_heat         = t_total_zero_carbon_heat + raw['exist_solar_thermal_annual_kwh']
          
          tecpkwh                          = tecpkwh              + raw['baseline_elec_cost_per_kwh'] 
          tgcpkwh                          = tgcpkwh              + raw['baseline_gas_cost_per_kwh']
          tdecscore                        = tdecscore            + raw['dec_score']
          t_total_area_m2                  = t_total_area_m2      + raw['gia_m2']

          update_with_gia                  = {raw['uprn'] : raw['gia_m2']}
          bidgiadict.update(update_with_gia)
          update_with_dec                  = {raw['uprn'] : raw['dec_score']}
          biddecdict.update(update_with_dec)
          
        print('This is total_energy_kwh')
        print(t_total_elec_kwh + t_total_gas_kwh + t_total_oil_kwh + t_total_lpg_kwh)
        
        t_elec_cost_kwh_gbp              = tecpkwh/t_total_estates_portfolio
        t_gas_cost_kwh_gbp               = tgcpkwh/t_total_estates_portfolio
        t_stock_ave_elec_use_kwh_m2      = t_total_elec_kwh/t_total_area_m2
        t_stock_ave_gas_use_kwh_m2       = t_total_gas_kwh/t_total_area_m2
        t_total_energy_kwh               = t_total_elec_kwh + t_total_gas_kwh + t_total_oil_kwh + t_total_lpg_kwh
        t_stock_ave_total_use_kwh_m2     = t_total_energy_kwh/t_total_area_m2        
        
        #t_average_dec_score              = tdecscore/t_total_estates_portfolio
  
        for res in cesres:
          
#          t_total_energy_kwh               = t_total_energy_kwh    + res['total_kwh']

          t_co2_scope_1                    = t_co2_scope_1         + res['total_scope_1']
          t_co2_scope_2                    = t_co2_scope_2         + res['total_scope_2']
          t_co2_scope_3                    = t_co2_scope_3         + res['total_scope_3']
          t_co2_total                      = t_co2_total           + res['total_co2_tco2e']
          
          bid                              = res['uprn']
          gia_4_building                   = bidgiadict[bid]
          dec_4_building                   = biddecdict[bid]
          
          if dec_4_building > 0:
            cd1                              = cd1 + gia_4_building
            
          dec_x_gia                        = dec_4_building * gia_4_building
          sum_dec_x_gia                    = sum_dec_x_gia + dec_x_gia
          
          bm_elec_kwh_m2b                  = res['bmark_elec_kwh_m2b']
          bm_gas_kwh_m2b                   = res['bmark_gas_kwh_m2b']

          t_use_good_bm_elec_use_kwh_m2    = t_use_good_bm_elec_use_kwh_m2 + (bm_elec_kwh_m2b * (gia_4_building/t_total_area_m2))
          t_use_good_bm_gas_use_kwh_m2     = t_use_good_bm_gas_use_kwh_m2 + (bm_gas_kwh_m2b * (gia_4_building/t_total_area_m2))
          t_elec_2b_saved_2get_good        = t_elec_2b_saved_2get_good + res['elec_2b_saved_2_typical']
          t_gas_2b_saved_2get_good         = t_gas_2b_saved_2get_good  + res['gas_2b_saved_2_typical']           
        
        t_use_good_bm_total_use_kwh_m2   = t_use_good_bm_elec_use_kwh_m2 + t_use_good_bm_gas_use_kwh_m2
        t_stock_ave_vs_good_elec_kwh_m2  = t_stock_ave_elec_use_kwh_m2 - t_use_good_bm_elec_use_kwh_m2
        t_stock_ave_vs_good_gas_kwh_m2   = t_stock_ave_gas_use_kwh_m2 - t_use_good_bm_gas_use_kwh_m2
        t_stock_ave_vs_good_total_kwh_m2 = t_stock_ave_total_use_kwh_m2 - t_use_good_bm_total_use_kwh_m2
        if cd1 > 0:  
          t_average_dec_score              = sum_dec_x_gia/cd1
        else:
          t_average_dec_score              = 0
          
        t_average_dec_rating             = kc.get_dec_letter(t_average_dec_score)
        
        # Calculate the percentages of total energy
      
        total_zero_and_nonzero_energy            = t_total_zero_carbon_elec + t_total_zero_carbon_heat + t_total_energy_kwh        
        t_pc_tot_energy_elec                     = (t_total_elec_kwh/total_zero_and_nonzero_energy) * 100
        t_pc_tot_energy_gas                      = (t_total_gas_kwh/total_zero_and_nonzero_energy) * 100                    
        t_pc_tot_energy_oil                      = (t_total_oil_kwh/total_zero_and_nonzero_energy) * 100
        t_pc_tot_energy_lpg                      = (t_total_lpg_kwh/total_zero_and_nonzero_energy) * 100

        t_pc_tot_energy_zero_carbon_elec         = (t_total_zero_carbon_elec /total_zero_and_nonzero_energy ) * 100
        t_pc_tot_energy_zero_carbon_heat         = (t_total_zero_carbon_heat /total_zero_and_nonzero_energy ) * 100
        
        summary['total_elec_kwh']                = t_total_elec_kwh
        summary['total_gas_kwh']                 = t_total_gas_kwh
        summary['total_oil_kwh']                 = t_total_oil_kwh
        summary['total_lpg_kwh']                 = t_total_lpg_kwh
        summary['total_solar_pv_kwh']            = t_total_zero_carbon_elec
        summary['total_solar_thermal_kwh']       = t_total_zero_carbon_heat        
        summary['total_energy_kwh']              = t_total_energy_kwh
        summary['total_area_m2']                 = t_total_area_m2
        summary['total_estates_portfolio']       = t_total_estates_portfolio
        summary['elec_cost_kwh_gbp']             = t_elec_cost_kwh_gbp
        summary['gas_cost_kwh_gbp']              = t_gas_cost_kwh_gbp
        summary['average_dec_score']             = t_average_dec_score
        summary['average_dec_rating']            = t_average_dec_rating
        summary['co2_scope_1']                   = t_co2_scope_1
        summary['co2_scope_2']                   = t_co2_scope_2
        summary['co2_scope_3']                   = t_co2_scope_3
        summary['co2_total']                     = t_co2_total
        summary['stock_ave_elec_use_kwh_m2']     = t_stock_ave_elec_use_kwh_m2
        summary['stock_ave_gas_use_kwh_m2']      = t_stock_ave_gas_use_kwh_m2
        summary['stock_ave_total_use_kwh_m2']    = t_stock_ave_total_use_kwh_m2
        summary['use_good_bm_elec_use_kwh_m2']   = t_use_good_bm_elec_use_kwh_m2
        summary['use_good_bm_gas_use_kwh_m2']    = t_use_good_bm_gas_use_kwh_m2
        summary['use_good_bm_total_use_kwh_m2']  = t_use_good_bm_total_use_kwh_m2
        summary['stock_ave_vs_good_elec_kwh_m2'] = t_stock_ave_vs_good_elec_kwh_m2
        summary['stock_ave_vs_good_gas_kwh_m2']  = t_stock_ave_vs_good_gas_kwh_m2
        summary['stock_ave_vs_good_total_kwh_m2']= t_stock_ave_vs_good_total_kwh_m2
        summary['elec_2b_saved_2get_good']       = t_elec_2b_saved_2get_good
        summary['gas_2b_saved_2get_good']        = t_gas_2b_saved_2get_good
        summary['pc_tot_energy_elec']            = t_pc_tot_energy_elec
        summary['pc_tot_energy_gas']             = t_pc_tot_energy_gas
        summary['pc_tot_energy_oil']             = t_pc_tot_energy_oil
        summary['pc_tot_energy_lpg']             = t_pc_tot_energy_lpg
        summary['pc_tot_energy_zero_carbon_elec']= t_pc_tot_energy_zero_carbon_elec
        summary['pc_tot_energy_zero_carbon_heat']= t_pc_tot_energy_zero_carbon_heat

        return summary
    
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    summary['ef'] = 2
    summary['em'] = f"**** A fatal error has occured while calulating summary for the controlled estate - please report to your support team: -\n{msg}"
  return summary    

def initialise_building_projects(conn, d, entity_number, name_id, start_date):
  # conn is the database connection object. d is a row (building) from the projects upload dataframe for entity 
  # with entity number entity_number. name_id is a list of dicts containing keys as names of project type and values as the associated id.
  # start_date is the programme start date specified for this entity used to calculate the assessed_delivery_date from the 'assessed' field.
  # The projects that have been assessed for the building are identified and for each project an entry is created in the projects table.
  # IMPORTANT NOTE: Projects for this building already in the database (having been processed by a previous upload) that are also on this new  
  # upload record will have their assessed status updated but will otherwise be un-touched. Projects for this building already in the database 
  # which are NOT on this new upload record will be deleted. Projects for this building which are not already in the database will be added to the projects table.
  
  print('In initialise_building_projects ')
  
  # Returns ret_mess where: -
  
  # ef          = error flag, 0 = success, 1 = warnings, 2 = error
  # em          = error message if ef not equal to zero
  # npdeleted   = number of projects already in database for this building that were deleted
  # npupdated   = number of projects already in database for this building that have been updated
  # npcreated   = number of new projects for this building that have been created
    
  ret_mess = {'ef':0, 'em':'', 'npdeleted':0, 'npupdated':0, 'npcreated':0} 
  
  sql2 = f"INSERT INTO projects (project_id,project_status) VALUES (LAST_INSERT_ID(),'Planned');"
  import datetime as dt
  
  try:
  # Set default for delivery_date
    ddd        = dt.datetime(1900,1,1)
  # Set counters
    npdeleted = 0
    npcreated = 0
    npupdated = 0
    
  # Remove unwanted columns
    uprn  =  d['Building ID']
    del d['excel_row_num']
    del d['Building ID']
    del d['Building name']
    del d['Building type']
    
  # For this building delete project records from the projects table if the project is not in this new upload record.
  
    with conn.cursor() as cursor:
      
      # Get list of projects already in the projects table for this building uprn and entity_number
      
      try:
        sqlr = f"SELECT project_id,project_type_id FROM projects WHERE entity_number = {entity_number} AND uprn = {uprn}; "
        cursor.execute(sqlr)
        t_projects_in_db   = cursor.fetchall() # A list of tuples project_id, project_type_id 
        keys               = ("project_id","project_type_id")
        projects_in_db     = [dict(zip(keys, values)) for values in t_projects_in_db] # Convert tuples to dicts
        
      except (pyodbc.Error)  as e:
        exnum            = 1
        ret_mess['ef']   = 2
        ret_mess['em']   = f"Exception number {exnum} - **** error in initialise_building_projects retrieving projects from projects table \n {e}"
        return ret_mess
       
      for key in d: #  Iterate across projects in input record for this building

        project_type       = key
        assessed           = d[key]
        # Find the project type by name in name_id and extract the associated project_type_id
        res                = next(item for item in name_id if item["name"] == project_type)
        project_type_id    = res['project_type_id'] 
        
        # Does this project_type_id already have a project in the projects table?
        
        type_in_db         = False
        
        for n in projects_in_db:
          if n['project_type_id'] == project_type_id: # Found there is a project of this type in the projects table so set get the project_id and set the flag
            project_id   = n['project_id']
            type_in_db   = True
            break
            
        # If this project type for this building has a null in the assessed field in the upload but it is already in the projects table then delete it.
        
        if assessed == '': # Project type is NOT assessed on input - if there is a project for this project_type_id already in the projects table then delete it
                           # otherwise no action taken.
          if type_in_db:
            sqld         = f"DELETE FROM projects WHERE project_id = {project_id};"
            try:
              cursor.execute(sqld)
              conn.commit()
              npdeleted    = npdeleted + 1
            except (pyodbc.Error) as e:
                exnum             = 2
                conn.rollback()
                ret_mess['ef']        = 2
                ret_mess['em']        = up_log + f"Exception number {exnum} - DELETE on projects table - ****ERROR on project id {project_id}. DB returned: - \n{e}\n"
                return ret_mess
        else: # Project type is assessed on input. If a project of this type already exists in the projects table then update its 'assessed' field and 'assessed_delivery_date' fields.
              # If there is no project of this type in the projects table then insert one.
          
          # Calculate the assessed_delivery_date based on start_date and assessed value.
          asdd = calculate_assessed_delivery_date(assessed, start_date)
          # Set the default delivery date
          
          if type_in_db:

            sqlup         = f"UPDATE projects SET assessed = \'{assessed}\', assessed_delivery_date = \'{asdd}\' WHERE project_id = {project_id};"

            try:
              cursor.execute(sqlup)
              conn.commit()
              npupdated           = npupdated + 1
                
            except (pyodbc.Error) as e:
                exnum             = 3
                conn.rollback()
                ret_mess['ef']        = 2
                ret_mess['em']        = f"Exception number {exnum} - UPDATE projects table - ****ERROR on project id {project_id}. DB returned: - \n{e}\n"
                return ret_mess
#                   summary           = summary + f"****ERROR - a database error has occured during UPDATE raw data. See upload log for details\n"

          else:  # There is not already an entry in the projects table so create one for this project. At the same time if the project_type_id is 20 (Solar PV)
                 # initialise the projects.solar_pv_area_m2 field to the value in the raw_estate_data.roof_space_m2 field.
            #print('In  ELSE type_in_db')
            sqlr    = f"SELECT roof_space_m2 FROM raw_estate_data WHERE entity_number = {entity_number} AND uprn = {uprn}; "
            cursor.execute(sqlr)
            t_oput    = cursor.fetchall()
            keys      = ("roof_space_m2","dummy_key")
            oput      = [dict(zip(keys, values)) for values in t_oput]
            dic       = oput[0]
            solpvm2   = dic['roof_space_m2']
            
            sqli = f"INSERT INTO projects (uprn, entity_number, project_type_id, assessed, assessed_delivery_date, delivery_date_mode, solar_pv_area_m2) VALUES ({uprn},{entity_number},{project_type_id},\'{assessed}\',\'{asdd}\',\'{ddd}\', {solpvm2});"
            #print('sqli')
            #print(sqli)
            try:
              cursor.execute(sqli)
              conn.commit()
              npcreated = npcreated + 1
            except (pyodbc.Error)  as e:
              ret_mess['ef']   = 4
              exnum            = 6
              ret_mess['em']   = f"Exception number {exnum} - error in initialise_building_projects inserting into projects table for uprn {uprn} entity_number {entity_number} project_type_id {project_type_id} \n {e}"
              conn.rollback()
              return ret_mess

  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef']    = 1
    ret_mess['em']    = msg
    return ret_mess
    
  ret_mess['npupdated'] = npupdated
  ret_mess['npdeleted'] = npdeleted
  ret_mess['npcreated'] = npcreated
  return ret_mess

def calculate_assessed_delivery_date(assessed, start_date):
# Assessed value    Delivered within     Value used in assessed_delivery_date
# --------------    ----------------     ------------------------------------
# FIRM              within 1 year         6 months (183 days) from start date
# LIKELY            1 - 2 years          18 months (548 days) from start date
# POSSIBLE          2 - 3 years          30 months (913 days) from start date
# POTENTIAL         3 - 5 years          48 months (1460 days) from start date
# IN PLACE          N/A                  Null - status set to 'Completed'
# ASSESSED/NV       N/A                  Null - status set to 'Cancelled'
# FTHR IMPV*        Unknown but assume   48 (1460 days) months from start date
#                   same as POTENTIAL
# * - to be clarified by JH

  import datetime as dt
  
  sd        = dt.datetime(1900,1,1)
#  bdate     = dt.datetime.date(sd).strftime("%d/%m/%Y")
  bdate     = dt.datetime.date(sd).strftime("%d-%b-%Y")
  if isinstance(start_date, str):
    start_date = dt.datetime.strptime(start_date, "%d-%b-%Y")
    
  deltas = {'FIRM':183, 'LIKELY':548, 'POSSIBLE':913, 'POTENTIAL':1460, 'FTHR IMPV':1460}
  vlist = ['FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL', 'FTHR IMPV']
  #print('++++++In calculate_assessed_delivery_date')
  #print(f"start_date : {start_date}")
  #print(f"Type of start date - {type(start_date)}")
  for i in vlist:
    if(i == assessed) :
      ddays     = deltas[assessed]
      assdate   = start_date + timedelta(days = ddays)
      break
    else:
      assdate   = bdate

  return assdate
def gaselec_savings_check(conn,entity_number):
  # conn          - database connection object
  # entity_number - unique entity identification number
  #
  # Checks the total gas and electricity savings, summed across all projects within a building, do not exceed 100%. 
  # Buildings where gas savings or electricity savings exceed 100% are logged in the return messages and have their g_saving_flag and e_saving_flag in the raw_estate_data table, 
  # respectively, set to 1.
  # Buildings where gas savings or electricity savings do NOT exceed 100% have their g_saving_flag and e_saving_flag in the raw_estate_data table, respectively, set to 0.
  # Updated 06/02/2024 to add building name to uprn on return
  
  try:
    checks       = {'ef':0, 'em':'', 'gas_uprn_list':[], 'gas_build_list':[], 'elec_uprn_list':[], 'elec_build_list':[], 'ngasfails':0, 'nelecfails':0} 
    e_uprn_list  = []
    g_uprn_list  = []
    e_build_list = []
    g_build_list = []
    
    # Get all uprns for this entity
    with conn.cursor() as cursor:
      sqlu            = f"SELECT uprn, building_name FROM raw_estate_data WHERE entity_number = {entity_number};"
  #    print('In checks top')
      ngasfails  = 0
      nelecfails = 0
      
      cursor.execute(sqlu)
      t_ulist   = cursor.fetchall()
      keys      = ("uprn","building_name","dummy_key") 
      ulist     = [dict(zip(keys, values)) for values in t_ulist]
      
#      print('About to loop through ulist')
      # Loop through the uprns and retrieve utility and saving percent for all projects for each uprn (building)
      for n in ulist:
        uprn         = n['uprn']
        build_name   = n['building_name']
        setgflag     = f"UPDATE raw_estate_data SET g_saving_flag = 1 WHERE ((entity_number = {entity_number} and uprn = {uprn}));"
        seteflag     = f"UPDATE raw_estate_data SET e_saving_flag = 1 WHERE ((entity_number = {entity_number} and uprn = {uprn}));"
        unsetgflag   = f"UPDATE raw_estate_data SET g_saving_flag = 0 WHERE ((entity_number = {entity_number} and uprn = {uprn}));"
        unseteflag   = f"UPDATE raw_estate_data SET e_saving_flag = 0 WHERE ((entity_number = {entity_number} and uprn = {uprn}));"         
        e_saving     = 0
        g_saving     = 0
                
   #     sqlp           = f"SELECT project_type_id, utility, saving_percent FROM projects WHERE (uprn = {uprn}) AND (entity_number = {entity_number}) AND (project_type_id <> 20) AND (project_type_id <> 21);"
        sqlp           = f"SELECT project_type_id, utility, saving_percent FROM projects WHERE (uprn = {uprn}) AND (entity_number = {entity_number}) AND (project_type_id NOT IN (16,20,21,25));"
   #     print('sqlp--------')
   #     print(sqlp)
        cursor.execute(sqlp)
        t_plist = cursor.fetchall()
        keys    = ("project_type_id","utility","saving_percent")
        plist   = [dict(zip(keys, values)) for values in t_plist]
    #    print(f" uprn = {uprn}\n")
    #    print(f" plist = {plist}\n")
        
        # Loop through projects for each uprn calculating total gas and electricity savings for each building. If building savings are > 100% add
        # uprn to relevant list of problem buildings and set appropriate flag (g_saving_flag and e_saving_flag) to 1 for building in raw_estate_data table.
        # Where the building totals are <= 100% set appropriate flag (g_saving_flag and e_saving_flag) to 0 for building in raw_estate_data table.

        if len(plist) > 0:
          for p in plist:
            utility = p['utility']
            saving  = p['saving_percent']

            if utility == 'GAS':
              g_saving = g_saving + saving
            if utility == 'ELEC':
              e_saving = e_saving + saving

          if g_saving > 1.0:
            g_uprn_list.append(uprn)
            g_build_list.append(build_name)
            ngasfails  = ngasfails + 1
            cursor.execute(setgflag)
            conn.commit()
          else:
            cursor.execute(unsetgflag)
            conn.commit()            
            
          if e_saving > 1.0:
            e_uprn_list.append(uprn)
            e_build_list.append(build_name)
            nelecfails  = nelecfails + 1
            cursor.execute(seteflag)
            conn.commit()
          else:
            cursor.execute(unseteflag)
            conn.commit() 
            
      checks       = {'ef':0, 'em':'', 'gas_uprn_list':g_uprn_list, 'gas_build_list':g_build_list, 'elec_uprn_list':e_uprn_list, 'elec_build_list':e_build_list, 'ngasfails':ngasfails, 'nelecfails':nelecfails}       
      return checks
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__)) 
    checks['ef']  = 2
    checks['em']  = msg
  return checks

def update_non_solar_project_details(conn, dfns, entity_number, output_pt):

# At this point we have a dataframe with at worst warnings and no errors so now we can UPDATE non solar project records in the database.
  num_rows_read          = dfns.shape[0]
  ret_mess               = {'ef':0, 'em':'', 'up_log': '', 'summary': ''}
  summary                = ''
  up_log                 = ''
  try:  
    
    with conn.cursor() as cursor:

    # Iterate down the down the dataframe (by project)
      rec_num    = 0
      nupdates   = 0
      nupdrawwa  = 0
      nupdrawer  = 0
      for d4 in dfns.to_dict(orient="records"): 
        rec_num = rec_num + 1

      # If delivery date is '0000-00-00'(which it can be when a blank cell is read in from Excel) convert it to 01/01/1900 before updating database'
      # to prevent giving Excel a date it doesn't understand when producing the project details data collection form

        if d4['Delivery date'] == '0000-00-00':
          sd                  = dt.datetime(1900,1,1)
          bdate               = dt.datetime.date(sd)
          d4['Delivery date'] = bdate
        
        # Find the project_type_id so we can find the project_id and use that to update the project and results_projects tables

        ptype          = d4['Project type']

        for n in output_pt:
          if n['name'] == ptype:
            ptypeid    = n['project_type_id']
            break  
        uprn = d4['Building ID']

        fisql = f"SELECT project_id FROM projects WHERE ((entity_number = {entity_number}) AND (uprn = {uprn}) AND (project_type_id = {ptypeid}));" 
        cursor.execute(fisql)
        t_lodics          = cursor.fetchall()
        keys              = ("project_id","dummy_key")
        lodics            = [dict(zip(keys, values)) for values in t_lodics]
        dic               = lodics[0]
        pr_id             = dic['project_id']

        upsql = f"UPDATE projects \
                SET assessed                        = \'{d4['Assessed']}\',\
                project_status                      = \'{d4['Status']}\',\
                utility                             = \'{d4['Utility']}\',\
                salix_pf                            = {d4['Lifetime (yrs)']},\
                saving_percent                      = {d4['Saving %']},\
                hp_scop                             = {d4['Heat pump scop']},\
                hp_elec_add_kwh_pa                  = {d4['Heat pump elec add kWh pa']},\
                cost_capex_mode                     = {d4['CAPEX']},\
                delivery_date_mode                  = \'{d4['Delivery date']}\'\
                WHERE project_id = {pr_id};"            

        try:
          cursor.execute(upsql)
          conn.commit()
          nupdates         = nupdates + 1
            
        except (pyodbc.Error) as e:

          nupdrawer         = nupdrawer + 1
          conn.rollback()
          up_log            = up_log + f"******* Database Exception updating non solar projects - ****ERROR on record number {rec_num}. DB returned: - \n{e}\n"
                          
      fs1                   = ''  
      fs                    =      f"Number of non solar project records read from upload file : {num_rows_read} \n"
      fs                    = fs + f"Number of non solar projects that were updated            : {nupdates} \n"
      if nupdrawwa > 0:
        fs                  = fs + f"Number of database update warnings                       : {nupdrawwa}\n"
        ret_mess['ef']      = 1
      
      if nupdrawer > 0:  
       fs                  = fs + f"Number of database update errors                         : {nupdrawer}\n"
       ret_mess['ef']      = 2     
      
      if nupdrawwa > 0 or nupdrawer > 0:
        fs1                 = fs + f"Please see the upload log for details of warnings and errors. \n"
    
      summary               = summary + fs1
      up_log                = up_log  + fs
      ret_mess['summary']   = summary
      ret_mess['up_log']    = up_log
    
      return ret_mess

  except Exception as e: 
    summary = '******An exception has occurred in update_non_solar_project_details - please see upload log for details'

    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))

    up_log  = f"Exception exit 1 \n {msg}"
    ret_mess['summary'] = summary
    ret_mess['up_log']  = up_log
    ret_mess['ef']      = 2
    
    return ret_mess
  
  
def update_solar_pv_project_details(conn, dfsp, entity_number, output_pt):

# At this point we have a dataframe with at worst warnings and no errors so now we can UPDATE solar pv project records in the database.
  num_rows_read          = dfsp.shape[0]
  ret_mess               = {'ef':0, 'em':'', 'up_log': '', 'summary': ''}
  summary                = ''
  up_log                 = ''
  
  try:  
    print('In update_solar_pv_project_details at top =====')
    print(dfsp.to_string())
    with conn.cursor() as cursor:
      
    # Get the list of project type ids and names
    
      sqlt              = f"SELECT project_type_id, name FROM project_types;"
      cursor.execute(sqlt)
      t_output_pt       = cursor.fetchall()
      keys              = ("project_type_id","name") 
      output_pt         = [dict(zip(keys, values)) for values in t_output_pt]
      
    # Get a list of uprn, gia_m2, baseline_annual_elec_kwh and listed for the estate
    
      sqls              = f"SELECT uprn, gia_m2, roof_space_m2, baseline_annual_elec_kwh, listed FROM raw_estate_data WHERE entity_number = {entity_number};"
      cursor.execute(sqls)
      t_output_ugl      = cursor.fetchall()
      keys              = ("uprn","gia_m2","roof_space_m2", "baseline_annual_elec_kwh","listed") 
      output_ugl        = [dict(zip(keys, values)) for values in t_output_ugl]
      
    # Iterate down the down the dataframe (by project)
      rec_num    = 0
      nupdates   = 0
      nupdrawwa  = 0
      nupdrawer  = 0
      
      for d4 in dfsp.to_dict(orient="records"): 
        rec_num = rec_num + 1

        # If delivery date is '0000-00-00'(which it can be when a blank cell is read in from Excel) convert it to 01/01/1900 before updating database'
        # to prevent giving Excel a date it doesn't understand when producing the project details data collection form
        
        if d4['Delivery date'] == '0000-00-00':
          sd                  = dt.datetime(1900,1,1)
          bdate               = dt.datetime.date(sd)

          d4['Delivery date'] = bdate
        # Find the project_type_id so we can find the project_id and use that to update the project and results_projects tables

        ptype          = d4['Project type']

        for n in output_pt:
          if n['name'] == ptype:
            ptypeid    = n['project_type_id']
            break  
        uprn = d4['Building ID']
        
        # Get the gia_m2, roof_space_m2, listed and baseline_annual_elec_kwh for the building
        
        founduprn = False
        for g in output_ugl:
          if g['uprn'] == uprn:
            gia_m2          = g['gia_m2']
            roof_space_m2   = g['roof_space_m2']
            listed          = g['listed']
            baseline_annual_elec_kwh = g['baseline_annual_elec_kwh']
            founduprn = True
            break
        
        if not founduprn:
          ret_mess['ef']  = 2
          ret_mess['em']  = f"****Error in update_solar_pv_project_details unable to find uprn - {uprn}\n"
          return ret_mess
        
        # Calculate the percentage savings
        
        spv_pc_saving  = 0
 
        solarkwpeak    = d4['Solar KW peak']
        
#        if d4['uprn'] == 42013472:
#          print('Solar kwpeak')
#          print(d4['solar_kw_peak'])
          
        if solarkwpeak == 0:
        
          sp_ret = calc_solar_pv_kwpeak(d4['Solar angle'], d4['Solar area m2'])
          
          if sp_ret['ef'] > 0:
            ms = sp_ret['em']
            em = f"Exception returned from calc_solar_pv_kwpeak \n {ms}\n"
            ret_mess['ef'] = sp_ret['ef']
            ret_mess['em'] = em
            return ret_mess
            
          solarkwpeak = sp_ret['kwpeak']
          
#          if d4['uprn'] == 42013472:
#            print('If solarkwpeak is zero')
#            print(solarkwpeak)
          
         
        coranngen      = d4['Corrected annual gen kWh']

#        if d4['Building ID'] == 42013472:
#          print('coranngen')
#          print(coranngen)  
          
        if coranngen   == 0:
        
          sp_ret = calc_solar_pv_corrected_annual_generation(solarkwpeak)
          
          if sp_ret['ef'] > 0:
            ms = sp_ret['em']
            em = f"Exception returned from calc_solar_pv_corrected_annual_generation \n {ms}\n"
            ret_mess['ef'] = sp_ret['ef']
            ret_mess['em'] = em
            return ret_mess          
          
          coranngen = sp_ret['coranngen']
          
#          if d4['uprn'] == 42013472:
#            print('coranngen calced when zero')
#            print(coranngen)   
        panel_m2 = 0

        if d4['Solar area m2'] > 0:
          panel_m2      = d4['Solar area m2']
        elif roof_space_m2 > 0:
          panel_m2      = roof_space_m2
        else:
          panel_m2      = gia_m2/3 # Very rough estimate in lack of anything better

#        if d4['uprn'] == 42013472:
#          print('before calc_solar_pv_total_zero_carbon_generation_v2')
#          print(f"assessed - {d4['assessed']}, listed - {listed}, panel_m2 - {panel_m2}, coranngen - {coranngen}")
          
        sp_ret = calc_solar_pv_total_zero_carbon_generation_v2(d4['Assessed'], listed, panel_m2, coranngen)

        if sp_ret['ef'] > 0:
            ms = sp_ret['em']
            em = f"Exception returned from calc_solar_pv_total_zero_carbon_generation \n {ms}\n"
            ret_mess['ef'] = sp_ret['ef']
            ret_mess['em'] = em
            return ret_mess

        totzerocarbgen = sp_ret['totzerocarbgen']

#        if d4['uprn'] == 42013472:
#          print('after calc_solar_pv_total_zero_carbon_generation_v2')
#          print(f" totzerocarbgen - {totzerocarbgen}")

        # Do not calculate a percent saving for Solar PV because we have already calculated the saving in kWh
        spv_pc_saving = 0
        fisql = f"SELECT project_id FROM projects WHERE ((entity_number = {entity_number}) AND (uprn = {uprn}) AND (project_type_id = {ptypeid}));" 
        cursor.execute(fisql)
        t_lodics        = cursor.fetchall()
        keys            = ("project_id","dummy_key")
        lodics          = [dict(zip(keys, values)) for values in t_lodics]       
        if len(lodics) == 0:
          print(f"In update_solar_pv_project_details ----=====\n UPRN : {uprn} \n Project type id : {ptypeid} \n dfsp  : \n {dfsp.to_string()}\n")
          
        dic             = lodics[0]
        pr_id           = dic['project_id']
#            print(f"----In finding project id - {pr_id}")
        upsql = f"UPDATE projects \
                SET assessed                        = \'{d4['Assessed']}\',\
                project_status                      = \'{d4['Status']}\',\
                utility                             = \'{d4['Utility']}\',\
                salix_pf                            = {d4['Lifetime (yrs)']},\
                cost_capex_mode                     = {d4['CAPEX']},\
                delivery_date_mode                  = \'{d4['Delivery date']}\',\
                saving_percent                      = {spv_pc_saving},\
                solar_roof_type                     = \'{d4['Solar roof type']}\',\
                solar_angle                         = \'{d4['Solar angle']}\',\
                solar_pv_area_m2                    = {d4['Solar area m2']},\
                solar_kw_peak                       = {d4['Solar KW peak']},\
                solar_pv_corrected_ann_gen_kwh      = {d4['Corrected annual gen kWh']},\
                solar_pv_tot_zero_carb_gen_kwh      = {totzerocarbgen} \
                WHERE project_id = {pr_id};"            

        try:
          cursor.execute(upsql)
          conn.commit()
          nupdates         = nupdates + 1
        
        except (pyodbc.Error) as e:

          nupdrawer         = nupdrawer + 1
          conn.rollback()
          up_log            = up_log + f"******* Database Exception updating solar pv projects - ****ERROR on record number {rec_num}. DB returned: - \n{e}\n"
                      
      fs1                   = ''  
      fs                    =      f"Number of solar pv project records read from upload file : {num_rows_read} \n"
      fs                    = fs + f"Number of solar pv projects that were updated            : {nupdates} \n"
      if nupdrawwa > 0:
        fs                  = fs + f"Number of database update warnings                       : {nupdrawwa}\n"
        ret_mess['ef']      = 1
        
      if nupdrawer > 0:  
        fs                  = fs + f"Number of database update errors                         : {nupdrawer}\n"
        ret_mess['ef']      = 2     
        
      if nupdrawwa > 0 or nupdrawer > 0:
        fs1                 = fs + f"Please see the upload log for details of warnings and errors. \n"

      summary               = summary + fs1
      up_log                = up_log  + fs
      ret_mess['summary']   = summary
      ret_mess['up_log']    = up_log

      return ret_mess

  except Exception as e: 
    print('#####In exception in update_solar_pv_project_details')
    
    summary = '******An exception has occurred in update_solar_pv_project_details - please see upload log for details'

    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))

    up_log  = f"Exception exit 1 \n {msg}"
    ret_mess['summary'] = summary
    ret_mess['up_log']  = f"up_log \n {msg}" 
    ret_mess['ef']      = 2
    ret_mess['em']      = msg

    return ret_mess
  
def update_solar_thermal_project_details(conn, dfst, entity_number, output_pt):

# At this point we have a dataframe with at worst warnings and no errors so now we can UPDATE solar thermal project records in the database.
  num_rows_read          = dfst.shape[0]
  ret_mess               = {'ef':0, 'em':'', 'up_log': '', 'summary': ''}
  summary                = ''
  up_log                 = ''  
  
  try:  
    
    with conn.cursor() as cursor:
      
    # Get the list of project type ids and names
    
      sqlt              = f"SELECT project_type_id, name FROM project_types;"
      cursor.execute(sqlt)
      t_output_pt       = cursor.fetchall()
      keys              = ("project_type_id","name")
      output_pt         = [dict(zip(keys, values)) for values in t_output_pt]
      
    # Get a list of uprn, controlled, baseline_annual_gas_kwh 
    
      sqls              = f"SELECT uprn, under_control, baseline_annual_gas_kwh FROM raw_estate_data WHERE entity_number = {entity_number};"
      cursor.execute(sqls)
      t_output_ugl      = cursor.fetchall()
      keys              = ("uprn", "under_control", "baseline_annual_gas_kwh")
      output_ugl        =  [dict(zip(keys, values)) for values in t_output_ugl]

    # Iterate down the down the dataframe (by project)
      rec_num    = 0
      nupdates   = 0
      nupdrawwa  = 0
      nupdrawer  = 0
      
      for d4 in dfst.to_dict(orient="records"): 
        rec_num = rec_num + 1

        # If delivery date is '0000-00-00'(which it can be when a blank cell is read in from Excel) convert it to 01/01/1900 before updating database'
        # to prevent giving Excel a date it doesn't understand when producing the project details data collection form
        
        if d4['delivery_date'] == '0000-00-00':
          sd                  = dt.datetime(1900,1,1)
          bdate               = dt.datetime.date(sd)

          d4['delivery_date'] = bdate
        # Find the project_type_id so we can find the project_id and use that to update the project and results_projects tables

        ptype          = d4['Project type']

        for n in output_pt:
          if n['name'] == ptype:
            ptypeid    = n['project_type_id']
            break  
        uprn = d4['uprn']
        
        # Get the controlled and baseline_annual_gas_kwh for the building
        
        founduprn = False
        for g in output_ugl:
          if g['uprn'] == uprn:
            controlled      = g['under_control']
            baseline_annual_gas_kwh  = g['baseline_annual_gas_kwh']
            founduprn = True
            break
        
        if not founduprn:
          ret_mess['ef']  = 2
          ret_mess['em']  = f"****Error in update_solar_thermal_project_details unable to find uprn - {uprn}\n"
          return ret_mess
        
        # Calculate the percentage savings
        
        sthermal_pc_saving  = 0
        
        coranngen      = d4['Corrected annual gen kWh']
        
        if coranngen   == 0:

          sp_ret = calc_solar_thermal_corrected_annual_generation(d4['Solar area m2'])
          
          if sp_ret['ef'] > 0:
            ms = sp_ret['em']
            em = f"Exception returned from calc_solar_thermal_corrected_annual_generation \n {ms}\n"
            ret_mess['ef'] = sp_ret['ef']
            ret_mess['em'] = em
            return ret_mess          
          
          coranngen = sp_ret['coranngen']


#       We do not need to calculate a percent saving for Solar Thermal as we have already claculated a saving in kWh
        sthermal_pc_saving = 0

        fisql = f"SELECT project_id FROM projects WHERE ((entity_number = {entity_number}) AND (uprn = {uprn}) AND (project_type_id = {ptypeid}));" 
        cursor.execute(fisql)
        t_lodics        = cursor.fetchall() 
        keys            = ("project_id","dummy_key")    
        lodics          = [dict(zip(keys, values)) for values in t_lodics]
        
        dic             = lodics[0]
        pr_id           = dic['project_id']
#            print(f"----In finding project id - {pr_id}")
        upsql = f"UPDATE projects \
                SET assessed                        = \'{d4['Assessed']}\',\
                project_status                      = \'{d4['Status']}\',\
                utility                             = \'{d4['Utility']}\',\
                salix_pf                            = {d4['Lifetime (yrs)']},\
                cost_capex_mode                     = {d4['CAPEX']},\
                delivery_date_mode                  = \'{d4['Delivery date']}\',\
                saving_percent                      = {sthermal_pc_saving},\
                solar_thermal_area_m2               = {d4['Solar area m2']},\
                solar_thermal_corrected_ann_gen_kwh = {coranngen}\
                WHERE project_id = {pr_id};"            

        try:
          cursor.execute(upsql)
          conn.commit()
          nupdates         = nupdates + 1
        
        except (pyodbc.Error) as e:

          nupdrawer         = nupdrawer + 1
          conn.rollback()
          up_log            = up_log + f"******* Database Exception updating solar thermal projects - ****ERROR on record number {rec_num}. DB returned: - \n{e}\n"
                      
      fs1                   = ''  
      fs                    =      f"Number of solar thermal project records read from upload file : {num_rows_read} \n"
      fs                    = fs + f"Number of solar thermal projects that were updated            : {nupdates} \n"
      if nupdrawwa > 0:
        fs                  = fs + f"Number of database update warnings                       : {nupdrawwa}\n"
        ret_mess['ef']      = 1
        
      if nupdrawer > 0:  
        fs                  = fs + f"Number of database update errors                         : {nupdrawer}\n"
        ret_mess['ef']      = 2     
        
      if nupdrawwa > 0 or nupdrawer > 0:
        fs1                 = fs + f"Please see the upload log for details of warnings and errors. \n"

      summary               = summary + fs1
      up_log                = up_log  + fs
      ret_mess['summary']   = summary
      ret_mess['up_log']    = up_log

      return ret_mess

  except Exception as e: 
    summary = '******An exception has occurred in update_solar_thermal_project_details - please see upload log for details'

    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))

    up_log  = f"Exception exit 1 \n {msg}"
    ret_mess['summary'] = summary
    ret_mess['up_log']  = up_log
    ret_mess['ef']      = 2

    return ret_mess
    
def calc_solar_pv_kwpeak(roof_angle, solar_pv_area_m2):
# In the event that the client does not provide their own KW Peak this function calculates
# a value based on the roof angle and roof area (square metres) available for solar pv.
# sf is imported server module Solar_factors
  try:
    ret_mess         = {'ef':0, 'em':'', 'kwpeak':0}
    kwpeak           = 0
    valid_roof_angle = ['Pitched','Flat','']

    if roof_angle not in valid_roof_angle:
      ret_mess['ef'] = 2
      ret_mess['em'] = '****Roof angle input is not valid'
      return ret_mess
  
    if solar_pv_area_m2 <= 0:
      return ret_mess
  
    if roof_angle == 'Pitched':
      kwpeak = (solar_pv_area_m2/sf.PitchCF) * sf.RoofCF
    if roof_angle == 'Flat':
      kwpeak = (solar_pv_area_m2/sf.FlatCF) * sf.RoofCF
  
    ret_mess['kwpeak'] = kwpeak
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess    
def calc_solar_pv_corrected_annual_generation(kwpeak):
# In the event that the client does not provide their own Corrected Annual Generation this function calculates
# a value based on the kwpeak and the Generation Factor for solar pv.
# sf is imported server module Solar_factors
  try:
    ret_mess              = {'ef':0, 'em':'', 'coranngen':0}
    coranngen             = kwpeak * sf.Generation_Factor
    
    ret_mess['coranngen'] = coranngen
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess    

def calc_solar_thermal_corrected_annual_generation(solar_thermal_area_m2):
# In the event that the client does not provide their own Solar Thermal Corrected Annual Generation this function calculates
# a value based on the solar_thermal_area_m2 
# sf is imported server module Solar_factors
  try:
    ret_mess              = {'ef':0, 'em':'', 'coranngen':0}
    coranngen             = solar_thermal_area_m2 * sf.Solar_thermal_kwh_per_panel * sf.RoofCF
    ret_mess['coranngen'] = coranngen
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess    

def calc_solar_thermal_percent_saving(controlled, solar_thermal_corrected_annual_generation, annual_gas_kwh):
# Calculates the percentage solar thermal saving for a controlled building based on the solar thermal corrected annual generation as a percentage of the annual gas kwh usage. 

  try:
    ret_mess              = {'ef':0, 'em':'', 'pcsothsaving':0}
    
    if controlled != 'YES':
      return ret_mess
    
    if solar_thermal_corrected_annual_generation == 0 or annual_gas_kwh == 0:
      return ret_mess
    
    pcsothsaving = solar_thermal_corrected_annual_generation /  annual_gas_kwh
    ret_mess['pcsothsaving'] = pcsothsaving
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess    

def calc_solar_pv_total_zero_carbon_generation(assessed, listed, gia_m2, solar_pv_corrected_ann_gen_kwh):
# Calculates the total zero carbon generation by solar PV. 
#
# If a non zero value is provided for the solar pv corrected annual generation then that value will be returned.
# If the assessed state is 'ASSESSED/NV', 'IN PLACE' or Null then zero is returned.
# If listed is 'Yes' then zero is returned.
# If the assessed state is 'FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL' then the following calculation is used: -
# ((gia_m2 * sf.Roof_GIA_factor)/sf.Kwpeak_factor) * sf.RoofCF * 800
#    where gia_m2 is Gross Internal Area of the building in square metres 
# sf is imported server module Solar_factors
  try:
    totzerocarbgen        = 0
    ret_mess              = {'ef':0, 'em':'', 'totzerocarbgen':0}
    
    if listed == 'YES':
      return ret_mess
    
    if solar_pv_corrected_ann_gen_kwh > 0:
      ret_mess['totzerocarbgen']  = solar_pv_corrected_ann_gen_kwh
      return ret_mess
    
    if (assessed == 'ASSESSED/NV') or (assessed == 'IN PLACE') or (assessed == ''):
      return ret_mess
      
    if (assessed == 'FIRM') or (assessed == 'LIKELY') or (assessed == 'POSSIBLE') or (assessed == 'POTENTIAL'):
      totzerocarbgen        = ((gia_m2 * sf.Roof_GIA_factor)/sf.Kwpeak_factor) * sf.RoofCF * 800
      ret_mess['totzerocarbgen'] = totzerocarbgen
      return ret_mess
      
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess  

def calc_solar_pv_total_zero_carbon_generation_v2(assessed, listed, panel_m2, solar_pv_corrected_ann_gen_kwh):
# Calculates the total zero carbon generation by solar PV. 
#
# If a non zero value is provided for the solar pv corrected annual generation then that value will be returned.
# If the assessed state is 'ASSESSED/NV', 'IN PLACE' or Null then zero is returned.
# If listed is 'Yes' then zero is returned.
# If the assessed state is 'FIRM', 'LIKELY', 'POSSIBLE', 'POTENTIAL' then the following calculation is used: -
# ((panel_m2 * sf.Roof_GIA_factor)/sf.Kwpeak_factor) * sf.RoofCF * 800
#    where panel_m2 is estimated area in square metres available for solar pv panels
# sf is imported server module Solar_factors
  try:
    totzerocarbgen        = 0
    ret_mess              = {'ef':0, 'em':'', 'totzerocarbgen':0}
    
    if listed == 'YES':
      return ret_mess
    
    if solar_pv_corrected_ann_gen_kwh > 0:
      ret_mess['totzerocarbgen']  = solar_pv_corrected_ann_gen_kwh
      return ret_mess
    
    if (assessed == 'ASSESSED/NV') or (assessed == 'IN PLACE') or (assessed == ''):
      return ret_mess
      
    if (assessed == 'FIRM') or (assessed == 'LIKELY') or (assessed == 'POSSIBLE') or (assessed == 'POTENTIAL'):
      totzerocarbgen        = ((panel_m2 * sf.Roof_GIA_factor)/sf.Kwpeak_factor) * sf.RoofCF * 800
      ret_mess['totzerocarbgen'] = totzerocarbgen
      return ret_mess
      
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess    
def calc_solar_pv_percent_saving(solar_pv_corrected_annual_generation,  total_zero_carbon_generation, annual_elec_kwh):
# Calculates the percentage solar pv saving for a building (percentage of annual electricity kwh). If solar_pv_corrected_annual_generation is > zero then 
# this is used in the calculation otherwise the total_zero_carbon_generation is used.

  try:
    ret_mess              = {'ef':0, 'em':'', 'pcsopvsaving':0}

    if annual_elec_kwh == 0:
      return ret_mess
    
    if solar_pv_corrected_annual_generation > 0:
      ret_mess['pcsopvsaving'] = solar_pv_corrected_annual_generation / annual_elec_kwh
      return ret_mess
    else:
      ret_mess['pcsopvsaving'] = total_zero_carbon_generation / annual_elec_kwh
      return ret_mess

  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
  return ret_mess
  
def calc_solar_summary(conn, entity_number):
  # Calculates the Solar Summary - Total PV Generation (Identified)  (MWH), TOTAL PV PEAK (MW) IDENTIFIED, Total (PV) Potential Remaining (MWH),
  # TOTAL PEAK (MW) REMAINING, TOTAL SOLAR THERMAL (MWH)
  # for all controlled buildings in entity with number entity_number that have solar projects (project_type_id = 20 or 21).
  
  ret_mess = {'ef':0, 'em':'','summary':'', 'up_log':''}
  summary  = ''
  up_log   = ''
  try:
    with conn.cursor() as cursor:
      
      # Get uprns and control status for all buildings in entity and convert to a dataframe (dfcont)
      sqlup = f"SELECT uprn, under_control FROM raw_estate_data WHERE entity_number = {entity_number} AND under_control = \'Yes\';"
      cursor.execute(sqlup)
      t_output_up = cursor.fetchall()
      keys        = ("uprn","under_control")
      output_up   = [dict(zip(keys, values)) for values in t_output_up]
      print('In calc solar summary loc 1 dfcont')
            
      dfcont = pd.DataFrame.from_dict(output_up)
      print(dfcont.to_string())
      
      total_pv_gen_id_mwh                 = 0
      total_pv_peak_id_mwh                = 0
      total_pv_potential_remain_mwh       = 0
      total_pv_peak_remain_kwh            = 0
      total_thermal_kwh                   = 0     
       
    # Get all the solar projects for this entity and covert to a dataframe (dfsolp). Calculate total_pv_gen_id_mwh, total_pv_peak_id_mwh, total_pv_potential_remain_mwh , total_pv_peak_remain_kwh , total_thermal_kwh and write these to the projects table
        
      sqlpr = f"SELECT uprn, assessed, solar_kw_peak, solar_thermal_corrected_ann_gen_kwh, solar_pv_corrected_ann_gen_kwh, solar_pv_tot_zero_carb_gen_kwh FROM projects \
              WHERE entity_number = {entity_number} AND (project_type_id = 20 OR project_type_id = 21) AND assessed != 'ASSESSED/NV' AND assessed != 'IN PLACE';"
      cursor.execute(sqlpr)
      t_output_prd = cursor.fetchall()
      keys         = ("uprn", "assessed", "solar_kw_peak", "solar_thermal_corrected_ann_gen_kwh", "solar_pv_corrected_ann_gen_kwh", "solar_pv_tot_zero_carb_gen_kwh")
      output_prd   = [dict(zip(keys, values)) for values in t_output_prd]
      
      dfsolp       = pd.DataFrame.from_dict(output_prd)
      print('====In calc_solar_summary dfsolp ')
      print(dfsolp.to_string())
      # Insert a column to hold the controlled status
      
      dfsolp.insert(loc=0,column    ='controlled', value = '')
    
      # Populate the controlled status column by uprn from the dfcont dataframe of uprn and controlled
      
      for index, row in dfsolp.iterrows():
        uprn       = row['uprn']
        controlled = dfcont[dfcont['uprn']==uprn]['under_control'].values[0]
        dfsolp['controlled'][index] = controlled

      print('=====dfsolp after  Populate the controlled status column by uprn from the dfcont dataframe of uprn and controlled')
      print(dfsolp.to_string())
      # Remove all rows (projects) where the building is not controlled (controlled = 'NO')
      
      dfsolp       = dfsolp[dfsolp.controlled == 'YES']
      
      print('====dfsolp after removing buildings not controlled')
      print(dfsolp.to_string())
      # Calculate solar summary

      total_pv_gen_id_mwh                 = dfsolp['solar_pv_corrected_ann_gen_kwh'].sum()/1000
      total_pv_peak_id_mwh                = total_pv_gen_id_mwh/800
      total_pv_potential_remain_mwh       = (dfsolp['solar_pv_tot_zero_carb_gen_kwh'].sum()/1000) - (dfsolp['solar_pv_corrected_ann_gen_kwh'].sum()/1000)
      total_pv_peak_remain_kwh            = total_pv_potential_remain_mwh/800
      total_thermal_kwh                   = dfsolp['solar_thermal_corrected_ann_gen_kwh'].sum()/1000     

      sqld    = f"DELETE FROM solar_estate_summary WHERE entity_number = {entity_number};" 
      cursor.execute(sqld)
      conn.commit

      sqlin = f"INSERT INTO solar_estate_summary (entity_number, total_pv_gen_id_mwh, total_pv_peak_id_mwh, total_pv_potential_remain_mwh, total_pv_peak_remain_kwh, total_thermal_kwh) VALUES ({entity_number}, {total_pv_gen_id_mwh}, {total_pv_peak_id_mwh}, {total_pv_potential_remain_mwh}, {total_pv_peak_remain_kwh}, {total_thermal_kwh});"
      print('sqlin')
      print(sqlin)
      try:
        cursor.execute(sqlin)
        conn.commit()
      except (pyodbc.Error) as e:
        conn.rollback()
        print('In pyodbc error')
        up_log            = up_log + f"****Error - calc_solar_summary DB Exception INSERT INTO solar_estate_summary: - \n{e}\n"
        summary           = summary +"****Error - see log for details"
        ret_mess          = {'ef': 2, 'em': f"Database error - \n {e}", 'summary':summary, 'up_log':up_log }
        return ret_mess
          
  except Exception as e:
    print('In Exception as e')
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = f"****Exception in calc_solar_summary: - \n {msg}\n"
  return ret_mess  

def write_raw_estate_lite_data_to_excel(dfl, dfr, entity_id, entity_number, dt_str):

# Reads the Estate Lite data capture template from Anvil file as media object and download it to an Excel file (.xlsx) in local disc storage
# Assumes entity_id and entity_number are already validated when this function is called.
  try:
    ret_mess  = {'ef':0, 'em':0, 'rmedia':''}
    print('____IN download_estate_lite_template_v001 at top')

    print(dt_str)
    
    output_filename   = f"OPF_estate_data_entry_for_entity_{entity_id} produced at {dt_str}"
    print('====Output_filename:')
    print(output_filename)

    # Read in the Estate Lite upload template from the documents table

    name              = 'EstateLite_Template_V14'
    row               = app_tables.documents.get(Name="EstateLite_Template_V14")
    content           = row['Document']
    file              = content.get_bytes()
    bytestream        = io.BytesIO(file)
    bytestream.seek(0)

    # From https://anvil.works/forum/t/upload-an-excel-to-a-datatable-media-object-column/16507/4
    book = load_workbook(bytestream)
    #print(book)
   # print(workbook.sheetnames)
    
# Create a Pandas Excel writer using XlsxWriter as the engine.
#    writer             = pd.ExcelWriter(bytestream, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')
#    writer             = pd.ExcelWriter(bytestream, engine = 'openpyxl',date_format = 'dd/mm/yyyy')    
#    writer.book        = book
# Convert the dataframes to XlsxWriter Excel objects.
#    dfl.to_excel(writer, sheet_name='Input sheet', index=False, header=False, startrow=1, startcol=0)
#    dfr.to_excel(writer, sheet_name='Input sheet', index=False, header=False, startrow=1, startcol=3)
#    writer.save()
#    writer.close()

    with pd.ExcelWriter(bytestream,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        dfl.to_excel(writer, sheet_name="Input sheet", index=False, header=False, startrow=1, startcol=0)
        dfr.to_excel(writer, sheet_name="Input sheet", index=False, header=False, startrow=1, startcol=3)   

    bytestream.seek(0)
# Get the xlsxwriter objects from the dataframe writer object.
#    workbook  = writer.book
#    file              = book.get_bytes()

#    bytestream.seek(0)
#    book              = openpyxl.load_workbook(bytestream)
#    abm               = anvil.BlobMedia(content=bytestream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = output_filename)
    abm               = anvil.BlobMedia(content=bytestream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = output_filename)
    ret_mess['rmedia']= abm
    
    return ret_mess
  except Exception as e:
    ret_mess['ef'] = 1
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['em'] = msg
    print(msg)
  return ret_mess
  
def write_raw_estate_data_to_excel(dfes_in, df_bt, entity):
  ret  = {'ef': 0, 'em' : '', 'abm' : ''}

  import datetime as dt
  now               = datetime.now()
  dt_str            = now.strftime("%D/%M/%Y")  
  try:  
    dfes      = dfes_in.copy()

    print('In write_raw_estate_data_to_excel')
    
    content = io.BytesIO()
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(content, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')

# Convert the dataframes to XlsxWriter Excel objects.
    dfes.to_excel(writer, sheet_name='estate_data', index=False)
    df_bt.to_excel(writer, sheet_name='reference', index=False, header=False, startcol=0)
    
# Get the xlsxwriter objects from the dataframe writer object.
    workbook  = writer.book
    worksheetes = writer.sheets['estate_data']
    worksheetbt = writer.sheets['reference']
    
# Get the dimensions of the estate dataframe.
    (max_row, max_col) = dfes.shape
    (bt_row,  bt_col)  = df_bt.shape

# Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in dfes.columns]

# Add the Excel table structure. Pandas will add the data.
    worksheetes.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

# Set up validation of columns in the Excel sheet 
    # Estate data

    # action
    worksheetes.data_validation('A2:A1048576', {'validate': 'list','source'  : ["UPDATE","DELETE"]})
    # delete reason
    worksheetes.data_validation('B2:B1048576', {'validate': 'list','source'  : ["FIRE","FLOOD","EXPLOSION","DEMOLISHED","TRANSFERRED","OTHER"]})
    # uprn
    worksheetes.data_validation('C2:C1048576', {'validate': 'integer','criteria' : 'between', 'minimum' : 1, 'maximum' : 999999999999999 })
    # building type
    worksheetes.data_validation('E2:E1048576', {'validate': 'list','source'  : '=reference!$A$1:$A$128'})
    # under control
    worksheetes.data_validation('H2:H1048576', {'validate': 'list','source'  : ["YES","NO"]})
    # remain in portfolio
    worksheetes.data_validation('I2:I1048576', {'validate': 'list','source'  : ["YES","NO"]})
    # listed
    worksheetes.data_validation('J2:J1048576', {'validate': 'list','source'  : ["YES","NO"]}) 
    # construction year
    worksheetes.data_validation('K2:K1048576', {'validate': 'list','source'  : ["2020","2015","2010","2005","2000","1990S","1980S","1970S","1960S","1950S","PRE 1950","PRE 1900S"]}) 
    # gia_m2
    worksheetes.data_validation('L2:L1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # roof space m2
    worksheetes.data_validation('M2:M1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # data year
    worksheetes.data_validation('N2:N1048576', {'validate': 'integer','criteria' : 'between', 'minimum' : 2020, 'maximum' : 2050 })
    # baseline elec annual kwh
    worksheetes.data_validation('O2:O1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # baseline gas annual kwh
    worksheetes.data_validation('P2:P1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # baseline oil annual kwh
    worksheetes.data_validation('Q2:Q1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # baseline lpg annual kwh
    worksheetes.data_validation('R2:R1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # source of heating
    worksheetes.data_validation('S2:S1048576', {'validate': 'list','source'  : [ "ELECTRICITY","GAS","OIL","LPG","COAL","BIOMASS"]})
    # source of dhw
    worksheetes.data_validation('T2:T1048576', {'validate': 'list','source'  : [ "ELECTRICITY","GAS","OIL","LPG","COAL","BIOMASS"]})
    # DEC score
    worksheetes.data_validation('U2:U1048576', {'validate': 'integer','criteria': '>=', 'value': 0 })
    # epc score
    worksheetes.data_validation('V2:V1048576', {'validate': 'integer','criteria': '>=', 'value': 0 })
    # baseline elec cost per kwh
    worksheetes.data_validation('W2:W1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # baseline gas cost per kwh
    worksheetes.data_validation('X2:X1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 }) 
    # baseline oil cost per kwh
    worksheetes.data_validation('Y2:Y1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 }) 
    # baseline lpg cost per kwh
    worksheetes.data_validation('Z2:Z1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 }) 
    # onsite generation annual kwh
    worksheetes.data_validation('AA2:AA1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 }) 
    # existing solar PV annual kwh
    worksheetes.data_validation('AB2:AB1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # existing solar thermal annual kwh
    worksheetes.data_validation('AC2:AC1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # existing non solar decarb heat annual kwh
    worksheetes.data_validation('AD2:AD1048576', {'validate': 'integer','criteria': '>=', 'value': 0 })
    # car park available
    worksheetes.data_validation('AE2:AE1048576', {'validate': 'list','source'  : ["YES","NO"]}) 
    # number of ev charging sockets
    worksheetes.data_validation('AF2:AF1048576', {'validate': 'integer','criteria' : 'between', 'minimum' : 0, 'maximum' : 99999 })
    # charging capacity kwh
    worksheetes.data_validation('AG2:AG1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })

#   Protect the worksheet without password so user can unprotect to filter and sort
    optionses = {
    'format_cells':          True,
    'format_columns':        True,
    'format_rows':           True,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           True,
    'select_locked_cells':   True,
    'sort':                  True,
    'autofilter':            True,
    'pivot_tables':          False,
    'select_unlocked_cells': True,
    }

    optionsbt = {
    'format_cells':          False,
    'format_columns':        False,
    'format_rows':           False,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           False,
    'select_locked_cells':   False,
    'sort':                  True,
    'autofilter':            False,
    'pivot_tables':          False,
    'select_unlocked_cells': False,
    }    
   
#   Set default of cells as unlocked
    unlocked = workbook.add_format({'locked': False})

#   Unprotect the cells where we allow the user to change values
    worksheetes.unprotect_range('A2:B1048576')
    worksheetes.unprotect_range('D2:AG1048576')

#   Shade the protected columns
    cell_format_prot = workbook.add_format()
    cell_format_prot.set_bg_color('#e0e0eb')
    worksheetes.set_column('C:C',10,cell_format_prot)

#   Set formats for numeric columns

    format2dp  = workbook.add_format({'num_format': '#,##0.00'})
    format1dp  = workbook.add_format({'num_format': '#,##0.0'})
    format0dp  = workbook.add_format({'num_format': '#,##0'})
    format0dpc = workbook.add_format({'num_format': '###0'})
#   Set column widths
    # Non solar
    worksheetes.set_column('A:A', 14)  
    worksheetes.set_column('B:B', 14) 
    worksheetes.set_column('C:C', 15, format0dpc)
    worksheetes.set_column('D:D', 45)
    worksheetes.set_column('E:E', 45)
    worksheetes.set_column('F:F', 50)
    worksheetes.set_column('G:G', 14) 
    worksheetes.set_column('H:H', 20) 
    worksheetes.set_column('I:I', 20) 
    worksheetes.set_column('J:J', 20) 
    worksheetes.set_column('K:K', 20)
    worksheetes.set_column('L:L', 15,format1dp)
    worksheetes.set_column('M:M', 15,format1dp)
    worksheetes.set_column('N:N', 12,format0dpc)
    worksheetes.set_column('O:O', 25,format0dp)
    worksheetes.set_column('P:P', 25,format0dp)
    worksheetes.set_column('Q:Q', 25,format0dp)
    worksheetes.set_column('R:R', 25,format0dp)
    worksheetes.set_column('S:S', 20)
    worksheetes.set_column('T:T', 18)
    worksheetes.set_column('U:U', 12, format1dp)
    worksheetes.set_column('V:V', 10, format1dp)
    worksheetes.set_column('W:W', 28, format2dp)
    worksheetes.set_column('X:X', 28, format2dp)
    worksheetes.set_column('Y:Y', 28, format2dp)
    worksheetes.set_column('Z:Z', 28, format2dp)
    worksheetes.set_column('AA:AA', 35, format0dp)
    worksheetes.set_column('AB:AB', 35, format0dp)
    worksheetes.set_column('AC:AC', 40, format0dp)
    worksheetes.set_column('AD:AD', 40, format0dp)
    worksheetes.set_column('AE:AE', 25, format1dp)
    worksheetes.set_column('AF:AF', 30, format1dp)
    worksheetes.set_column('AG:AG', 30, format1dp)
    
    worksheetes.protect('',optionses)
    worksheetbt.protect('',optionsbt)
    
    # Hide zeros to make sheet look neater and easier to identify where there are data missing

    worksheetes.hide_zero()

    # Protect workbook structure
    workbook.protect(password='OPFpw321')
    
    writer.close()
    xlsx_data = content.getvalue()
    print('past dfes.to_excel')

# Create anvil blobmedia object

    abm = anvil.BlobMedia(content=xlsx_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = f"estate_data_capture_entity_{entity} - {dt_str}")
    ret['abm'] = abm
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
  return ret

def write_raw_estate_data_to_excel_PC_01(dfes_in, df_bt, entity, partner, client):
  ret  = {'ef': 0, 'em' : '', 'abm' : ''}
# First version for Partner Channel
  import datetime as dt
  now               = datetime.now()
  dt_str            = now.strftime("%D/%M/%Y")  
  try:  
    dfes      = dfes_in.copy()
    cols = ['Building ID', 'Building Name', 'Building Type (choose from drop down)',
       'Address', 'Postcode', 'Latitude (decimal degrees)',
       'Longitude (decimal degrees)', 'Gross Internal Area (square metres)',
       'Roof space (square metres)',
       'Year (enter the year the energy usage relates to)',
       'Annual Electricity Usage (kWh)', 'Annual Gas Usage (kWh)',
       'Annual Oil Usage (kWh)', 'Annual LPG Usage (kWh)',
       'Total onsite generation (kWh)', 'Annual Solar PV usage (kWh)',
       'Annual Solar Thermal usage (kWh)',
       'Electricity purchased from REGO sources (kWh)',
       'Annual non-solar decarbonised heat usage (kWh)',
       'Electricity cost per kWh', 'Gas cost per kWh', 'Oil cost per kWh',
       'LPG cost per kWh', 'REGO electricity cost per kWh', 'Heating source',
       'Hot water source', 'DEC Score', 'EPC Score']
    dfes.columns = cols
    print('In write_raw_estate_data_to_excel_PC_01')
    print(dfes.to_string())
    print('END Print dfes---------')
    content = io.BytesIO()
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(content, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')

# Convert the dataframes to XlsxWriter Excel objects.
    dfes.to_excel(writer, sheet_name='Input Sheet', index=False, header = True, startrow = 1)
    df_bt.to_excel(writer, sheet_name='Reference', index=False, header=False, startcol=0)

# Get the xlsxwriter objects from the dataframe writer object.
    workbook    = writer.book
# Create worksheet objects
    worksheetes = writer.sheets['Input Sheet']
    worksheetbt = writer.sheets['Reference']
    worksheetau = workbook.add_worksheet('Auth')
    worksheetky = workbook.add_worksheet('Key') 
# Write the auth sheet
    ret_mess = write_auth_sheets(workbook, worksheetau, worksheetky, partner, client, entity)
    ef       = ret_mess['ef']
    em       = ret_mess['em']
    if ef == 2:
      ret['ef'] = ef
      ret['em'] = em
      return ret
# Get the dimensions of the estate dataframe.
    (max_row, max_col) = dfes.shape
    print(f"\nIn write_estate - max_row = {max_row}\n")
    

    # Write the data for row 1 which is the group headers. Then merge cells to create group headers.
    # Merge cells on row 1 to create column group headers
    worksheetes.merge_range('A1:C1','Building Identity')
    worksheetes.merge_range('D1:G1','Location')
    worksheetes.merge_range('H1:I1','Building Size')
    worksheetes.merge_range('L1:O1','Fosil Fuel Utilisation')
    worksheetes.merge_range('P1:S1','Green Energy Utilisation')
    worksheetes.merge_range('T1:X1','Energy Unit Costs')
    worksheetes.merge_range('Y1:Z1','Sources')
    worksheetes.merge_range('AA1:AB1','DEC/EPC')

    # Set up the group formats (Row 1)
    cell_format_BID     = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#305496', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_LOC     = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#7b7b7b', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_SIZE    = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#a6a6a6', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_YEAR    = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#bfbfbf', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_GREU    = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#bf8f00', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_FFU     = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#806000', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_GEEU    = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#70ad47', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_EUC     = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#b4c6e7', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_SOURCES = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#a6a6a6', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})
    cell_format_DECEPC  = workbook.add_format({'font_name': 'Calibri', 'font_size':14, 'font_color':'white', 'bg_color': '#bfbfbf', 'text_wrap':'true', 'bold': True, 'center_across':'True', 'valign':'top', 'align':'center', 'border_color':'white', 'right':1})

    # Write merged column groups Row 1 
    worksheetes.write('A1','Building Identity', cell_format_BID)
    worksheetes.write('D1','Location', cell_format_LOC)
    worksheetes.write('H1','Building Size', cell_format_SIZE)
    worksheetes.write('J1', 'Data Year', cell_format_YEAR)
    worksheetes.write('K1', 'Grid Electricity Utilisation', cell_format_GREU)    
    worksheetes.write('L1','Fossil Fuel Utilisation', cell_format_FFU)
    worksheetes.write('P1','Green Energy Utilisation', cell_format_GEEU)
    worksheetes.write('T1','Energy Unit Costs', cell_format_EUC)
    worksheetes.write('Y1','Sources', cell_format_SOURCES)
    worksheetes.write('AA1','DEC/EPC', cell_format_DECEPC)
    
    # Set up column header formats for Row 2
    cell_format_BID2    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#305496', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_LOC2    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#7b7b7b', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_SIZE2   = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#a6a6a6', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_YEAR2   = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#bfbfbf', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_GREU2   = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#bf8f00', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_FFU2    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#806000', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_GEEU2   = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#70ad47', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_EUC2    = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#b4c6e7', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_SOURCES2= workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#a6a6a6', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})
    cell_format_DECEPC2 = workbook.add_format({'font_name': 'Calibri', 'font_size':12, 'font_color':'white', 'bg_color': '#bfbfbf', 'text_wrap':'true', 'bold': True, 'valign':'vcenter', 'align':'center'})    
	

# Create a list of column headers, to use in add_table().
    header_format   = workbook.add_format({'text_wrap':True,'font_name': 'Calibri', 'font_size':12, 'valign':'vcenter', 'align':'center'})
    
    column_settings = [{'header': column,'header_format':header_format} for column in dfes.columns]

# Add the Excel table structure. Pandas will add the data.
    worksheetes.add_table(1, 0, max_row + 1, max_col - 1, {'columns': column_settings, 'header_row': True})

    # Set up validation of columns in the Excel sheet 
    # Partner Estate data

    # building id
    worksheetes.data_validation('A3:A1048576', {'validate': 'integer','criteria' : 'between', 'minimum' : 1, 'maximum' : 999999999999999 })
    # building type
    worksheetes.data_validation('C3:C1048576', {'validate': 'list','source'  : '=Reference!$A$1:$A$129'})
    # Latitude & Longitude
    worksheetes.data_validation('F3:F1048576', {'validate': 'decimal','criteria' : 'between', 'minimum' : -90, 'maximum' : +90})
    worksheetes.data_validation('G3:G1048576', {'validate': 'decimal','criteria' : 'between', 'minimum' : -180, 'maximum' : +180 })
    # gia_m2
    worksheetes.data_validation('H3:H1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # roof space m2
    worksheetes.data_validation('I3:I1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    #  Data year
    worksheetes.data_validation('J3:J1048576', {'validate': 'integer','criteria' : 'between', 'minimum' : 2020, 'maximum' : 2050 })    
    # Grid electricity utilisation
    worksheetes.data_validation('K3:K1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual gas usage
    worksheetes.data_validation('L3:L1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual oil usage
    worksheetes.data_validation('M3:M1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual lpg usage
    worksheetes.data_validation('N3:N1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual onsite generation
    worksheetes.data_validation('O3:O1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual Solar PV usage
    worksheetes.data_validation('P3:P1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual Solar Thermal
    worksheetes.data_validation('Q3:Q1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual Elec purchased from REGO
    worksheetes.data_validation('R3:R1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Annual non-solar decarbonised heat usage
    worksheetes.data_validation('S3:S1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })    
    # Electricity cost per kWh
    worksheetes.data_validation('T3:T1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Gas cost per kWh
    worksheetes.data_validation('U3:U1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Oil cost per kWh
    worksheetes.data_validation('V3:V1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # LPG cost per kWh
    worksheetes.data_validation('W3:W1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # REGO electricity cost per kWh
    worksheetes.data_validation('X3:X1048576', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    # Heating source
    worksheetes.data_validation('Y3:Y1048576', {'validate': 'list','source'  : [ "ELECTRICITY","GAS","OIL","LPG","COAL","BIOMASS"]}) 
    # Hot water source
    worksheetes.data_validation('Z3:Z1048576', {'validate': 'list','source'  : [ "ELECTRICITY","GAS","OIL","LPG","COAL","BIOMASS"]}) 
    # DEC score
    worksheetes.data_validation('AA3:AA1048576', {'validate': 'integer','criteria': '>=', 'value': 0 })
    # epc score
    worksheetes.data_validation('AB3:AB1048576', {'validate': 'integer','criteria': '>=', 'value': 0 })

#  Set options for worksheets
    optionses = {
    'format_cells':          False,
    'format_columns':        False,
    'format_rows':           False,
    'insert_columns':        False,
    'insert_rows':           True,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           True,
    'select_locked_cells':   False,
    'sort':                  True,
    'autofilter':            True,
    'pivot_tables':          False,
    'select_unlocked_cells': True,
    }

    optionsbt = {
    'format_cells':          False,
    'format_columns':        False,
    'format_rows':           False,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           False,
    'select_locked_cells':   False,
    'sort':                  True,
    'autofilter':            False,
    'pivot_tables':          False,
    'select_unlocked_cells': False,
    }    
   
#   Set formats for numeric columns

    format3dp  = workbook.add_format({'num_format': '#,##0.000'})
    format2dp  = workbook.add_format({'num_format': '#,##0.00'})
    format1dp  = workbook.add_format({'num_format': '#,##0.0'})
    format0dp  = workbook.add_format({'num_format': '#,##0'})
    format0dpc = workbook.add_format({'num_format': '###0'})
#   Set column widths
    
    worksheetes.set_column('A:A', 14)  
    worksheetes.set_column('B:B', 45) 
    worksheetes.set_column('C:C', 50)
    worksheetes.set_column('D:D', 45)
    worksheetes.set_column('E:E', 14)
    worksheetes.set_column('F:F', 14)
    worksheetes.set_column('G:G', 14) 
    worksheetes.set_column('H:H', 20,format0dp) 
    worksheetes.set_column('I:I', 20,format0dp) 
    worksheetes.set_column('J:J', 14) 
    worksheetes.set_column('K:K', 18,format0dp)
    worksheetes.set_column('L:L', 18,format0dp)
    worksheetes.set_column('M:M', 18,format0dp)
    worksheetes.set_column('N:N', 18,format0dp)
    worksheetes.set_column('O:O', 18,format0dp)
    worksheetes.set_column('P:P', 18,format0dp)
    worksheetes.set_column('Q:Q', 18,format0dp)
    worksheetes.set_column('R:R', 18,format0dp)
    worksheetes.set_column('S:S', 18,format0dp)
    worksheetes.set_column('T:T', 18,format3dp)
    worksheetes.set_column('U:U', 18,format3dp)
    worksheetes.set_column('V:V', 18,format3dp)
    worksheetes.set_column('W:W', 18,format3dp)
    worksheetes.set_column('X:X', 18,format3dp)
    worksheetes.set_column('Y:Y', 15)
    worksheetes.set_column('Z:Z', 15)
    worksheetes.set_column('AA:AA',15,format0dp)
    worksheetes.set_column('AB:AB',15,format0dp)
    
    # Hide zeros to make sheet look neater and easier to identify where there are data missing

    worksheetes.hide_zero()

#  Freeze top row and 1st 3 columns
    
    worksheetes.freeze_panes(1, 3)      

    wspwd = anvil.secrets.get_secret('protect_workbook')
    # Protect worksheets (except worksheetes where we need to allow users to sort, filter and add rows to table - these cant been done on a protected sheet.)
    worksheetbt.protect(wspwd,optionsbt)
#    worksheetau.protect(wspwd)

#  Write and close    
    writer.close()
    xlsx_data = content.getvalue()
    print('past dfes.to_excel')

# Create anvil blobmedia object

    abm = anvil.BlobMedia(content=xlsx_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = f"Estate_data_capture for Estate {entity} - created {dt_str}")
    ret['abm'] = abm
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
  return ret
  
def write_project_details_to_excel(df_innos, df_inspv, df_insth, entity, partner, client):
  ret  = {'ef': 0, 'em' : '', 'abm' : ''}
  import datetime as dt
  now               = datetime.now()
  dt_str            = now.strftime("%D/%M/%Y")  
  try:  
    dfnonsolar      = df_innos.copy()
    dfsolarpv       = df_inspv.copy()
    dfsolarthermal  = df_insth.copy()

    # Remove the baselined column (not wanted for upload)
    dfnonsolar.drop('baselined', axis=1, inplace=True)
    dfsolarpv.drop('baselined', axis=1, inplace=True)
    dfsolarthermal.drop('baselined', axis=1, inplace=True)
    
    print('In write_project_details_to_excel')
    
# Where delivery date is set to default base date for Excel (1/1/1900) replace with Null so output sheet looks cleaner

    sd        = dt.datetime(1900,1,1)
    bdate     = dt.datetime.date(sd)
    dfnonsolar.loc[dfnonsolar['delivery_date'] == bdate, 'delivery_date'] = ''
    dfsolarpv.loc[dfsolarpv['delivery_date'] == bdate, 'delivery_date'] = ''
    dfsolarthermal.loc[dfsolarthermal['delivery_date'] == bdate, 'delivery_date'] = ''
    content = io.BytesIO()
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(content, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')
# Convert the dataframe to an XlsxWriter Excel object.
    dfnonsolar.to_excel(writer, sheet_name='Non Solar', index=False)
    dfsolarpv.to_excel(writer, sheet_name='Solar PV', index=False)
    dfsolarthermal.to_excel(writer, sheet_name='Solar Thermal', index=False)
# Get the xlsxwriter objects from the dataframe writer object.
    workbook  = writer.book
    worksheetns = writer.sheets['Non Solar']
    worksheetsp = writer.sheets['Solar PV']
    worksheetst = writer.sheets['Solar Thermal']
    worksheetau = workbook.add_worksheet('Auth')
    worksheetky = workbook.add_worksheet('Key') 
# Write the auth sheets
    ret_mess = write_auth_sheets(workbook, worksheetau, worksheetky, partner, client, entity)
    ef       = ret_mess['ef']
    em       = ret_mess['em']
    if ef == 2:
      ret['ef'] = ef
      ret['em'] = em
      return ret 

# Set header formats
    #cols 0-4
    header_format1 = workbook.add_format({'bold': True, 'text_wrap': True, 'bg_color':'#c4d79b'})
    header_format1.set_align('center')
    header_format1.set_align('vcenter')
    #cols 5 -> 
    header_format2 = workbook.add_format({'bold': True, 'text_wrap': True, 'bg_color':'#0070c0', 'font_color':'#ffffff'})
    header_format2.set_align('center')
    header_format2.set_align('vcenter')  

# Set capex format
    format0dpc = workbook.add_format({'num_format': '"£" #,##0'})
# Set numeric fields with no decimal places or currency symbol but 100;s commas
    format0dpcthcom = workbook.add_format({'num_format': '#,##0'})
    
# Set date formats    
    formatdict = {'num_format':'dd/mm/yyyy'}
    fmt = workbook.add_format(formatdict)
    worksheetns.set_column('M:M', None, fmt)
    worksheetsp.set_column('J:J', None, fmt)
    worksheetst.set_column('J:J', None, fmt)
# Set percentage format
    percentage = workbook.add_format({'num_format': '0.0%'})
    worksheetns.set_column('I:I',None , percentage )
# Set up validation of columns 
    # Non solar
#    worksheetns.data_validation('E2:E50000', {'validate': 'list','source'  : ["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FURTHER IMPV","ASSESSED/NV"]})
    worksheetns.data_validation('F2:F50000', {'validate': 'list','source'  : ["Concept","Feasibility","Business Case","Procurement","Abandoned","Benefits Realisation"]})
    worksheetns.data_validation('G2:G50000', {'validate': 'list','source'  : ["GAS","ELEC"]})
    worksheetns.data_validation('H2:H50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetns.data_validation('I2:I50000', {'validate': 'decimal','criteria' : 'between', 'minimum' : -1, 'maximum' : 1 })
    worksheetns.data_validation('J2:J50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetns.data_validation('K2:K50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetns.data_validation('L2:L50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetns.data_validation('M2:M50000', {'validate': 'date', 'criteria' : 'between', 'minimum' : date(1900,1,1), 'maximum' : date(2052, 12, 31) })

    # Solar PV
#    worksheetsp.data_validation('E2:E50000', {'validate': 'list','source'  : ["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FURTHER IMPV","ASSESSED/NV"]})
    worksheetsp.data_validation('F2:F50000', {'validate': 'list','source'  : ["Concept","Feasibility","Business Case","Procurement","Abandoned","Benefits Realisation"]})
    worksheetsp.data_validation('G2:G50000', {'validate': 'list','source'  : ["GAS","ELEC"]})
    worksheetsp.data_validation('H2:H50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetsp.data_validation('I2:I50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetsp.data_validation('J2:J50000', {'validate': 'date', 'criteria' : 'between', 'minimum' : date(1900,1,1), 'maximum' : date(2052, 12, 31) })
    worksheetsp.data_validation('K2:K50000', {'validate': 'list','source'  : ["Flat Roof (Membrane)","Flat Roof (Deck)","Profile sheet","Concrete Tile","Clay tile","Slate"]})
    worksheetsp.data_validation('L2:L50000', {'validate': 'list','source'  : ["Flat", "Pitched"]})
    worksheetsp.data_validation('M2:M50000', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    worksheetsp.data_validation('N2:N50000', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    worksheetsp.data_validation('O2:O50000', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    
    # Solar thermal
#    worksheetst.data_validation('E2:E50000', {'validate': 'list','source'  : ["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FURTHER IMPV","ASSESSED/NV"]})
    worksheetst.data_validation('F2:F50000', {'validate': 'list','source'  : ["Concept","Feasibility","Business Case","Procurement","Abandoned","Benefits Realisation"]})
    worksheetst.data_validation('G2:G50000', {'validate': 'list','source'  : ["GAS","ELEC"]})
    worksheetst.data_validation('H2:H50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetst.data_validation('I2:I50000', {'validate': 'decimal','criteria': '>', 'value': 0 })
    worksheetst.data_validation('J2:J50000', {'validate': 'date', 'criteria' : 'between', 'minimum' : date(1900,1,1), 'maximum' : date(2052, 12, 31) })
    worksheetst.data_validation('K2:K50000', {'validate': 'decimal','criteria': '>=', 'value': 0 })
    worksheetst.data_validation('L2:L50000', {'validate': 'decimal','criteria': '>=', 'value': 0 })    
    
#   Protect the worksheet without password so user can unprotect to filter and sort
    options = {
    'format_cells':          True,
    'format_columns':        True,
    'format_rows':           True,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           True,
    'select_locked_cells':   True,
    'sort':                  True,
    'autofilter':            True,
    'pivot_tables':          False,
    'select_unlocked_cells': True,
    }
    worksheetns.protect('',options)
    worksheetsp.protect('',options)
    worksheetst.protect('',options)
  
#   Set default of cells as unlocked
    unlocked = workbook.add_format({'locked': False})

#   Unprotect the cells where we allow the user to change values
    worksheetns.unprotect_range('F2:M50000')
    worksheetsp.unprotect_range('F2:O50000')
    worksheetst.unprotect_range('F2:L50000')

#   Shade the protected columns
    cell_format_prot = workbook.add_format()
    cell_format_prot.set_bg_color('#e0e0eb')
    worksheetns.set_column('A:E',10,cell_format_prot)
    worksheetsp.set_column('A:E',10,cell_format_prot)
    worksheetst.set_column('A:E',10,cell_format_prot)

#   Set column widths
    # Non solar
    worksheetns.set_column('A:A', 14)  
    worksheetns.set_column('B:B', 45) 
    worksheetns.set_column('C:C', 35)
    worksheetns.set_column('D:D', 40)
    worksheetns.set_column('E:E', 15)
    worksheetns.set_column('F:F', 25) 
    worksheetns.set_column('G:G', 8) 
    worksheetns.set_column('H:H', 20) 
    worksheetns.set_column('I:I', 20)
    worksheetns.set_column('J:J', 12)
    worksheetns.set_column('K:K', 30, format0dpcthcom)
    worksheetns.set_column('L:L', 15, format0dpc)
    worksheetns.set_column('M:M', 15)
    # Solar PV
    worksheetsp.set_column('A:A', 14)  
    worksheetsp.set_column('B:B', 45) 
    worksheetsp.set_column('C:C', 35)
    worksheetsp.set_column('D:D', 40)
    worksheetsp.set_column('E:E', 15)
    worksheetsp.set_column('F:F', 25) 
    worksheetsp.set_column('G:G', 8) 
    worksheetsp.set_column('H:H', 20) 
    worksheetsp.set_column('I:I', 15, format0dpc) 
    worksheetsp.set_column('J:J', 22)
    worksheetsp.set_column('K:K', 25)
    worksheetsp.set_column('L:L', 20)
    worksheetsp.set_column('M:M', 20, format0dpcthcom)
    worksheetsp.set_column('N:N', 20, format0dpcthcom)
    worksheetsp.set_column('O:O', 32, format0dpcthcom)
    # Solar Thermal   
    worksheetst.set_column('A:A', 14)  
    worksheetst.set_column('B:B', 45) 
    worksheetst.set_column('C:C', 35)
    worksheetst.set_column('D:D', 40)
    worksheetst.set_column('E:E', 15)
    worksheetst.set_column('F:F', 25) 
    worksheetst.set_column('G:G', 8) 
    worksheetst.set_column('H:H', 20) 
    worksheetst.set_column('I:I', 15, format0dpc) 
    worksheetst.set_column('J:J', 22)
    worksheetst.set_column('K:K', 24, format0dpcthcom)
    worksheetst.set_column('L:L', 37, format0dpcthcom)
    worksheetst.set_column('M:M', 20)

# Hide zeros to make sheet look neater and easier to identify where there are data missing

    worksheetns.hide_zero()
    worksheetsp.hide_zero()
    worksheetst.hide_zero()

# Re-write and format the column headers
    colnames = ['Building ID', 'Building name','Building type', 'Project type','Assessed','Status','Utility','Lifetime (yrs)','Saving %','Heat pump scop','Heat pump elec add kWh pa','CAPEX','Delivery date']
    for col_num, value in enumerate(colnames):
      if col_num <=4:  
        worksheetns.write(0, col_num , value, header_format1)
      else:
        worksheetns.write(0, col_num , value, header_format2)   

    colnames = ['Building ID', 'Building name','Building type', 'Project type','Assessed','Status','Utility','Lifetime (yrs)','CAPEX','Delivery date','Solar roof type','Solar angle','Solar area m2','Solar KW peak','Corrected annual gen kWh']
    for col_num, value in enumerate(colnames):
      if col_num <=4:  
        worksheetsp.write(0, col_num , value, header_format1)
      else:
        worksheetsp.write(0, col_num , value, header_format2)

    colnames = ['Building ID', 'Building name','Building type', 'Project type','Assessed','Status','Utility','Lifetime (yrs)','CAPEX','Delivery date','Solar area m2','Corrected annual gen kWh']
    for col_num, value in enumerate(colnames):
      if col_num <=4:  
        worksheetst.write(0, col_num , value, header_format1)
      else:
        worksheetst.write(0, col_num , value, header_format2) 

#  Freeze top row and 1st 5 columns
    
    worksheetns.freeze_panes(1, 5)
    worksheetsp.freeze_panes(1, 5)
    worksheetst.freeze_panes(1, 5)
    
    writer.close()
    xlsx_data = content.getvalue()
    print('past df.to_excel')

# Create anvil blobmedia object

    abm = anvil.BlobMedia(content=xlsx_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = f"project_details_form_entity_{entity} - {dt_str}")
    ret['abm'] = abm
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
  return ret

def write_project_initialisation_to_excel(df_inass, entity, partner, client):
  ret  = {'ef': 0, 'em' : '', 'abm' : ''}
  import datetime as dt
  now               = datetime.now()
  dt_str            = now.strftime("%D/%M/%Y")  
  try:  
    dfin            = df_inass.copy()
    nr              = dfin.shape[0] + 1
    vrange          = f"D2:Z{nr}"
    
# Remove the 'Heat Pump (ELEC ADD)' column as not needed
    dfin.drop('Heat Pump (ELEC ADD)', axis=1, inplace=True)

# Convert Building ID to string format
    dfin['Building ID'] = dfin['Building ID'].apply(str)
    
    print('dfin = ')
    print(dfin.to_string())
    content = io.BytesIO()
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(content, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')
# Convert the dataframe to an XlsxWriter Excel object.
    dfin.to_excel(writer, sheet_name='Projects', index=False)

# Get the xlsxwriter objects from the dataframe writer object.
    workbook    = writer.book
    worksheetns = writer.sheets['Projects']
    worksheetau = workbook.add_worksheet('Auth')
    worksheetky = workbook.add_worksheet('Key') 
# Write the auth sheets
    ret_mess = write_auth_sheets(workbook, worksheetau, worksheetky, partner, client, entity)
    ef       = ret_mess['ef']
    em       = ret_mess['em']
    if ef == 2:
      ret['ef'] = ef
      ret['em'] = em
      return ret    
    header_format1 = workbook.add_format({'bold': True, 'text_wrap': True, 'bg_color':'#c4d79b'})
    header_format1.set_align('center')
    header_format1.set_align('vcenter')

    header_format2 = workbook.add_format({'bold': True, 'text_wrap': True, 'bg_color':'#0070c0', 'font_color':'#ffffff'})
    header_format2.set_align('center')
    header_format2.set_align('vcenter')    

    vrange_formateven = workbook.add_format({'bold': False, 'text_wrap': True})
    vrange_formateven.set_align('center')
    vrange_formateven.set_align('vcenter')
    vrange_formateven.set_top(3)
    vrange_formateven.set_left(0)
    vrange_formateven.set_right(0)
    vrange_formateven.set_bottom(0)
    
    vrange_formatodd  = workbook.add_format({'bold': False, 'text_wrap': True})
    vrange_formatodd.set_align('center')
    vrange_formatodd.set_align('vcenter')
    vrange_formatodd.set_top(6)
    vrange_formatodd.set_left(0)
    vrange_formatodd.set_right(0)
    vrange_formatodd.set_bottom(0)

    col1_formateven   = workbook.add_format({'bg_color':'#e0e0eb'})
    col1_formateven.set_align('right')
    col1_formateven.set_top(3)
    col1_formateven.set_left(0)
    col1_formateven.set_right(0)
    col1_formateven.set_bottom(0)

    col1_formatodd   = workbook.add_format({'bg_color':'#e0e0eb'})
    col1_formatodd.set_align('right')
    col1_formatodd.set_top(6)
    col1_formatodd.set_left(0)
    col1_formatodd.set_right(0)
    col1_formatodd.set_bottom(0)

    col23_formateven   = workbook.add_format({'bg_color':'#e0e0eb'})
    col23_formateven.set_align('left')
    col23_formateven.set_top(3)
    col23_formateven.set_left(0)
    col23_formateven.set_right(0)
    col23_formateven.set_bottom(0)

    col23_formatodd   = workbook.add_format({'bg_color':'#e0e0eb'})
    col23_formatodd.set_align('left')
    col23_formatodd.set_top(6)
    col23_formatodd.set_left(0)
    col23_formatodd.set_right(0)
    col23_formatodd.set_bottom(0)

    vrange_formateven_lr = workbook.add_format({'bold': False, 'text_wrap': True})
    vrange_formateven_lr.set_align('center')
    vrange_formateven_lr.set_align('vcenter')
    vrange_formateven_lr.set_top(3)
    vrange_formateven_lr.set_left(0)
    vrange_formateven_lr.set_right(0)
    vrange_formateven_lr.set_bottom(5)
    
    vrange_formatodd_lr  = workbook.add_format({'bold': False, 'text_wrap': True})
    vrange_formatodd_lr.set_align('center')
    vrange_formatodd_lr.set_align('vcenter')
    vrange_formatodd_lr.set_top(6)
    vrange_formatodd_lr.set_left(0)
    vrange_formatodd_lr.set_right(0)
    vrange_formatodd_lr.set_bottom(5)

    col1_formateven_lr   = workbook.add_format({'bg_color':'#e0e0eb'})
    col1_formateven_lr.set_align('right')
    col1_formateven_lr.set_top(3)
    col1_formateven_lr.set_left(0)
    col1_formateven_lr.set_right(0)
    col1_formateven_lr.set_bottom(5)

    col1_formatodd_lr   = workbook.add_format({'bg_color':'#e0e0eb'})
    col1_formatodd_lr.set_align('right')
    col1_formatodd_lr.set_top(6)
    col1_formatodd_lr.set_left(0)
    col1_formatodd_lr.set_right(0)
    col1_formatodd_lr.set_bottom(5)

    col23_formateven_lr   = workbook.add_format({'bg_color':'#e0e0eb'})
    col23_formateven_lr.set_align('left')
    col23_formateven_lr.set_top(3)
    col23_formateven_lr.set_left(0)
    col23_formateven_lr.set_right(0)
    col23_formateven_lr.set_bottom(5)

    col23_formatodd_lr   = workbook.add_format({'bg_color':'#e0e0eb'})
    col23_formatodd_lr.set_align('left')
    col23_formatodd_lr.set_top(6)
    col23_formatodd_lr.set_left(0)
    col23_formatodd_lr.set_right(0)
    col23_formatodd_lr.set_bottom(5)
    
# Overwrite both the value and the format of each header cell
    for col_num, value in enumerate(dfin.columns.values):
      if col_num <=2:  
        worksheetns.write(0, col_num , value, header_format1)
      else:
        worksheetns.write(0, col_num , value, header_format2)

# Write the main dataframe and re-format columns 3 onwards
    nr          = nr - 1
    for index, row in dfin.iterrows():
      rv        = row.values
      xl_row    = index + 1
      print(f"xl_row = {xl_row} nr = {nr}")
    #  if xl_row > 2:
    #    print(f"Index = {index}, xl_row = {xl_row}")  
    #    print(f"rv = {rv}")
      for col_num, value in enumerate(rv):
        if col_num == 0:
          if xl_row == 2:
            worksheetns.write(xl_row, col_num , value, col1_formateven) 
          if xl_row  % 2:
            #odd
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, col1_formatodd_lr)
            else:
              worksheetns.write(xl_row, col_num , value, col1_formatodd)                   
          else:
            #even
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, col1_formateven_lr) 
            else:
              worksheetns.write(xl_row, col_num , value, col1_formateven) 

        if col_num == 1 or col_num == 2:
          if xl_row == 2:
            worksheetns.write(xl_row, col_num , value, col23_formateven) 
          if xl_row  % 2:
            #odd 
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, col23_formatodd_lr)
            else:
              worksheetns.write(xl_row, col_num , value, col23_formatodd)
          else:
            #even 
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, col23_formateven_lr)
            else:
              worksheetns.write(xl_row, col_num , value, col23_formateven) 
        
        if col_num >= 3:
          if xl_row == 2:
            worksheetns.write(xl_row, col_num , value, vrange_formateven) 
          if xl_row  % 2:
            #odd
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, vrange_formatodd_lr)
            else:
              worksheetns.write(xl_row, col_num , value, vrange_formatodd)
          else:
            #even 
            if xl_row == nr:
              worksheetns.write(xl_row, col_num , value, vrange_formateven_lr)
            else:
              worksheetns.write(xl_row, col_num , value, vrange_formateven)
  
# Set up validation of columns 
    nr          = nr + 1
    worksheetns.data_validation(vrange, {'validate': 'list','source'  : ["FIRM","LIKELY","POSSIBLE","POTENTIAL","IN PLACE","FTHR IMPV","ASSESSED/NV"]})

#   Protect the worksheet without password so user can unprotect to filter and sort
    options = {
    'format_cells':          True,
    'format_columns':        True,
    'format_rows':           True,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           True,
    'select_locked_cells':   True,
    'sort':                  True,
    'autofilter':            True,
    'pivot_tables':          False,
    'select_unlocked_cells': True,
    }
    worksheetns.protect('',options)
  
#   Set default of cells as unlocked
    unlocked = workbook.add_format({'locked': False})

#   Unprotect the cells where we allow the user to change values
    worksheetns.unprotect_range(vrange)

#   Shade the protected columns
    cell_format_prot = workbook.add_format()
    cell_format_prot.set_bg_color('#e0e0eb')
    worksheetns.set_column('A:C',10,cell_format_prot)

#  Set column widths
    
    worksheetns.set_column('A:A', 14)
    worksheetns.set_column('B:B', 45) 
    worksheetns.set_column('C:C', 100)
    worksheetns.set_column('D:Z', 20)

# Conditional formating background colour depending on dropdown contents of cells

    format_firm = workbook.add_format({'bg_color':'#00b0f0'})
    format_like = workbook.add_format({'bg_color':'#92d050'})
    format_poss = workbook.add_format({'bg_color':'#ffe699'})
    format_pote = workbook.add_format({'bg_color':'#ffc000'})
    format_inpl = workbook.add_format({'bg_color':'#ffffff'})
    format_impv = workbook.add_format({'bg_color':'#a6a6a6','font_color':'#f1c025'})
    format_asnv = workbook.add_format({'bg_color':'#002060','font_color':'#ffffff'})

    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'FIRM',
									                         'format' : format_firm}) 
    
    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'LIKELY',
									                         'format' : format_like}) 
    
    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'POSSIBLE',
									                         'format' : format_poss}) 

    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'POTENTIAL',
									                         'format' : format_pote})

    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'IN PLACE',
									                         'format' : format_inpl}) 

    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'FTHR IMPV',
									                         'format' : format_impv}) 

    worksheetns.conditional_format(vrange,{'type': 'text',
                                           'criteria' : 'containing',
									                         'value' : 'ASSESSED/NV',
									                         'format' : format_asnv}) 

# Hide zeros to make sheet look neater and easier to identify where there are data missing

    worksheetns.hide_zero()

#  Freeze top row and 1st 3 columns
    
    worksheetns.freeze_panes(1, 3)    
    
    writer.close()
    xlsx_data = content.getvalue()
    print('past df.to_excel')

# Create anvil blobmedia object

    abm = anvil.BlobMedia(content=xlsx_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = f"project_assessment_form_entity_{entity} - {dt_str}")
    ret['abm'] = abm
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
  return ret


def write_all_results_to_excel(df_projects, df_estate, df_estatesummary, df_solarsummary, df_bec, df_wfallps, entity_name, user_name, dt_str):
# Writes the dataframes containing results to Excel tables each on a separate worksheet. The worksheets are then protected to prevent overwriting.
  ret  = {'ef': 0, 'em' : '', 'abm' : ''}
  print('In write all results entity_name:')
  print(entity_name)
  import datetime as dt
  now               = datetime.now()
  dt_str            = now.strftime("%D/%M/%Y")  
  try:  
    dfprojects      = df_projects.copy()
    dfestate        = df_estate.copy()
    dfestatesummary = df_estatesummary.copy()
    dfsolarsummary  = df_solarsummary.copy()
    dfbec           = df_bec.copy()
    dfwfallps       = df_wfallps.copy()
    
# Where delivery date is set to default base date for Excel (1/1/1900) replace with Null so output sheet looks cleaner

    sd        = dt.datetime(1900,1,1)
    bdate     = dt.datetime.date(sd)
    dfprojects.loc[dfprojects['delivery_date_mode'] == bdate, 'delivery_date_mode'] = ''

    content = io.BytesIO()
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(content, engine = 'xlsxwriter',date_format = 'dd/mm/yyyy')
#    writer = pd.ExcelWriter(content, engine = 'xlsxwriter')
# Convert the dataframe to an XlsxWriter Excel object.
    dfprojects.to_excel(writer, sheet_name='Project_savings', startrow=1, header = False, index=False)
    dfestate.to_excel(writer, sheet_name='Estate',startrow=1, header = False, index=False)
    dfestatesummary.to_excel(writer, sheet_name='Estate_summary',startrow=1, header = False, index=False)
    dfsolarsummary.to_excel(writer, sheet_name='Solar_summary', startrow=1,header = False, index=False)
    dfbec.to_excel(writer, sheet_name='Build_energy_cost', startrow=1,header = False, index=False)
    dfwfallps.to_excel(writer, sheet_name='Waterfall_project_savings', startrow=1,header = False, index=False)
    
# Get the xlsxwriter objects from the dataframe writer object.
    workbook  = writer.book
    worksheetps = writer.sheets['Project_savings']
    worksheetes = writer.sheets['Estate']
    worksheetesum = writer.sheets['Estate_summary']
    worksheetssum = writer.sheets['Solar_summary']
    worksheetbec  = writer.sheets['Build_energy_cost']
    worksheetwfps = writer.sheets['Waterfall_project_savings']
    
    formatdict = {'num_format':'dd/mm/yyyy'}
    fmt = workbook.add_format(formatdict)
    worksheetps.set_column('O:O', None, fmt)
    
# Get the dimensions of the dataframe.
    (max_row_ps, max_col_ps) = dfprojects.shape
    (max_row_es, max_col_es) = dfestate.shape
    (max_row_esum, max_col_esum) = dfestatesummary.shape
    (max_row_ssum, max_col_ssum) = dfsolarsummary.shape
    (max_row_bec, max_col_bec) = dfbec.shape
    (max_row_wfps, max_col_wfps) = dfwfallps.shape
    
# Create a list of column headers, to use in add_table().
    column_settingsps = []
    for header in dfprojects.columns:
      column_settingsps.append({'header': header}) 
      
    column_settingses = []
    for header in dfestate.columns:
      column_settingses.append({'header': header}) 
      
    column_settingsesum = []
    for header in dfestatesummary.columns:
      column_settingsesum.append({'header': header}) 
      
    column_settingsssum = []
    for header in dfsolarsummary.columns:
      column_settingsssum.append({'header': header})
      
    column_settingsbec = []
    for header in dfbec.columns:
      column_settingsbec.append({'header': header})
      
    column_settingswfps = []
    for header in dfwfallps.columns:
      column_settingswfps.append({'header': header}) 
      
# Add the table.
#    worksheetps.add_table(0, 0, max_row_ps, max_col_ps - 1, {'columns': column_settingsps, 'data': dfprojects, 'name': 'Project_savings','banded_rows' : True})      
    worksheetps.add_table(0, 0, max_row_ps, max_col_ps - 1, {'columns': column_settingsps, 'name': 'Project_savings'}) 
    worksheetes.add_table(0, 0, max_row_es, max_col_es - 1, {'columns': column_settingses, 'name': 'Estate'})
    worksheetesum.add_table(0, 0, max_row_esum, max_col_esum - 1, {'columns': column_settingsesum, 'name': 'Estate_summary'})    
    worksheetssum.add_table(0, 0, max_row_ssum, max_col_ssum - 1, {'columns': column_settingsssum, 'name': 'Solar_summary'})
    worksheetbec.add_table(0, 0, max_row_bec, max_col_bec - 1, {'columns': column_settingsbec, 'name': 'Build_energy_cost'})
    worksheetwfps.add_table(0, 0, max_row_wfps, max_col_wfps - 1, {'columns': column_settingswfps, 'name': 'Waterfall_project_savings'})
    
#   Protect the worksheets with password so user can unprotect to filter and sort
    options = {
    'format_cells':          True,
    'format_columns':        True,
    'format_rows':           True,
    'insert_columns':        False,
    'insert_rows':           False,
    'insert_hyperlinks':     False,
    'delete_columns':        False,
    'delete_rows':           True,
    'select_locked_cells':   True,
    'sort':                  True,
    'autofilter':            True,
    'pivot_tables':          False,
    'select_unlocked_cells': True,
    }
    worksheetps.protect('',options)
    worksheetes.protect('',options)
    worksheetesum.protect('',options)    
    worksheetssum.protect('',options)
    worksheetbec.protect('',options)
    worksheetwfps.protect('',options)
    
# Write the dataframes

    writer.close()
    xlsx_data = content.getvalue()
    print('past df.to_excel')
#    worksheetps.add_table('A1:O1',{'data': dfprojects, 'name': 'Project_savings','banded_rows' : True})
    
    # Create anvil blobmedia object

    abm = anvil.BlobMedia(content=xlsx_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", name = f"OnePointFive - all results for {entity_name} - {dt_str}")
    ret['abm'] = abm
    return ret    
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef'] = 2
    ret['em'] = msg
  return ret 

def validate_forecast_actuals_buildings(conn, entity, entity_number, df):
  print('At top validate_forecast_actuals_buildings')
  # Validates buildings dataframe (df) from forecast or actuals upload workbook for entity with entity_number. Conn is the connection object to the Onepointfive database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}
    vm                   = ''
    df_in                = df.copy()
    df_original          = df.copy()

    nvw                  = 0
    nve                  = 0
    uprn_mess            = ''
    nuprnerrs            = 0
    num_rows             = df_in.shape[0]

    # Find duplicated uprns and produce warnings
    
    ids                 = df_in["uprn"]
    ddf                 = df_in
    tdf                 = ddf[ids.isin(ids[ids.duplicated()])].sort_values(by="uprn")
    num_dup_uprn        = tdf.shape[0]
    if num_dup_uprn > 0:
      dups                = tdf[["uprn", "building_name"]]
      dups_noi            = dups.to_string(index=False)
      dup_mess            = f"-----WARNING - {num_dup_uprn} occurrences of duplicate UPRNs have been found - review advised\n"
      dup_mess            = dup_mess + dups_noi
      vm                  = f" {vm + dup_mess}\n"

    nvw                   = num_dup_uprn
    nve                   = nuprnerrs
    print('==In validate buildings==')
    print('nvw and nve')
    print(nvw)
    print(nve)
    validation['validated_df']        = df_original
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def validate_forecast_actuals_usage_v2(conn, entity, entity_number, df, en_type):
  print('At top validate_forecast_actuals_usage_v2')
  print(en_type)
  # Validates the Elec, Gas, Oil, LPG, Solar PV and Solar Thermal usage dataframes (df) from forecast or actuals upload workbook for entity with entity_number. Conn is the connection object to the Onepointfive database.
  # Assumes uprn and building name columns have been removed and columns are years.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}

    vm                   = ''
    df_in                = df.copy()

    nvw                  = 0
    nve                  = 0

    # Extract column headers as year string in format 'YYYY' eg: '2022'. '2023' etc.
    # Note at this point the column headers are strings 

    col_dates            = list(df_in)
    set_dates            = set(col_dates)

    # Validate year columns are unique

    if len(col_dates) != len(set_dates):
      validation['ef']   = 2
      validation['em']   = "****Error - duplicate years found - please resolve and re-submit"
      validation['nve']  = 1
      return validation

    # Validate there are no duplicate uprns
    
    set_uprn             = set(df_in['uprn'])
    list_uprn            = df_in['uprn']

    if len(set_uprn) != len(list_uprn):
      validation['ef']   = 2
      validation['em']   = "****Error - duplicate uprns found - please resolve and re-submit"
      validation['nve']  = 1
      return validation    

    # Validate values are numeric and > zero

  #  for ic = 
    nvw                  = 0
    nve                  = 0
    num_rows             = df_in.shape[0]

    for year, values in df.iteritems():
      ir = 1

      for v in values:
        ir = ir + 1
        if pd.isna(v):
          vm  = vm + f"Invalid non numeric usage value {v} in year {year} row {ir}\n"
          nve = nve + 1
        elif  type(v) == str:
          vm  = vm + f"Invalid non numeric usage value {v} in year {year} row {ir}\n"
          nve = nve + 1          
          
        elif v < 0:
            vm  = vm + f"Invalid negative usage value {v} in year {year} row {ir}\n"
            nve = nve + 1
        
    validation['validated_df']        = df_in # uprn and building stripped - just data and date col headings
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def validate_forecast_actuals_cost_v2(conn, entity, entity_number, df, en_type):
  print('At top validate_forecast_actuals_cost_v2')
  print(en_type)
  # Validates the Elec, Gas, Oil, LPG, Solar PV and Solar Thermal cost dataframes (df) from forecast or actuals cost upload workbook for entity with entity_number. Conn is the connection object to the Onepointfive database.
  # Assumes uprn and building name columns have been removed and columns are years.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}

    vm                   = ''
    df_in                = df.copy()

    nvw                  = 0
    nve                  = 0

    # Extract column headers as year string in format 'YYYY' eg: '2022'. '2023' etc.
    # Note at this point the column headers are strings 

    col_dates            = list(df_in)
    set_dates            = set(col_dates)

    # Validate year columns are unique

    if len(col_dates) != len(set_dates):
      validation['ef']   = 2
      validation['em']   = "****Error - duplicate years found - please resolve and re-submit"
      validation['nve']  = 1
      return validation

    # Validate there are no duplicate uprns
    
    set_uprn             = set(df_in['uprn'])
    list_uprn            = df_in['uprn']

    if len(set_uprn) != len(list_uprn):
      validation['ef']   = 2
      validation['em']   = "****Error - duplicate uprns found - please resolve and re-submit"
      validation['nve']  = 1
      return validation    
      
   
    # Validate values are numeric and > zero

  #  for ic = 
    nvw                  = 0
    nve                  = 0
    num_rows             = df_in.shape[0]

    for year, values in df.iteritems():
      ir = 1

      for v in values:
        ir = ir + 1
        if pd.isna(v):
          vm  = vm + f"Invalid non numeric cost value {v} in year {year} row {ir}\n"
          nve = nve + 1
        elif  type(v) == str:
          vm  = vm + f"Invalid non numeric cost value {v} in year {year} row {ir}\n"
          nve = nve + 1          
          
        elif v < 0:
            vm  = vm + f"Invalid negative cost value {v} in year {year} row {ir}\n"
            nve = nve + 1
        
    validation['validated_df']        = df_in # uprn and building stripped - just data and date col headings
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def convert_datestrings_to_ints(ms, syr):
# Convert a list on 3 letter month codes (e.g. 'Jan', 'JAN','jan') to equivalent integers 1-12
  ret  = {'ef':0, 'em':'', 'intmonths':[], 'intyears':[]}
  valc = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
  im   = []
  iy   = []
  msin = list(map(lambda x: x.lower(), ms ))
  #print('lower ms')
  #print(msin)
  try:
    for s in msin:
      fm  = s
      mi  = valc.index(fm) + 1 
      im.append(mi)
      
    ret['intmonths'] = im

    for s in syr:
      fm  = s
      yi  = int(fm) + 2000
      iy.append(yi)
      
    ret['intyears'] = iy
    return ret
 
  except Exception as e:
    ret['ef']  = 2
    ret['em']  = "Error - input contains either an invalid 3-letter month code or a year string that cannot be converted to int"
    return ret

@anvil.server.callable
def validate_forecast_actuals_usage(conn, entity, entity_number, df, en_type):
  print('At top validate_forecast_actuals_usage')
  print(en_type)
  # Validates the Elec, Gas, Oil, LPG, Solar PV and Solar Thermal usage dataframes (df) from forecast or actuals upload workbook for entity with entity_number. Conn is the connection object to the Onepointfive database.
  try:
    validation           = {'ef':0,'em':'Validation completed successfully','validated_df':'','validation_messages':'','nvw':0,'nve':0}

    vm                   = ''
    df_in                = df.copy()

    nvw                  = 0
    nve                  = 0

    # Extract column headers as string dates in format 'Mmm-yy' eg: 'Oct-22'. 'May-24' etc.
    # Note at this point the column headers are strings 

    col_dates            = list(df)
    set_dates            = set(col_dates)

    # Validate month/year combos are unique

    if len(col_dates) != len(set_dates):
      validation['ef']   = 2
      validation['em']   = "****Error - duplicate dates found - please resolve and re-submit"
      validation['nve']  = 1
      return validation
   
    # Validate values are numeric and > zero

  #  for ic = 
    nvw                  = 0
    nve                  = 0
    num_rows             = df_in.shape[0]
    
 #   print('In validate forecast usage')
 #   print('DF with uprn and building name removed')
 #   print(df_in.to_string())
    
    
   # for index, row in df_in.iterrows():

    validation['validated_df']        = df_in # uprn and building stripped - just data and date col headings
    validation['validation_messages'] = vm
    validation['nvw']                 = nvw
    validation['nve']                 = nve
    return validation
  
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    validation['ef'] = 2
    validation['em'] = f"****ERROR - validation terminated with exception: \n {msg}"
    validation['validated_df']        = ''
    validation['validation_messages'] = ''
    validation['nvw']                 = 0
    validation['nve']                 = 0
    return validation

def convert_datestrings_to_ints(ms, syr):
# Convert a list on 3 letter month codes (e.g. 'Jan', 'JAN','jan') to equivalent integers 1-12
  ret  = {'ef':0, 'em':'', 'intmonths':[], 'intyears':[]}
  valc = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
  im   = []
  iy   = []
  msin = list(map(lambda x: x.lower(), ms ))
  #print('lower ms')
  #print(msin)
  try:
    for s in msin:
      fm  = s
      mi  = valc.index(fm) + 1 
      im.append(mi)
      
    ret['intmonths'] = im

    for s in syr:
      fm  = s
      yi  = int(fm) + 2000
      iy.append(yi)
      
    ret['intyears'] = iy
    return ret
 
  except Exception as e:
    ret['ef']  = 2
    ret['em']  = "Error - input contains either an invalid 3-letter month code or a year string that cannot be converted to int"
    return ret

@anvil.server.callable
def calc_project_energy_carbon_savings_v4(conn, entity_number):
  # Calculates the Project annual total energy saving, electricity saving, gas saving, oil saving, lpg saving, Project carbon saving (broken down into Scope 1, Scope 2 and Scope 3) and Project tonne CO2 lifetime cost.
  # Calculates annual energy cost savings (broken down to elec, gas, oil and lpg), Lifetime Tonnes CO2 equivalent (t CO2-e) saved per project, Average Annual Abatement (tCO2-e/Year) per project and 
  # Annual Cost of Abatement per Tonne per project (£/tCO2-e). These are calculated for all controlled buildings in entity with number entity_number.
  # NOTE: Savings for Solar PV and Solar Thermal are not calculated by use of a 'percent_saving' since the amount generated in kWh is already a saving against imported (carbon bearing) energy and can be subtracted.
    
  ret_mess = {'ef':0, 'em':''}
  
  try:
    with conn.cursor() as cursor:
      
      # Get building level details for all controlled buildings in entity
      sqlrd = f"SELECT uprn, data_year, baseline_annual_elec_kwh, baseline_annual_gas_kwh, baseline_annual_oil_kwh, baseline_annual_lpg_kwh, \
              baseline_elec_cost_per_kwh, baseline_gas_cost_per_kwh, baseline_oil_cost_per_kwh, baseline_lpg_cost_per_kwh \
              FROM raw_estate_data WHERE entity_number = {entity_number} AND under_control = \'Yes\';"
      cursor.execute(sqlrd)
      t_output_rd      = cursor.fetchall()
      keys             = ("uprn", "data_year", "baseline_annual_elec_kwh", "baseline_annual_gas_kwh", "baseline_annual_oil_kwh", "baseline_annual_lpg_kwh",\
                         "baseline_elec_cost_per_kwh", "baseline_gas_cost_per_kwh", "baseline_oil_cost_per_kwh", "baseline_lpg_cost_per_kwh")
      output_rd        = [dict(zip(keys, values)) for values in t_output_rd]
      
      # Loop through each building in estate and get required building data and emission factors
      num_builds           = len(output_rd)
      num_projs            = 0
      
      for bui in output_rd:
        uprn              = bui['uprn']
        data_year         = bui['data_year']
        annual_elec_kwh   = bui['baseline_annual_elec_kwh']
        annual_gas_kwh    = bui['baseline_annual_gas_kwh']
        annual_oil_kwh    = bui['baseline_annual_oil_kwh']
        annual_lpg_kwh    = bui['baseline_annual_lpg_kwh']
        elec_unit_cost    = bui['baseline_elec_cost_per_kwh']
        gas_unit_cost     = bui['baseline_gas_cost_per_kwh']
        oil_unit_cost     = bui['baseline_oil_cost_per_kwh']
        lpg_unit_cost     = bui['baseline_lpg_cost_per_kwh']        
        
        # Get Emission factors for building
        
        sqlemf = f"SELECT elec_consumed, gas_consumed, oil_consumed, lpg_consumed,  electricity_t_d,  electricity_t_d_wtt,  electricity_gen_wtt, overall_elec_scope_3, gas_wtt, heating_oil_wtt, lpg_wtt  FROM emission_factors WHERE data_year = {data_year};"
        cursor.execute(sqlemf)
        t_output_emf         = cursor.fetchall() 
        keys                 = ("elec_consumed", "gas_consumed", "oil_consumed", "lpg_consumed",  "electricity_t_d",  "electricity_t_d_wtt",  "electricity_gen_wtt", "overall_elec_scope_3", "gas_wtt", "heating_oil_wtt", "lpg_wtt")
        output_emf           = [dict(zip(keys, values)) for values in t_output_emf]
        
        emf                  = output_emf[0] 
        elec_consumed        = emf['elec_consumed']
        gas_consumed         = emf['gas_consumed']
        oil_consumed         = emf['oil_consumed']
        lpg_consumed         = emf['lpg_consumed']
        electricity_t_d      = emf['electricity_t_d']
        electricity_t_d_wtt  = emf['electricity_t_d_wtt']
        electricity_gen_wtt  = emf['electricity_gen_wtt']
        overall_elec_scope_3 = emf['overall_elec_scope_3']
        gas_wtt              = emf['gas_wtt']
        heating_oil_wtt      = emf['heating_oil_wtt']
        lpg_wtt              = emf['lpg_wtt']
        
        # Get all the projects for this building. Loop through the projects calculating energy and carbon savings and write these to the project_results table
        
        sqlpr = f"SELECT project_id, project_type_id, assessed, saving_percent, cost_capex_mode, salix_pf, utility, solar_pv_tot_zero_carb_gen_kwh, solar_thermal_corrected_ann_gen_kwh, hp_scop, hp_elec_add_kwh_pa FROM projects WHERE ((entity_number = {entity_number}) AND (uprn = {uprn}));"
        cursor.execute(sqlpr)
        t_output_prd         = cursor.fetchall()
        keys                 = ("project_id", "project_type_id", "assessed", "saving_percent", "cost_capex_mode", "salix_pf", "utility", "solar_pv_tot_zero_carb_gen_kwh", "solar_thermal_corrected_ann_gen_kwh", "hp_scop","hp_elec_add_kwh_pa")
        output_prd           = [dict(zip(keys, values)) for values in t_output_prd]
        
        num_projs            = num_projs + len(output_prd)
        
        for pr in output_prd:

          project_id              = pr['project_id']
          project_type_id         = pr['project_type_id']
          assessed                = pr['assessed']          
          saving_percent          = pr['saving_percent']  
          capex                   = pr['cost_capex_mode']
          capex                   = float(capex)
          
          salix_pf                = pr['salix_pf']
          salix_pf                = float(salix_pf)

          utility                 = pr['utility']
          solar_pv_saving         = pr['solar_pv_tot_zero_carb_gen_kwh']
          solar_thermal_saving    = pr['solar_thermal_corrected_ann_gen_kwh']
          hp_scop                 = pr['hp_scop']
          hp_elec_add_kwh_pa      = pr['hp_elec_add_kwh_pa']
          
          if assessed != 'ASSESSED/NV' and assessed != 'IN PLACE':
            proj_elec_savings  = 0
            proj_gas_savings   = 0
            proj_oil_savings   = 0
            proj_lpg_savings   = 0
            proj_elec_annual_cost_saving = 0
            proj_gas_annual_cost_saving = 0
            proj_oil_annual_cost_saving = 0
            proj_lpg_annual_cost_saving = 0
            project_cost_savings = 0
            project_financial_value = 0
            ROI_percent        = 0 
            project_energy_savings = 0
            elec_carbon        = 0
            gas_carbon         = 0
            project_carbon_savings = 0
            tonne_co2_lifetime_cost = 0
            lifetime_tonnes_CO2e = 0
            annual_abatement_cost_tCO2e = 0

            if project_type_id == 20:  #Solar PV
              proj_elec_savings      = solar_pv_saving
              proj_gas_saving        = 0

            elif project_type_id == 21:  # Solar Thermal
              if utility == 'ELEC':
                proj_elec_savings      = solar_thermal_saving
                proj_gas_savings       = 0
    
              if utility == 'GAS':
                proj_gas_savings       = solar_thermal_saving
                proj_elec_savings      = 0            

            elif (project_type_id == 14) or (project_type_id == 12): #Heat Pump (GAS Saving)  or Heat Network

              proj_gas_savings       = annual_gas_kwh * saving_percent
              # Calc negative savings on electricity
              if hp_elec_add_kwh_pa != 0:
                proj_elec_savings    = -(abs(hp_elec_add_kwh_pa)) #If a value has been provided for hp_elec_add_kwh_pa use its negative absolute value
              elif hp_scop != 0:
                proj_elec_savings    = -proj_gas_savings/abs(hp_scop) # If hp_scop provided but no hp_elec_add_kwh_pa then calculate using negative absolute hp_scop
              else:
                proj_elec_savings    = -proj_gas_savings/3 
            
            else:
              if utility == 'ELEC':
                proj_elec_savings      = annual_elec_kwh * saving_percent
                proj_gas_savings       = 0
              if utility == 'GAS':
                proj_gas_savings       = annual_gas_kwh * saving_percent
                proj_elec_savings      = 0
                
              proj_oil_savings       = annual_oil_kwh * saving_percent
              proj_lpg_savings       = annual_lpg_kwh * saving_percent
              
            project_energy_savings   = proj_elec_savings + proj_gas_savings + proj_oil_savings + proj_lpg_savings

            # Calculate annual energy cost savings

            proj_elec_annual_cost_saving = proj_elec_savings * elec_unit_cost
            proj_gas_annual_cost_saving  = proj_gas_savings * gas_unit_cost
            proj_oil_annual_cost_saving  = proj_oil_savings * oil_unit_cost
            proj_lpg_annual_cost_saving  = proj_lpg_savings * lpg_unit_cost 
        
            project_cost_savings         = proj_elec_annual_cost_saving + proj_gas_annual_cost_saving + proj_oil_annual_cost_saving + proj_lpg_annual_cost_saving  
            project_financial_value      = project_cost_savings * decimal.Decimal(salix_pf) #Not adjusted for inflation
            
            # Calculate Return on investment percentage
            
            if capex > 0:
              ROI_percent                = ((project_financial_value - decimal.Decimal(capex)) / decimal.Decimal(capex) ) * 100
            else:
              ROI_percent                = 0            

            # Emission factors are in kgCO2 per kwh. Divide by 1000 to derive Tonnes CO2 equivalent.
  
            elec_co2             = (elec_consumed  * proj_elec_savings)/1000
            gas_co2              = (gas_consumed  *  proj_gas_savings)/1000
            oil_co2              = (oil_consumed  *  proj_oil_savings)/1000
            lpg_co2              = (lpg_consumed  *  proj_lpg_savings)/1000
  
            gas_wtt_scope_3      = (proj_gas_savings *  gas_wtt)/1000
            elec_t_d_scope_3     = (proj_elec_savings *  electricity_t_d)/1000
            elec_wtt_t_d_scope_3 = (proj_elec_savings * electricity_t_d_wtt)/1000
            elec_wtt_gen_scope_3 = (proj_elec_savings * electricity_gen_wtt)/1000
            oil_wtt_scope_3      = (proj_oil_savings * heating_oil_wtt)/1000
            lpg_wtt_scope_3      = (proj_lpg_savings * lpg_wtt)/1000
      
            total_scope_1        = gas_co2 + oil_co2 + lpg_co2
            total_scope_2        = elec_co2
            total_scope_3        = gas_wtt_scope_3 + elec_t_d_scope_3 + elec_wtt_t_d_scope_3 + elec_wtt_gen_scope_3 + oil_wtt_scope_3 + lpg_wtt_scope_3
      
            total_co2_tco2e      = total_scope_3 + total_scope_2 + total_scope_1	#Also equals average annual abatement when emission factors assumed equal over lifetime
                                                                                  # NOTE: saved in database as carbon_savings

            lifetime_tonnes_CO2e = total_co2_tco2e * decimal.Decimal(salix_pf)            
        
            if (total_co2_tco2e * decimal.Decimal(salix_pf)) > 0:
              tonne_co2_lifetime_cost	= decimal.Decimal(capex)/(total_co2_tco2e * decimal.Decimal(salix_pf))
            else:
              tonne_co2_lifetime_cost	= 0
 
            # Calculate the annual abatement cost (£/tCO2e)
            
            if salix_pf > 0:
              annual_abatement_cost       = (decimal.Decimal(capex)/decimal.Decimal(salix_pf)) - project_cost_savings
            else:
              annual_abatement_cost       = 0

            if total_co2_tco2e > 0:
              annual_abatement_cost_tCO2e = annual_abatement_cost / total_co2_tco2e # Annual abatement cost per Tonne CO2
            else:
              annual_abatement_cost_tCO2e = 0
              
            # Delete the record for this project in project_results table and re-insert with latest values calculated here
            
            sqldp = f"DELETE FROM project_results WHERE project_id = {project_id};"
            cursor.execute(sqldp)  
            conn.commit()

            sqlpri1 = f"INSERT INTO project_results (project_id, project_type_id, uprn, entity_number, energy_savings, gas_savings, electric_savings, oil_savings, lpg_savings, carbon_savings, tonne_co2_lifetime_cost, scope_1_savings, scope_2_savings, scope_3_savings, hp_elec_add_kwh_pa, \
            proj_elec_annual_cost_saving, proj_gas_annual_cost_saving, proj_oil_annual_cost_saving, proj_lpg_annual_cost_saving, project_cost_savings, project_financial_value, ROI_percent, lifetime_tonnes_CO2e, annual_abatement_cost_tCO2e) VALUES ("
            sqlpri2 = f"{project_id}, {project_type_id}, {uprn}, {entity_number}, {project_energy_savings}, {proj_gas_savings}, {proj_elec_savings},{proj_oil_savings}, {proj_lpg_savings}, {total_co2_tco2e},{tonne_co2_lifetime_cost}, {total_scope_1}, {total_scope_2}, {total_scope_3},{hp_elec_add_kwh_pa},\
            {proj_elec_annual_cost_saving}, {proj_gas_annual_cost_saving}, {proj_oil_annual_cost_saving}, {proj_lpg_annual_cost_saving}, {project_cost_savings}, {project_financial_value}, {ROI_percent}, {lifetime_tonnes_CO2e}, {annual_abatement_cost_tCO2e});"
            sqlpri  = f"{sqlpri1} {sqlpri2}"
            
            cursor.execute(sqlpri)
            conn.commit
            #ret_mess['ef'] = 0
            #ret_mess['em'] = sqlpri
    return ret_mess     
          
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
    return ret_mess

@anvil.server.callable
def calc_project_energy_carbon_savings_v5_PC01(conn, entity_number):
  # Calculates: -
  #  - Project annual total energy savings (broken down into electricity saving, gas saving, oil saving, lpg saving) 
  #  - Lifetime Tonnes CO2 equivalent (t CO2-e) saved per project
  #  - Project lifetime cost
  #  - Project energy cost savings
  #  - Project financial value
  #  - Project ROI
  #  - Average Annual Abatement (tCO2-e/Year) per project 
  #  - Annual Cost of Abatement per Tonne per project (£/tCO2-e). 
  
  # These are calculated for all projects (except IN PLACE asn ASSESSED_NV) for all controlled buildings in entity with number entity_number.
  
  # NOTES: 
  # 1 - Savings for Solar PV and Solar Thermal are not calculated by use of a 'percent_saving' since the amount generated in kWh is already a saving against imported (carbon bearing) energy and can be subtracted.
  # 2 = From version 5 the lifetime carbon savings are calculated by summing the annual carbon savings over the lifetime (salix_persistence factor) of the project starting at the delivery date of the project.
  #     The annual carbon savings for each year are calulated by multiplying the energy savings for that year by the emission factors for that year. For full explanation of Methods
  #     see 'OnePointFive - method definitions - Lifetime Tonnes CO2 e, abatement costs per tonne, annual abatement'
  
  ret_mess     = {'ef':0, 'em':''}
  ret_ltc_mess = {'ef':0, 'em':0, 'total_scope_1':0, 'total_scope_2':0, 'total_scope_3':0}
  try:
    with conn.cursor() as cursor:
      print('calc_project_energy_carbon_savings_v5_PC01 has been called %%%%%')
      # Create a dataframe from the emission_factors table containing the factors required across all years
      sqlemf = f"SELECT data_year,elec_consumed, gas_consumed, oil_consumed, lpg_consumed,  electricity_t_d,  electricity_t_d_wtt,  electricity_gen_wtt, overall_elec_scope_3, gas_wtt, heating_oil_wtt, lpg_wtt  FROM emission_factors ;"
      cursor.execute(sqlemf)
      t_output_emf         = cursor.fetchall() 
      keys                 = ("data_year", "elec_consumed", "gas_consumed", "oil_consumed", "lpg_consumed",  "electricity_t_d",  "electricity_t_d_wtt",  "electricity_gen_wtt", "overall_elec_scope_3", "gas_wtt", "heating_oil_wtt", "lpg_wtt")
      output_emf           = [dict(zip(keys, values)) for values in t_output_emf] 

      if len(output_emf) == 0:
        ret_mess['em'] = "****Error - No records found on emissions_factors table"
        ret_mess['ef'] = 2
        return ret_mess
     
      # Get building level details for all controlled buildings in entity
      sqlrd = f"SELECT uprn, data_year, baseline_annual_elec_kwh, baseline_annual_gas_kwh, baseline_annual_oil_kwh, baseline_annual_lpg_kwh, \
              baseline_elec_cost_per_kwh, baseline_gas_cost_per_kwh, baseline_oil_cost_per_kwh, baseline_lpg_cost_per_kwh \
              FROM raw_estate_data WHERE entity_number = {entity_number} AND under_control = \'Yes\';"
      cursor.execute(sqlrd)
      t_output_rd      = cursor.fetchall()
      keys             = ("uprn", "data_year", "baseline_annual_elec_kwh", "baseline_annual_gas_kwh", "baseline_annual_oil_kwh", "baseline_annual_lpg_kwh",\
                         "baseline_elec_cost_per_kwh", "baseline_gas_cost_per_kwh", "baseline_oil_cost_per_kwh", "baseline_lpg_cost_per_kwh")
      output_rd        = [dict(zip(keys, values)) for values in t_output_rd]
      
      # Loop through each building in estate and get required building data
      num_builds           = len(output_rd)
      num_projs            = 0
      build_count          = 0
      
      for bui in output_rd:
        build_count       = build_count + 1
        uprn              = bui['uprn']
        data_year         = bui['data_year']
        annual_elec_kwh   = bui['baseline_annual_elec_kwh']
        annual_gas_kwh    = bui['baseline_annual_gas_kwh']
        annual_oil_kwh    = bui['baseline_annual_oil_kwh']
        annual_lpg_kwh    = bui['baseline_annual_lpg_kwh']
        elec_unit_cost    = bui['baseline_elec_cost_per_kwh']
        gas_unit_cost     = bui['baseline_gas_cost_per_kwh']
        oil_unit_cost     = bui['baseline_oil_cost_per_kwh']
        lpg_unit_cost     = bui['baseline_lpg_cost_per_kwh']        
 
#        if build_count > 3:
#          return ret_mess
        print(f">>>>>>>>>>> Building uprn = {uprn} \n")
        # Get all the projects for this building. Loop through the projects calculating energy and carbon savings and write these to the project_results table
        
        sqlpr = f"SELECT project_id, project_type_id, assessed, assessed_delivery_date, delivery_date_mode, saving_percent, cost_capex_mode, salix_pf, utility, solar_pv_tot_zero_carb_gen_kwh, solar_thermal_corrected_ann_gen_kwh, hp_scop, hp_elec_add_kwh_pa FROM projects WHERE ((entity_number = {entity_number}) AND (uprn = {uprn}));"
        cursor.execute(sqlpr)
        t_output_prd         = cursor.fetchall()
        keys                 = ("project_id", "project_type_id", "assessed", "assessed_delivery_date", "delivery_date_mode", "saving_percent", "cost_capex_mode", "salix_pf", "utility", "solar_pv_tot_zero_carb_gen_kwh", "solar_thermal_corrected_ann_gen_kwh", "hp_scop","hp_elec_add_kwh_pa")
        output_prd           = [dict(zip(keys, values)) for values in t_output_prd]
        
        if len(output_prd) == 0:
          print(f"Skipped building uprn {uprn} - NO PROJECTS")
          continue
        num_projs            = num_projs + len(output_prd)
        pr_count             = 0
        
        for pr in output_prd:
          pr_count                = pr_count + 1
          project_id              = pr['project_id']
          project_type_id         = pr['project_type_id']
          assessed                = pr['assessed'] 
          assessed_delivery_date  = pr['assessed_delivery_date']
          delivery_date_mode      = pr['delivery_date_mode']
          saving_percent          = pr['saving_percent']  
          capex                   = pr['cost_capex_mode']
          capex                   = float(capex)
          
          salix_pf                = pr['salix_pf']
          salix_pf                = float(salix_pf)

          utility                 = pr['utility']
          solar_pv_saving         = pr['solar_pv_tot_zero_carb_gen_kwh']
          solar_thermal_saving    = pr['solar_thermal_corrected_ann_gen_kwh']
          hp_scop                 = pr['hp_scop']
          hp_elec_add_kwh_pa      = pr['hp_elec_add_kwh_pa']

          # Select the delivery date to be used for start of project savings
          print(f"\n Evaluate delivery_date - delivery_date_mode = {delivery_date_mode}, assessed _delivery_date = {assessed_delivery_date}")
          if delivery_date_mode.year == 1900:
            delivery_date = assessed_delivery_date
          else:
            delivery_date = delivery_date_mode
          print(f"Delivery date selected = {delivery_date} for project_id = {project_id}. Lifetime = {salix_pf}")
          if assessed != 'ASSESSED/NV' and assessed != 'IN PLACE':
            proj_elec_savings                = 0
            proj_gas_savings                 = 0
            proj_oil_savings                 = 0
            proj_lpg_savings                 = 0
            proj_elec_annual_cost_saving     = 0
            proj_gas_annual_cost_saving      = 0
            proj_oil_annual_cost_saving      = 0
            proj_lpg_annual_cost_saving      = 0
            project_cost_savings             = 0
            project_financial_value          = 0
            ROI_percent                      = 0 
            project_energy_savings           = 0
            elec_carbon                      = 0
            gas_carbon                       = 0
            project_carbon_savings           = 0
            tonne_co2_lifetime_cost          = 0
            lifetime_tonnes_CO2e             = 0
            annual_abatement_cost_tCO2e      = 0
            annual_co2_tco2e                 = 0
            total_scope_1                    = 0
            total_scope_2                    = 0
            total_scope_3                    = 0

            if project_type_id == 20:  #Solar PV
              proj_elec_savings      = solar_pv_saving
              proj_gas_saving        = 0

            elif project_type_id == 21:  # Solar Thermal
              if utility == 'ELEC':
                proj_elec_savings      = solar_thermal_saving
                proj_gas_savings       = 0
    
              if utility == 'GAS':
                proj_gas_savings       = solar_thermal_saving
                proj_elec_savings      = 0            

            elif (project_type_id == 14) or (project_type_id == 12): #Heat Pump (GAS Saving)  or Heat Network

              proj_gas_savings       = annual_gas_kwh * saving_percent
              # Calc negative savings on electricity
              if hp_elec_add_kwh_pa != 0:
                proj_elec_savings    = -(abs(hp_elec_add_kwh_pa)) #If a value has been provided for hp_elec_add_kwh_pa use its negative absolute value
              elif hp_scop != 0:
                proj_elec_savings    = -proj_gas_savings/abs(hp_scop) # If hp_scop provided but no hp_elec_add_kwh_pa then calculate using negative absolute hp_scop
              else:
                proj_elec_savings    = -proj_gas_savings/3 
            
            else:
              if utility == 'ELEC': # Select 'ELEC' when project saves electricity. Projects that save electricity (even when used for heating) can't save on fossil fuel usage
                proj_elec_savings      = annual_elec_kwh * saving_percent
                proj_gas_savings       = 0
                proj_oil_savings       = 0
                proj_lpg_savings       = 0
              if utility == 'GAS': # Select 'GAS' when project saves use of fossil fuels (e.g. insulation projects) for heating
                proj_elec_savings      = 0
                proj_gas_savings       = annual_gas_kwh * saving_percent
                proj_oil_savings       = annual_oil_kwh * saving_percent
                proj_lpg_savings       = annual_lpg_kwh * saving_percent
              
            project_energy_savings   = proj_elec_savings + proj_gas_savings + proj_oil_savings + proj_lpg_savings

            # Calculate annual energy cost savings

            proj_elec_annual_cost_saving = proj_elec_savings * elec_unit_cost
            proj_gas_annual_cost_saving  = proj_gas_savings * gas_unit_cost
            proj_oil_annual_cost_saving  = proj_oil_savings * oil_unit_cost
            proj_lpg_annual_cost_saving  = proj_lpg_savings * lpg_unit_cost 
        
            project_cost_savings         = proj_elec_annual_cost_saving + proj_gas_annual_cost_saving + proj_oil_annual_cost_saving + proj_lpg_annual_cost_saving  
            project_financial_value      = project_cost_savings * decimal.Decimal(salix_pf) #Not adjusted for inflation
            
            # Calculate Return on investment percentage
            
            if capex > 0:
              ROI_percent                = ((project_financial_value - decimal.Decimal(capex)) / decimal.Decimal(capex) ) * 100
            else:
              ROI_percent                = 0            

            # Calculate project lifetime carbon savings - total all scopes, total scope 1, total scope 2 and total scope 3.
            print(f"About to call calculate_lifetime_carbon for project_id : {project_id}")
            ret_ltc_mess                 = calculate_lifetime_carbon(output_emf, delivery_date, salix_pf, proj_elec_savings, proj_gas_savings, proj_oil_savings, proj_lpg_savings)
            
            print('Return from call to calculate_lifetime_carbon ret_ltc_mess:')
            print(ret_ltc_mess)
#            if (build_count == 1) and (pr_count ==1 ):
#              return ret_mess
            
            if ret_ltc_mess['ef'] == 2:
              ret_mess['ef'] = 2
              ret_mess['em'] = f"{ret_mess['em']} \n ***ERROR from calculate_lifetime_carbon - {ret_ltc_mess['em'] } \n"
              return ret_mess
            if ret_ltc_mess['ef'] == 1:
              ret_mess['ef'] = 1
              ret_mess['em'] = f"{ret_mess['em']} \n Warning from calculate_lifetime_carbon - {ret_ltc_mess['em']} \n"

            total_scope_1        = ret_ltc_mess['total_scope_1'] # scope 1 savings over lifetime of project
            total_scope_2        = ret_ltc_mess['total_scope_2'] # scope 2 savings over lifetime of project
            total_scope_3        = ret_ltc_mess['total_scope_3'] # scope 3 savings over lifetime of project
            lifetime_tonnes_CO2e = total_scope_1 + total_scope_2 + total_scope_3

            if lifetime_tonnes_CO2e > 0:
              tonne_co2_lifetime_cost	= decimal.Decimal(capex)/lifetime_tonnes_CO2e
            else:
              tonne_co2_lifetime_cost	= 0
 
            # Calculate the annual abatement (average) and annual abatement cost (£/tCO2e)
            
            if salix_pf > 0:
              annual_abatement_cost       = (decimal.Decimal(capex)/decimal.Decimal(salix_pf)) - project_cost_savings
              annual_co2_tco2e            = lifetime_tonnes_CO2e/decimal.Decimal(salix_pf) # Average annual abatement - saved in database as carbon_savings (historical reasons)
            else:
              annual_abatement_cost       = 0
              annual_co2_tco2e            = 0

            if annual_co2_tco2e > 0:
              annual_abatement_cost_tCO2e = annual_abatement_cost / annual_co2_tco2e # Annual abatement cost per Tonne CO2
            else:
              annual_abatement_cost_tCO2e = 0
              
            # Delete the record for this project in project_results table and re-insert with latest values calculated here
            
            sqldp = f"DELETE FROM project_results WHERE project_id = {project_id};"
            cursor.execute(sqldp)  
            conn.commit()

            sqlpri1 = f"INSERT INTO project_results (project_id, project_type_id, uprn, entity_number, energy_savings, gas_savings, electric_savings, oil_savings, lpg_savings, carbon_savings, tonne_co2_lifetime_cost, scope_1_savings, scope_2_savings, scope_3_savings, hp_elec_add_kwh_pa, \
            proj_elec_annual_cost_saving, proj_gas_annual_cost_saving, proj_oil_annual_cost_saving, proj_lpg_annual_cost_saving, project_cost_savings, project_financial_value, ROI_percent, lifetime_tonnes_CO2e, annual_abatement_cost_tCO2e) VALUES ("
            sqlpri2 = f"{project_id}, {project_type_id}, {uprn}, {entity_number}, {project_energy_savings}, {proj_gas_savings}, {proj_elec_savings},{proj_oil_savings}, {proj_lpg_savings}, {annual_co2_tco2e},{tonne_co2_lifetime_cost}, {total_scope_1}, {total_scope_2}, {total_scope_3},{hp_elec_add_kwh_pa},\
            {proj_elec_annual_cost_saving}, {proj_gas_annual_cost_saving}, {proj_oil_annual_cost_saving}, {proj_lpg_annual_cost_saving}, {project_cost_savings}, {project_financial_value}, {ROI_percent}, {lifetime_tonnes_CO2e}, {annual_abatement_cost_tCO2e});"
            sqlpri  = f"{sqlpri1} {sqlpri2}"
            
            cursor.execute(sqlpri)
            conn.commit
            #ret_mess['ef'] = 0
            #ret_mess['em'] = sqlpri
    return ret_mess     
          
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
    return ret_mess
    
def calculate_lifetime_carbon(output_emf, delivery_date, salix_pf, proj_elec_savings, proj_gas_savings, proj_oil_savings, proj_lpg_savings):
# Calculates the total lifetime tCO2e savings for a project and the breakdown into Scopes 1,2 & 3.
# For explanation of method see 'OnePointFive - method definitions - Lifetime Tonnes CO2 e, abatement costs per tonne, annual abatement'
  
  ret_mess             = {'ef':0, 'em':'Normal return', 'lifetime_tonnes_CO2e':0, 'total_scope_1':0, 'total_scope_2':0, 'total_scope_3':0}
 # print('calculate_lifetime_carbon at top')
  try:
    nr                 = len(output_emf)
    if nr == 0:
      ret_mess['ef']   = 2
      ret_mess['em']   = '****ERROR - emission factor table is empty'
      return ret_mess

    min_dict = min(output_emf, key=lambda x: x['data_year'])
    min_year = min_dict['data_year']

    if delivery_date.year < min_year:
      ret_mess['ef']   = 2
      ret_mess['em']   = '****ERROR - delivery date out of bounds - must be >= 2017'
      return ret_mess      
 
    if salix_pf <= 1:
      ret_mess['ef']   = 2
      ret_mess['em']   = '****ERROR - persistence factor (lifetime) must be > 1'
      return ret_mess

    # Get last year for which we have emission factors. Projects with a lifetime extending beyond this point will use the emission factors for this last year
    # for every year beyond.

    max_dict = max(output_emf, key=lambda x: x['data_year'])
    max_data_year = max_dict['data_year']
  
    # Calculate the number of days left in delivery year, end of lifetime date and number of days in end of lifetime year to end of lifetime date,
    # Calculate year fractions for delivery year and end of life year using days (as opposed to months or weeks)
    
    delivery_day_of_year = delivery_date.timetuple().tm_yday
    delivery_year        = delivery_date.year
#    print(f"Delivery date = {delivery_date}")
#    print(f"Delivery date - day of year = {delivery_day_of_year}")
#    print(f"Delivery year = {delivery_year}")
#    print(f"Salix_pf = {salix_pf}")
    #import calendar
    if calendar.isleap(delivery_date.year):
      numdindy = 366
    else:
      numdindy = 365
#    print(f"Number of days in delivery year = {numdindy}")
    
    # Calculate the end of life date      
    life_time = salix_pf
    ly        = np.floor(life_time)          #whole number of years in life time
    lf        = life_time - ly               #remainder of time in lifetime as fraction of a year
    ld        = round_half_up(lf * numdindy) #remainder of time in lifetime as days (used to calculate end of life date)
    eol_date  = delivery_date + relativedelta(years = ly, days = ld)
    eol_year  = eol_date.year
    
    # Calculate year fractions for delivery year and end of life year
    delivery_year_fraction =  (numdindy - delivery_day_of_year)/numdindy
    
    eol_day_of_year = eol_date.timetuple().tm_yday
    if calendar.isleap(eol_date.year):
      numdiney = 366
    else:
      numdiney = 365
    eol_year_fraction      = eol_day_of_year/numdiney
#    print(f"Number of days in EOL year = {numdiney}")
#    print(f"EOL day of year = {eol_day_of_year}")
#    print('EOL date')
#    print(eol_date)
 #   print(f"EOL year = {eol_year}")
#    print(f"Delivery year fraction = {delivery_year_fraction}")
#    print(f"EOL year fraction = {eol_year_fraction}")
#    print(f"Max data year = {max_data_year}")

    # Build the year and year fraction lists for the lifetime of the project

    year_list          = list()
    year_fraction_list = list()

    year_list.append(delivery_year)
    year_fraction_list.append(delivery_year_fraction)

    yr  = delivery_year
    irange = (eol_year-1) - delivery_year
    for x in range(irange):
      yr = yr + 1
      year_list.append(yr)    
      year_fraction_list.append(1)

    year_list.append(eol_year)
    year_fraction_list.append(eol_year_fraction)

 #   print('Created year and year fraction lists')
    print(year_list)
    print(year_fraction_list)

 #   print('AT test of dateutil end ----')
    total_annual_tco2e     = 0
    total_scope_1          = 0
    total_scope_2          = 0
    total_scope_3          = 0
    
    # Loop through the years in lifetime, calculate total carbon and scopes 1,2,3 in each year and sum as we go.
    ind = 0
    for yr in year_list:
      year_fraction = year_fraction_list[ind]
      ind = ind + 1
#      print(f"yr is - {yr}")
      if yr > max_data_year: #If yr exceeds maximum year in the emissions factor dataframe then set yr to maximum year
        use_year = max_data_year
      else:
        use_year = yr
        
    # Select the dict (emf_row) for the appropriate year

      emf_row = next(item for item in output_emf if item["data_year"] == use_year)

      elec_co2             = ((emf_row['elec_consumed']  * proj_elec_savings)/1000) * decimal.Decimal(year_fraction)
      gas_co2              = ((emf_row['gas_consumed']  *  proj_gas_savings)/1000) * decimal.Decimal(year_fraction)
      oil_co2              = ((emf_row['oil_consumed']  *  proj_oil_savings)/1000) * decimal.Decimal(year_fraction)
      lpg_co2              = ((emf_row['lpg_consumed']  *  proj_lpg_savings)/1000) * decimal.Decimal(year_fraction)
  
      gas_wtt_scope_3      = ((proj_gas_savings *  emf_row['gas_wtt'])/1000) * decimal.Decimal(year_fraction)
      elec_t_d_scope_3     = ((proj_elec_savings * emf_row['electricity_t_d'])/1000) * decimal.Decimal(year_fraction)
      elec_wtt_t_d_scope_3 = ((proj_elec_savings * emf_row['electricity_t_d_wtt'])/1000) * decimal.Decimal(year_fraction)
      elec_wtt_gen_scope_3 = ((proj_elec_savings * emf_row['electricity_gen_wtt'])/1000) * decimal.Decimal(year_fraction)
      oil_wtt_scope_3      = ((proj_oil_savings * emf_row['heating_oil_wtt'])/1000) * decimal.Decimal(year_fraction)
      lpg_wtt_scope_3      = ((proj_lpg_savings * emf_row['lpg_wtt'])/1000) * decimal.Decimal(year_fraction)
      
      total_scope_1        = total_scope_1 + gas_co2 + oil_co2 + lpg_co2
      total_scope_2        = total_scope_2 + elec_co2
      total_scope_3        = total_scope_3 + gas_wtt_scope_3 + elec_t_d_scope_3 + elec_wtt_t_d_scope_3 + elec_wtt_gen_scope_3 + oil_wtt_scope_3 + lpg_wtt_scope_3    

      print(f"In calc calculate lifetime carbon at end \n Year = {yr} \n use_year = {use_year} \n Total Scope 1 = {total_scope_1} \n Total Scope 2 = {total_scope_2} \n Total Scope 3 = {total_scope_3} \n total_annual_tco2e = {total_annual_tco2e} \n")

    ret_mess['total_scope_1']        = total_scope_1
    ret_mess['total_scope_2']        = total_scope_2
    ret_mess['total_scope_3']        = total_scope_3
    return ret_mess
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret_mess['ef'] = 2
    ret_mess['em'] = msg
    return ret_mess
    
@anvil.server.background_task
def create_pbi_tables_v3( conn, entity_number):
# THIS VERSION IS CALLABLE FROM ANOTHER BACKGROUND TASK
# This subroutine assembles the dataframes and writes to the database tables for use in Power BI for entity with entity_number using database connection conn.
# 
# The following tables are updated: -
#    pbi_estate, pbi_waterfall_project_savings.
#    Power BI can use the estate_summary and solar_summary tables without further modification.
  print('At start of create_pbi_tables')
  ret_mess  = {'ef':0, 'em':0}
  summary   = ''
  up_log    = ''

  try:
    
    odbc_str = Connections.connection_string
    print('****Connection string')
    print(odbc_str)
    
    with conn.cursor() as cur:

    #====================================================================================================
    #
    # Assemble estate results 
    #
    #=====================================================================================================    

      sqle                         = f"SELECT ra.uprn, ra.building_name, ra.latitude_dd, ra.longitude_dd, ra.building_type, ra.under_control, ra.listed, ra.baseline_annual_elec_kwh,ra.baseline_annual_gas_kwh,ra.baseline_annual_oil_kwh,ra.baseline_annual_lpg_kwh, ra.dec_score, ra.epc, \
                                      rr.elec_co2, rr.gas_co2, rr.oil_co2, rr.lpg_co2, rr.gas_wtt_scope_3, rr.elec_t_d_scope_3, rr.elec_wtt_t_d_scope_3, rr.elec_wtt_gen_scope_3, rr.oil_wtt_scope_3, rr.lpg_wtt_scope_3,\
                                      rr.total_scope_1, rr.total_scope_2, rr.total_scope_3, rr.total_co2_tco2e, rr.annual_elec_cost, rr.annual_gas_cost, rr.annual_oil_cost, rr.annual_lpg_cost, rr.annual_energy_cost, \
                                      rr.total_kwh, rr.elec_kwh_m2, rr.gas_kwh_m2, rr.bmark_elec_kwh_m2b, rr.bmark_gas_kwh_m2b, rr.elec_2b_saved_2_typical, rr.gas_2b_saved_2_typical, rr.baseline_flag, ra.gia_m2 \
                                      FROM raw_estate_data AS ra \
                                      LEFT JOIN results_raw_estate_data AS rr \
                                      ON (ra.uprn = rr.uprn) AND (ra.entity_number = rr.entity_number)\
                                      WHERE ra.entity_number = {entity_number};"

      dfestate            = pd.read_sql_query(sqle, conn) 
      print("=======>>>>>dfestate on creation")
      print(dfestate.to_string()) 
      print(" ")
      print(sqle)
      print(' ')

    # Add columns to hold entity_number & DEC rating

      dfestate.insert(loc=0, column='entity_number', value=entity_number, allow_duplicates=True)
      dfestate.insert(loc=13, column='dec_rating', value='', allow_duplicates=True)

    # Calculate DEC rating letters and insert in dec_rating column
      
      for index,row in dfestate.iterrows():
        ds                       = row['dec_score']
        row['dec_rating']        = get_dec_letter(ds)
        dfestate.iloc[index]     = row

 #  Get total energy consumption for estate from the controlled_estate_summary table
        
      sqles                      = f"SELECT total_elec_kwh, total_gas_kwh, total_oil_kwh, total_lpg_kwh, total_energy_kwh, co2_scope_1, co2_scope_2, co2_scope_3, co2_total FROM controlled_estate_summary WHERE entity_number = {entity_number};"
      cur.execute(sqles)
      keys                 = [column[0] for column in cur.description]
      t_output_es          = cur.fetchall() 
      output_es            = [dict(zip(keys, values)) for values in t_output_es]
      
      dfestatesummary    = pd.DataFrame.from_dict (output_es)         
#      print('In create_pbi_tables forming estate table 1, Here is dfestate: -----------------------')
#      print(dfestate.to_string())
#      return
    #  Delete all records for this entity_number (for this scenario) in pbi_waterfall_project_savings table then insert the waterfall dataframe

      sqlde  = f"DELETE FROM pbi_estate WHERE entity_number = {entity_number} ;"
      cur.execute(sqlde)
      conn.commit()

      print('Creating engine and doing to_sql')
      print(Connections.connection_string)
      connect_str = 'mssql+pyodbc:///?odbc_connect=' + urllib.parse.quote_plus(odbc_str)
      print('======connect_str')
      print(connect_str)

      engine      = salch.create_engine(connect_str)
      with engine.connect().execution_options(autocommit=False) as conn2:
        print('dfestate before to_sql==========')
        print(dfestate.to_string())
        txn = conn2.begin()
        dfestate.to_sql('pbi_estate', con=conn2, if_exists='append', index= False)
#        print(dir(conn2))
        txn.commit()      
    #====================================================================================================
    #
    # Assemble waterfalls table for waterfall visualisations of project savings
    #
    #=====================================================================================================

    # Get data for all projects for this entity_number from projects and project_results tables. Filter out projects where assessed is 'IN PLACE' or 'ASSESSED/NV'

      sqlj               = f"SELECT  projects.entity_number, projects.uprn, projects.project_id, projects.assessed, projects.assessed_delivery_date, \
                                     projects.cost_capex_mode, projects.delivery_date_mode, project_results.energy_savings, project_results.gas_savings, \
                                     project_results.electric_savings, project_results.oil_savings, project_results.lpg_savings,project_results.carbon_savings, project_results.tonne_co2_lifetime_cost, \
                                     project_results.scope_1_savings, project_results.scope_2_savings, project_results.scope_3_savings \
                                     FROM projects \
                                     INNER JOIN project_results ON projects.project_id = project_results.project_id \
                                     WHERE projects.entity_number = {entity_number} AND assessed != 'IN PLACE' AND assessed != 'ASSESSED/NV';"
    
      dfprojects         = pd.read_sql_query(sqlj, conn)

        
    # Sort the dataframe by uprn to get all projects for buildings together
    
      dfprojects                = dfprojects.sort_values(by='uprn')
      
    # If delivery_date_mode has not been explicitly set (i.e. it is still the default 1900-01-01) or is NULL then set it to the assessed_delivery_date
    
      tdate = pd.to_datetime('1900-01-01',format = "%Y-%m-%d" )
      dfprojects.loc[(dfprojects.delivery_date_mode == tdate), 'delivery_date_mode'] = dfprojects.assessed_delivery_date
#      dfprojects.loc[(dfprojects.delivery_date_mode == None), 'delivery_date_mode'] = dfprojects.assessed_delivery_date
      
    # Copy the project savings dataframe
    
      dfwf                       = dfprojects.copy()
    
    # Extract just the columns we need

      scenario_number            = 0
    
      dfwf1                      = dfwf[['entity_number', 'uprn', 'project_id', 'assessed','delivery_date_mode','energy_savings', 'gas_savings', 'electric_savings', 'oil_savings', 
                                         'lpg_savings', 'carbon_savings', 'tonne_co2_lifetime_cost', 'scope_1_savings', 'scope_2_savings', 'scope_3_savings']].copy()
      dfwfall                    = pd.DataFrame(columns=[ 'entity_number', 'uprn', 'project_id', 'likelihood', 'ordering','delivery_date', 'param_name', 'param_value', 'title', 'scenario_number'])
#      print('dfwf1-----')
#      print(dfwf1.to_string())
      total_energy_savings       = dfwf1['energy_savings'].sum()
      total_electric_savings     = dfwf1['electric_savings'].sum()
      total_gas_savings          = dfwf1['gas_savings'].sum()
      total_oil_savings          = dfwf1['oil_savings'].sum()
      total_lpg_savings          = dfwf1['lpg_savings'].sum()
      total_co2_savings          = dfwf1['carbon_savings'].sum()
      total_scope_1_savings      = dfwf1['scope_1_savings'].sum()
      total_scope_2_savings      = dfwf1['scope_2_savings'].sum()
      total_scope_3_savings      = dfwf1['scope_3_savings'].sum()
      
      baseline_energy            = dfestatesummary.at[0,'total_energy_kwh']
      baseline_electric          = dfestatesummary.at[0,'total_elec_kwh']
      baseline_gas               = dfestatesummary.at[0,'total_gas_kwh']
      baseline_oil               = dfestatesummary.at[0,'total_oil_kwh']
      baseline_lpg               = dfestatesummary.at[0,'total_lpg_kwh']
      baseline_co2_total         = dfestatesummary.at[0,'co2_total']
      baseline_co2_scope_1       = dfestatesummary.at[0,'co2_scope_1']
      baseline_co2_scope_2       = dfestatesummary.at[0,'co2_scope_2']
      baseline_co2_scope_3       = dfestatesummary.at[0,'co2_scope_3']

      remaining_energy_kwh       = baseline_energy         -       decimal.Decimal(total_energy_savings)
      remaining_electric_kwh     = baseline_electric       -       decimal.Decimal(total_electric_savings)
      remaining_gas_kwh          = baseline_gas            -       decimal.Decimal(total_gas_savings)
      remaining_oil_kwh          = baseline_oil            -       decimal.Decimal(total_oil_savings)
      remaining_lpg_kwh          = baseline_lpg            -       decimal.Decimal(total_lpg_savings)
      remaining_co2_total        = baseline_co2_total      -       decimal.Decimal(total_co2_savings)
      remaining_co2_scope_1      = baseline_co2_scope_1    -       decimal.Decimal(total_scope_1_savings)
      remaining_co2_scope_2      = baseline_co2_scope_2    -       decimal.Decimal(total_scope_2_savings)
      remaining_co2_scope_3      = baseline_co2_scope_3    -       decimal.Decimal(total_scope_3_savings)
      
      for index, row in dfwf1.iterrows():
        order                                 = 0
        assessed                              = ''
        ass                                   = row['assessed']
        entity_number                         = row['entity_number']
        uprn                                  = row['uprn']
        project_id                            = row['project_id']
        
        if ass == 'FIRM':
          assessed       = 'Firm (0-1yr)'
          order          = 2
        if ass == 'LIKELY':
          assessed       = 'Likely (1-2yrs)'
          order          = 3          
        if ass == 'POSSIBLE':
          assessed       = 'Possible (2-3yrs)'
          order          = 4                   
        if ass == 'POTENTIAL':
          assessed       = 'Potential (3-5yrs)'
          order          = 5  
        if ass == 'FTHR IMPV':
          assessed       = 'Further Improvement (3-5yrs)'
          order          = 6           
         
        delivery_date                         = row['delivery_date_mode']
        energy_savings                        = row['energy_savings']
        gas_savings                           = row['gas_savings'] 
        electric_savings                      = row['electric_savings']
        oil_savings                           = row['oil_savings'] 
        lpg_savings                           = row['lpg_savings']          
        total_co2_savings                     = row['carbon_savings']
        tonne_co2_lifetime_cost               = row['tonne_co2_lifetime_cost']
        scope_1_savings                       = row['scope_1_savings']
        scope_2_savings                       = row['scope_2_savings']
        scope_3_savings                       = row['scope_3_savings']
          
        # Temporary dataframe dft
        dft = pd.DataFrame(columns=[ 'entity_number', 'uprn', 'project_id', 'likelihood', 'ordering', 'delivery_date', 'param_name', 'param_value', 'title', 'scenario_number'],
                   index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9])
        
        dft.loc[[0],['entity_number']]         = entity_number
        dft.loc[[0],['uprn']]                  = uprn
        dft.loc[[0],['project_id']]            = project_id        
        dft.loc[[0],['likelihood']]            = assessed
        dft.loc[[0],['ordering']]              = order
        dft.loc[[0],['delivery_date']]         = delivery_date
        dft.loc[[0],['param_name']]            = 'Energy savings kWh pa'
        dft.loc[[0],['param_value']]           = round(-energy_savings,4)
        dft.loc[[0],['title']]                 = 'Interventions - energy savings - kWh pa'
        dft.loc[[0],['scenario_number']]       = scenario_number
        
        dft.loc[[1],['entity_number']]         = entity_number
        dft.loc[[1],['uprn']]                  = uprn
        dft.loc[[1],['project_id']]            = project_id
        dft.loc[[1],['likelihood']]            = assessed
        dft.loc[[1],['ordering']]              = order
        dft.loc[[1],['delivery_date']]         = delivery_date
        dft.loc[[1],['param_name']]            = 'Gas savings kWh pa'
        dft.loc[[1],['param_value']]           = round(-gas_savings,4)
        dft.loc[[1],['title']]                 = 'Interventions - gas savings - kWh pa'
        dft.loc[[1],['scenario_number']]       = scenario_number
      
        dft.loc[[2],['entity_number']]         = entity_number
        dft.loc[[2],['uprn']]                  = uprn
        dft.loc[[2],['project_id']]            = project_id
        dft.loc[[2],['likelihood']]            = assessed
        dft.loc[[2],['ordering']]              = order
        dft.loc[[2],['delivery_date']]         = delivery_date
        dft.loc[[2],['param_name']]            = 'Electricity savings kWh pa'
        dft.loc[[2],['param_value']]           = round(-electric_savings,4)    
        dft.loc[[2],['title']]                 = 'Interventions - electricity savings - kWh pa'
        dft.loc[[2],['scenario_number']]       = scenario_number
        
        dft.loc[[3],['entity_number']]         = entity_number
        dft.loc[[3],['uprn']]                  = uprn
        dft.loc[[3],['project_id']]            = project_id
        dft.loc[[3],['likelihood']]            = assessed
        dft.loc[[3],['ordering']]              = order
        dft.loc[[3],['delivery_date']]         = delivery_date
        dft.loc[[3],['param_name']]            = 'Oil savings kWh pa'
        dft.loc[[3],['param_value']]           = round(-oil_savings,4)    
        dft.loc[[3],['title']]                 = 'Interventions - oil savings - kWh pa'
        dft.loc[[3],['scenario_number']]       = scenario_number
        
        dft.loc[[4],['entity_number']]         = entity_number
        dft.loc[[4],['uprn']]                  = uprn
        dft.loc[[4],['project_id']]            = project_id
        dft.loc[[4],['likelihood']]            = assessed
        dft.loc[[4],['ordering']]              = order
        dft.loc[[4],['delivery_date']]         = delivery_date
        dft.loc[[4],['param_name']]            = 'LPG savings kWh pa'
        dft.loc[[4],['param_value']]           = round(-lpg_savings,4)    
        dft.loc[[4],['title']]                 = 'Interventions - lpg savings - kWh pa'
        dft.loc[[4],['scenario_number']]       = scenario_number
        
        dft.loc[[5],['entity_number']]         = entity_number
        dft.loc[[5],['uprn']]                  = uprn
        dft.loc[[5],['project_id']]            = project_id
        dft.loc[[5],['likelihood']]            = assessed
        dft.loc[[5],['ordering']]              = order
        dft.loc[[5],['delivery_date']]         = delivery_date
        dft.loc[[5],['param_name']]            = 'Carbon savings Tonnes pa'
        dft.loc[[5],['param_value']]           = round(-total_co2_savings,4)    
        dft.loc[[5],['title']]                 = 'Interventions - carbon savings - tonnes pa'
        dft.loc[[5],['scenario_number']]       = scenario_number
        
        dft.loc[[6],['entity_number']]         = entity_number
        dft.loc[[6],['uprn']]                  = uprn
        dft.loc[[6],['project_id']]            = project_id
        dft.loc[[6],['likelihood']]            = assessed
        dft.loc[[6],['ordering']]              = order
        dft.loc[[6],['delivery_date']]         = delivery_date
        dft.loc[[6],['param_name']]            = '£ CO2 tonnes lifetime cost savings'
        dft.loc[[6],['param_value']]           = round(-tonne_co2_lifetime_cost,4)    
        dft.loc[[6],['title']]                 = 'Interventions - carbon savings - £ CO2 tonnes lifetime '
        dft.loc[[6],['scenario_number']]       = scenario_number

        dft.loc[[7],['entity_number']]         = entity_number
        dft.loc[[7],['uprn']]                  = uprn
        dft.loc[[7],['project_id']]            = project_id
        dft.loc[[7],['likelihood']]            = assessed
        dft.loc[[7],['ordering']]              = order
        dft.loc[[7],['delivery_date']]         = delivery_date
        dft.loc[[7],['param_name']]            = 'Scope 1 savings Tonnes pa'
        dft.loc[[7],['param_value']]           = round(-scope_1_savings,4)    
        dft.loc[[7],['title']]                 = 'Interventions - scope 1 savings - tonnes pa'
        dft.loc[[7],['scenario_number']]       = scenario_number

        dft.loc[[8],['entity_number']]         = entity_number
        dft.loc[[8],['uprn']]                  = uprn
        dft.loc[[8],['project_id']]            = project_id
        dft.loc[[8],['likelihood']]            = assessed
        dft.loc[[8],['ordering']]              = order
        dft.loc[[8],['delivery_date']]         = delivery_date
        dft.loc[[8],['param_name']]            = 'Scope 2 savings Tonnes pa'
        dft.loc[[8],['param_value']]           = round(-scope_2_savings,4)    
        dft.loc[[8],['title']]                 = 'Interventions - scope 2 savings - tonnes pa'
        dft.loc[[8],['scenario_number']]       = scenario_number

        dft.loc[[9],['entity_number']]         = entity_number
        dft.loc[[9],['uprn']]                  = uprn
        dft.loc[[9],['project_id']]            = project_id
        dft.loc[[9],['likelihood']]            = assessed
        dft.loc[[9],['ordering']]              = order
        dft.loc[[9],['delivery_date']]         = delivery_date
        dft.loc[[9],['param_name']]            = 'Scope 3 savings Tonnes pa'
        dft.loc[[9],['param_value']]           = round(-scope_3_savings,4)    
        dft.loc[[9],['title']]                 = 'Interventions - scope 3 savings - tonnes pa' 
        dft.loc[[9],['scenario_number']]       = scenario_number
        
        dfwfall                                = dfwfall.append(dft)
        
    # Insert the baseline and remaining figures
      dft = pd.DataFrame(columns=[ 'entity_number', 'uprn', 'project_id', 'likelihood', 'ordering','delivery_date', 'param_name', 'param_value', 'title', 'scenario_number'],
                  index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17])    
      
      dft.loc[[0],['entity_number']]         = entity_number
      dft.loc[[0],['uprn']]                  = 0
      dft.loc[[0],['project_id']]            = 0
      dft.loc[[0],['likelihood']]            = 'BASELINE'
      dft.loc[[0],['ordering']]              = 1
      dft.loc[[0],['delivery_date']]         = '2022-1-1'
      dft.loc[[0],['param_name']]            = 'Energy savings kWh pa'
      dft.loc[[0],['param_value']]           = round(baseline_energy,4)
      dft.loc[[0],['title']]                 = 'Intervention energy savings - kWh pa'
      dft.loc[[0],['scenario_number']]       = 0
      
      dft.loc[[1],['entity_number']]         = entity_number
      dft.loc[[1],['uprn']]                  = 0
      dft.loc[[1],['project_id']]            = 0    
      dft.loc[[1],['likelihood']]            = 'BASELINE'
      dft.loc[[1],['ordering']]              = 1
      dft.loc[[1],['delivery_date']]         = '2022-1-1'
      dft.loc[[1],['param_name']]            = 'Gas savings kWh pa'
      dft.loc[[1],['param_value']]           = round(baseline_gas,4)
      dft.loc[[1],['title']]                 = 'Intervention gas savings - kWh pa'
      dft.loc[[1],['scenario_number']]       = 0
    
      dft.loc[[2],['entity_number']]         = entity_number
      dft.loc[[2],['uprn']]                  = 0
      dft.loc[[2],['project_id']]            = 0      
      dft.loc[[2],['likelihood']]            = 'BASELINE'
      dft.loc[[2],['ordering']]              = 1
      dft.loc[[2],['delivery_date']]         = '2022-1-1'
      dft.loc[[2],['param_name']]            = 'Electricity savings kWh pa'
      dft.loc[[2],['param_value']]           = round(baseline_electric,4)    
      dft.loc[[2],['title']]                 = 'Intervention electricity savings - kWh pa'
      dft.loc[[2],['scenario_number']]       = 0
      
      dft.loc[[3],['entity_number']]         = entity_number
      dft.loc[[3],['uprn']]                  = 0
      dft.loc[[3],['project_id']]            = 0      
      dft.loc[[3],['likelihood']]            = 'BASELINE'
      dft.loc[[3],['ordering']]              = 1
      dft.loc[[3],['delivery_date']]         = '2022-1-1'
      dft.loc[[3],['param_name']]            = 'Oil savings kWh pa'
      dft.loc[[3],['param_value']]           = round(baseline_oil,4)
      dft.loc[[3],['title']]                 = 'Intervention oil savings - kWh pa'
      dft.loc[[3],['scenario_number']]       = 0

      dft.loc[[4],['entity_number']]         = entity_number
      dft.loc[[4],['uprn']]                  = 0
      dft.loc[[4],['project_id']]            = 0      
      dft.loc[[4],['likelihood']]            = 'BASELINE'
      dft.loc[[4],['ordering']]              = 1
      dft.loc[[4],['delivery_date']]         = '2022-1-1'
      dft.loc[[4],['param_name']]            = 'LPG savings kWh pa'
      dft.loc[[4],['param_value']]           = round(baseline_lpg,4)
      dft.loc[[4],['title']]                 = 'Intervention lpg savings - kWh pa'
      dft.loc[[4],['scenario_number']]       = 0
     
      dft.loc[[5],['entity_number']]         = entity_number
      dft.loc[[5],['uprn']]                  = 0
      dft.loc[[5],['project_id']]            = 0      
      dft.loc[[5],['likelihood']]            = 'BASELINE'
      dft.loc[[5],['ordering']]              = 1
      dft.loc[[5],['delivery_date']]         = '2022-1-1'
      dft.loc[[5],['param_name']]            = 'Carbon savings Tonnes pa'
      dft.loc[[5],['param_value']]           = round(baseline_co2_total,4)    
      dft.loc[[5],['title']]                 = 'Intervention carbon savings - tonnes pa'
      dft.loc[[5],['scenario_number']]       = 0

      dft.loc[[6],['entity_number']]         = entity_number
      dft.loc[[6],['uprn']]                  = 0
      dft.loc[[6],['project_id']]            = 0      
      dft.loc[[6],['likelihood']]            = 'BASELINE'
      dft.loc[[6],['ordering']]              = 1
      dft.loc[[6],['delivery_date']]         = '2022-1-1'
      dft.loc[[6],['param_name']]            = 'Scope 1 savings Tonnes pa'
      dft.loc[[6],['param_value']]           = round(baseline_co2_scope_1,4)    
      dft.loc[[6],['title']]                 = 'Intervention scope 1 savings - tonnes pa'
      dft.loc[[6],['scenario_number']]       = 0
      
      dft.loc[[7],['entity_number']]         = entity_number
      dft.loc[[7],['uprn']]                  = 0
      dft.loc[[7],['project_id']]            = 0      
      dft.loc[[7],['likelihood']]            = 'BASELINE'
      dft.loc[[7],['ordering']]              = 1
      dft.loc[[7],['delivery_date']]         = '2022-1-1'
      dft.loc[[7],['param_name']]            = 'Scope 2 savings Tonnes pa'
      dft.loc[[7],['param_value']]           = round(baseline_co2_scope_2,4)    
      dft.loc[[7],['title']]                 = 'Intervention scope 2 savings - tonnes pa'
      dft.loc[[7],['scenario_number']]       = 0

      dft.loc[[8],['entity_number']]         = entity_number
      dft.loc[[8],['uprn']]                  = 0
      dft.loc[[8],['project_id']]            = 0      
      dft.loc[[8],['likelihood']]            = 'BASELINE'
      dft.loc[[8],['ordering']]              = 1
      dft.loc[[8],['delivery_date']]         = '2022-1-1'
      dft.loc[[8],['param_name']]            = 'Scope 3 savings Tonnes pa'
      dft.loc[[8],['param_value']]           = round(baseline_co2_scope_3,4)    
      dft.loc[[8],['title']]                 = 'Intervention scope 3 savings - tonnes pa'
      dft.loc[[8],['scenario_number']]       = 0
 
      dft.loc[[9],['entity_number']]         = entity_number
      dft.loc[[9],['uprn']]                  = 0
      dft.loc[[9],['project_id']]            = 0      
      dft.loc[[9],['likelihood']]            = 'REMAINING'
      dft.loc[[9],['ordering']]              = 7
      dft.loc[[9],['delivery_date']]         = '2022-1-1'
      dft.loc[[9],['param_name']]            = 'Energy savings kWh pa'
      dft.loc[[9],['param_value']]           = round(-remaining_energy_kwh,4)
      dft.loc[[9],['title']]                 = 'Intervention energy savings - kWh pa'
      dft.loc[[9],['scenario_number']]       = 0
      
      dft.loc[[10],['entity_number']]         = entity_number
      dft.loc[[10],['uprn']]                  = 0
      dft.loc[[10],['project_id']]            = 0      
      dft.loc[[10],['likelihood']]            = 'REMAINING'
      dft.loc[[10],['ordering']]              = 7
      dft.loc[[10],['delivery_date']]         = '2022-1-1'
      dft.loc[[10],['param_name']]            = 'Gas savings kWh pa'
      dft.loc[[10],['param_value']]           = round(-remaining_gas_kwh,4)
      dft.loc[[10],['title']]                 = 'Intervention gas savings - kWh pa'
      dft.loc[[10],['scenario_number']]       = 0

      dft.loc[[11],['entity_number']]         = entity_number
      dft.loc[[11],['uprn']]                  = 0
      dft.loc[[11],['project_id']]            = 0     
      dft.loc[[11],['likelihood']]            = 'REMAINING'
      dft.loc[[11],['ordering']]              = 7
      dft.loc[[11],['delivery_date']]         = '2022-1-1'
      dft.loc[[11],['param_name']]            = 'Electricity savings kWh pa'
      dft.loc[[11],['param_value']]           = round(-remaining_electric_kwh,4)    
      dft.loc[[11],['title']]                 = 'Intervention electricity savings - kWh pa'
      dft.loc[[11],['scenario_number']]       = 0

      dft.loc[[12],['entity_number']]         = entity_number
      dft.loc[[12],['uprn']]                  = 0
      dft.loc[[12],['project_id']]            = 0      
      dft.loc[[12],['likelihood']]            = 'REMAINING'
      dft.loc[[12],['ordering']]              = 7
      dft.loc[[12],['delivery_date']]         = '2022-1-1'
      dft.loc[[12],['param_name']]            = 'Oil savings kWh pa'
      dft.loc[[12],['param_value']]           = round(-remaining_oil_kwh,4)    
      dft.loc[[12],['title']]                 = 'Intervention oil savings - kWh pa'
      dft.loc[[12],['scenario_number']]       = 0

      dft.loc[[13],['entity_number']]         = entity_number
      dft.loc[[13],['uprn']]                  = 0
      dft.loc[[13],['project_id']]            = 0     
      dft.loc[[13],['likelihood']]            = 'REMAINING'
      dft.loc[[13],['ordering']]              = 7
      dft.loc[[13],['delivery_date']]         = '2022-1-1'
      dft.loc[[13],['param_name']]            = 'LPG savings kWh pa'
      dft.loc[[13],['param_value']]           = round(-remaining_lpg_kwh,4)    
      dft.loc[[13],['title']]                 = 'Intervention lpg savings - kWh pa'
      dft.loc[[13],['scenario_number']]       = 0

      dft.loc[[14],['entity_number']]         = entity_number
      dft.loc[[14],['uprn']]                  = 0
      dft.loc[[14],['project_id']]            = 0      
      dft.loc[[14],['likelihood']]            = 'REMAINING'
      dft.loc[[14],['ordering']]              = 7
      dft.loc[[14],['delivery_date']]         = '2022-1-1'
      dft.loc[[14],['param_name']]            = 'Carbon savings Tonnes pa'
      dft.loc[[14],['param_value']]           = round(-remaining_co2_total,4)    
      dft.loc[[14],['title']]                 = 'Intervention carbon savings - tonnes pa'
      dft.loc[[14],['scenario_number']]       = 0

      dft.loc[[15],['entity_number']]         = entity_number
      dft.loc[[15],['uprn']]                  = 0
      dft.loc[[15],['project_id']]            = 0      
      dft.loc[[15],['likelihood']]            = 'REMAINING'
      dft.loc[[15],['ordering']]              = 7
      dft.loc[[15],['delivery_date']]         = '2022-1-1'
      dft.loc[[15],['param_name']]            = 'Scope 1 savings Tonnes pa'
      dft.loc[[15],['param_value']]           = round(-remaining_co2_scope_1,4)    
      dft.loc[[15],['title']]                 = 'Intervention scope 1 savings - tonnes pa'
      dft.loc[[15],['scenario_number']]       = 0

      dft.loc[[16],['entity_number']]         = entity_number
      dft.loc[[16],['uprn']]                  = 0
      dft.loc[[16],['project_id']]            = 0      
      dft.loc[[16],['likelihood']]            = 'REMAINING'
      dft.loc[[16],['ordering']]              = 7
      dft.loc[[16],['delivery_date']]         = '2022-1-1'
      dft.loc[[16],['param_name']]            = 'Scope 2 savings Tonnes pa'
      dft.loc[[16],['param_value']]           = round(-remaining_co2_scope_2,4)    
      dft.loc[[16],['title']]                 = 'Intervention scope 2 savings - tonnes pa'
      dft.loc[[16],['scenario_number']]       = 0

      dft.loc[[17],['entity_number']]         = entity_number
      dft.loc[[17],['uprn']]                  = 0
      dft.loc[[17],['project_id']]            = 0      
      dft.loc[[17],['likelihood']]            = 'REMAINING'
      dft.loc[[17],['ordering']]              = 7
      dft.loc[[17],['delivery_date']]         = '2022-1-1'
      dft.loc[[17],['param_name']]            = 'Scope 3 savings Tonnes pa'
      dft.loc[[17],['param_value']]           = round(-remaining_co2_scope_3,4)    
      dft.loc[[17],['title']]                 = 'Intervention scope 3 savings - tonnes pa'
      dft.loc[[17],['scenario_number']]       = 0
      
      dfwfall                                 = dfwfall.append(dft)
     
      print('======dfwfall columns')
#      print(dfwfall.columns)
      
#      print("At end of create_pbi_tables")
#      print('====dfestate')
#      print(' ')
#      print(dfestate.to_string())
#      print(' ')
#      print('====dfwfall')
#      print(' ')
#      print(dfwfall.to_string())
#      print('**************** Number of rows in dfwfall df')
#      print(dfwfall.shape[0])
    #  Delete all records for this entity_number (for this scenario) in pbi_waterfall_project_savings table then insert the waterfall dataframe

      sqlde  = f"DELETE FROM pbi_waterfall_project_savings WHERE entity_number = {entity_number} ;"
      cur.execute(sqlde)
      conn.commit()

      print('Creating engine and doing to_sql')
      print(Connections.connection_string)
      connect_str = 'mssql+pyodbc:///?odbc_connect=' + urllib.parse.quote_plus(odbc_str)
      print('======connect_str')
      print(connect_str)


      engine      = salch.create_engine(connect_str)
      with engine.connect().execution_options(autocommit=False) as conn2:
        txn = conn2.begin()
        dfwfall.to_sql('pbi_waterfall_project_savings', con=conn2, if_exists='append', index= False)
#        print(dir(conn2))
        txn.commit()

    return ret_mess      
  except Exception as e: 
    summary =   summary + f"******An exception has occurred in 'create_pbi_tables_v3'- please see upload log for details"

    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    print(msg)
    ret_mess['ef']      = 2
    ret_mess['em']      = msg
    return ret_mess
 

@anvil.server.callable
def write_upload_log_2_db(conn, entity_number, user_email, dt_str, task_id, task_name, up_log):
  # This function writes the upload log to the task_logs table in the database;
  try:
    with conn.cursor() as cursor:
      dup_log = up_log.replace('\'', '"') #Single quotes in the text of the up_log will cause syntax error in SQL so replace with double quotes

      # Extract the real task_id and context (PCE) from the input task_id
      
      parts_list = task_id.split('/')
      task_id    = parts_list[0]
      partner_id = parts_list[1]
      client_id  = parts_list[2]
      entity_id  = parts_list[3]
      
      sql_in          = f"INSERT INTO task_logs (entity_number, user_email, dtstr, task_id, partner_id, client_id, entity_id, task_name, up_log) VALUES ({entity_number},\'{user_email}\',\'{dt_str}\', \'{task_id}\', \'{partner_id}\', \'{client_id}\', \'{entity_id}\',\'{task_name}\', \'{dup_log}\')"
      sql_up          = f"UPDATE task_logs SET entity_number = {entity_number}, user_email = \'{user_email}\', dtstr = \'{dt_str}\', task_id = \'{task_id}\', partner_id = \'{partner_id}\', client_id = \'{client_id}\', entity_id = \'{entity_id}\',task_name = \'{task_name}\', up_log = \'{dup_log}\' WHERE task_id = \'{task_id}\' "

      sql_pre         = f"IF EXISTS (SELECT task_id FROM task_logs WHERE task_id =  \'{task_id}\' ) BEGIN {sql_up} END ELSE BEGIN {sql_in} END;"
      cursor.execute(sql_pre)
      conn.commit()
     
  except Exception as e: 
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    print(f"Exception in write_upload_log_2_db \n {msg}")
    return


@anvil.server.callable   
def log_str_types_in_numeric_cols(df_in, column_list):
    # Used to return a formatted log of the location of occurences of values of type str appearing in the numeric columns of a dataframe. 
    # df_in        - input dataframe
    # column_list  - list of the column names in df_in to be searched for occurences of values of type str
    try:
      out_log     = ''
      ef          = 0
      em          = ''
      nerrs       = 0
      ret         = {'out_log' : out_log, 'ef' : 0, 'em' : '', 'nerrs' : 0}
      ncols       = len(column_list)
      shape       = df_in.shape
      nrows       = shape[0]
      
      # Validate column names in input list of columns
      
      colerr      = False
      for col in column_list:
        if not col in df_in.columns:
          colerr  = True
          em  = em + f"Column named {col} does not exist in input dataframe \n"
      if colerr:    
        ret['out_log'] = ''
        ret['ef']      = 2 
        ret['em']      = em
        return ret

      # Search each of the named columns for values of type str and log them
      out_log      = out_log + f"The following rows in numeric columns contain strings instead of numbers: \n"
      for col in column_list:
        for index, row in df_in.iterrows():
          excel_row = index + 2  #Indexing starts from 0 plus have excel header

          if type(row[col]) is str:
            nerrs   = nerrs + 1
            out_log = out_log + f"Value in column {col}, row {excel_row}  \n"
      
      ret['out_log']    = out_log
      ret['ef']         = 0 
      ret['em']         = ''
      ret['nerrs']      = nerrs
      return ret
      
    except Exception as e:
      msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
      print(f"Exception 'log_str_types_in_numeric_cols' \n {msg}")
      ret['ef']   = 2
      ret['em']   = msg
      return ret   

@anvil.server.callable 
def extract_dt_str_from_upload_log(log):
  ret_mess  = {'ef':0, 'em':''}
  # Extracts the run date and time from a log in the format '%d/%m/%Y %H:%M:%S' and returns it in a sortable format '%Y/%m/%d %H:%M:%S'
  try:
    sub1 = "run on "
    sub2 = " for ent"
    
    # getting index of substrings
    idx1 = log.index(sub1)
    idx2 = log.index(sub2)
    
    res = ''
    # getting elements in between = the run data and time string
    for idx in range(idx1 + len(sub1) , idx2):
        res = res + log[idx]
    
    # Convert date time string to datetime object and convert that object to a new date time string format
    
    old_date_time = datetime.strptime(res, '%d/%m/%Y %H:%M:%S')
    new_str       = datetime.strftime(old_date_time, '%Y/%m/%d %H:%M:%S')
    return new_str

  except Exception as e:
    return 'Date re-format error'

@anvil.server.callable
def get_partner_client_from_entity_number(conn, entity_number):
# This function takes an entity_number as input and navigates up the organisational hierarchy to find the owning Client
# and the Partner who owns that Client. The Client and Partner are returned together with an error flag (ef) and error 
# message (em) indicating the result of the search: -
# Error flag (ef) possible values: -
# 0  - Client and Partner successfully found
# 2  - Exception failure - Partner and Client returned as 'Undefined'
  ret            = {'partner' : 'Undefined', 'client' : 'Undefined', 'ef' : 0, 'em' : ''}

  try:
    with conn.cursor() as cursor:
      q1 = f"SELECT e.entity_number, e.entity_id, e.entity_name,  c.client_number, c.client_id, c.client_name,  \
      p.partner_number, p.partner_id, p.partner_name \
      FROM entities AS e \
      INNER JOIN client AS c \
      ON e.client_number = c.client_number \
      INNER JOIN partner AS p \
      ON c.partner_number = p.partner_number \
      WHERE e.entity_number = {entity_number};"

      print(f" q1 = {q1}\n")
      cursor.execute(q1)
      t_cn = cursor.fetchall()
      keys = ("entity_number","entity_id","entity_name","client_number","client_id","client_name","partner_number","partner_id","partner_name")
      cn   = [dict(zip(keys, values)) for values in t_cn]
      lpc  = cn[0]
      
      ret['partner'] = lpc['partner_id']
      ret['client']  = lpc['client_id']
      return ret

  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef']   = 2
    ret['em']   = msg
    return ret 
    
@anvil.server.callable
def get_partner_client_entity_from_key( key):
# This function takes the encrypted PCE key as input and decodes it back to Partner, Client and Entity.
# The Partner_id, Client_id and Entity_id are returned together with with an error flag (ef) and error 
# message (em) indicating the result of the search: -
# Error flag (ef) possible values: -
# 0  - Client and Partner successfully found
# 2  - Exception failure - Partner and Client returned as 'Undefined'
  
  ret            = {'partner' : 'Undefined', 'client' : 'Undefined', 'entity' : 'undefined', 'ef' : 0, 'em' : ''}

  try:

      return ret

  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef']   = 2
    ret['em']   = msg
    return ret   

@anvil.server.callable
def authenticate_workbook(auts, keys, context_partner, context_client, context_entity):
  
# Performs 3 stages of authentication of the input workbook identified by file_name. auts is the DataFrame for the Auth sheet and Keys is the Dataframe for the keys sheet.
# context_partner, context_client and context_entity define the context within which this function has been called in the application
# 
# Stage    Authentication performed
# -----    ------------------------
#  1       Checks that sheets 'Auth' and 'Key' exist in the workbook and neither are empty. It reads the contents of both sheets: -
#          Auths - reads wb_partner at iloc(2,0); wb_client at iloc(4,0) and wb_entity at iloc(6,0)
#                  wb_partner, wb_client and wb_entity define the context used to write the workbook in clear text.
#          Keys -  reads key at iloc(1,0) which is a 140 byte encrypted string encoding the value of 'partner_id/client_id/entity_id'
#                  as these were in the context when the workbook was created.
#  2       Reads the partner, client and entity read from the 'Auths' sheet, decodes the key on the 'Keys' sheet and compares the resulting partner, client and entity
#          values with the corresponding values read from the 'Auth' sheet.
#          If the values do not match this means there has been some interference with the authentication in the workbook and the workbook will be rejected.
#  3       Decrypts the key and compares the decrypted values of partner, client and entity against the equivalent values in the current context.
#          If the 2 sets of values do not match this means the user is trying to load the workbook into a context that does not match the context intended
#          when the workbook was created. The workbook will be rejected.
#
# Authentication errors will result in the value of ret['ef'] = 1 and ret['em'] will hold a message describing the authentication failure for display to the user.
# Exceptions will result in the value of ret['ef'] = 2 and ret[.em.] will hold the the exception message for logging.
#
  try:
    ret                 = {'ef':0, 'em': ''}
    # Check Auths and Key sheet are not empty. Read their contents.
#    auts                = pd.read_excel(file_name, sheet_name = 'Auth', dtype = object)
    nr                  = auts.shape[0]
    if nr == 0:
      ret['ef']   = 1 
      ret['em']   = "****Authentication failure - input Auths sheet is empty"
      return
    else:
      print(f"At auts = {auts.to_string()}\n")
      s_partner   = auts.iloc[1,0]
      s_client    = auts.iloc[3,0]
      s_entity    = auts.iloc[5,0]
      
 #   keys                = pd.read_excel(file_name, sheet_name = 'Key', dtype = object)
    nr                  = keys.shape[0]
    if nr == 0:
      ret['ef']   = 1 
      ret['em']   = "****Authentication failure - input Key sheet is empty"
      return      
    else:
      print(f"At keys = {keys.to_string()}\n")
      s_key       = keys.iloc[0,0]
      kp          = decode_upload_auth_key(s_key) # NOTE: If the key read is invalid because it has been edited then this statement throws an exception
      k_partner   = kp[0]
      k_client    = kp[1]
      k_entity    = kp[2]
      
      if (s_entity != k_entity) or (s_client != k_client) or (s_partner != k_partner):
        ret['ef']   = 1 
        ret['em']   = "****Authentication failure - inconsistency between key and partner, client and entity values in workbook Auths sheet\n Likely the workbook has been tampered with."
        return ret             

      if (context_entity != k_entity) or (context_client != k_client) or (context_partner != k_partner):
        ret['ef']   = 1 
        ret['em']   = "****Authentication failure - inconsistency between current context and the key stored in the workbook\n Possibly attempting to load the workbook to the wrong entity"
        return ret      
    return ret
  except Exception as e:
    msg = "".join(traceback.format_exception(type(e), e, e.__traceback__))
    ret['ef']   = 2
    ret['em']   = msg
    return ret       